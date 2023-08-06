#!/usr/bin/env python

"""
camcops_server/cc_modules/cc_tsv.py

===============================================================================

    Copyright (C) 2012-2020 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of CamCOPS.

    CamCOPS is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    CamCOPS is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with CamCOPS. If not, see <http://www.gnu.org/licenses/>.

===============================================================================

**Helper functions/classes for spreadsheet-style tab-separated value (TSV)
exports.**

"""

from collections import OrderedDict
import csv
import datetime
import io
import logging
import os
import random
import re
from typing import (Any, BinaryIO, Callable, Dict, Iterable, List, Optional,
                    Sequence, Union)
from unittest import TestCase
import zipfile

from cardinal_pythonlib.datetimefunc import (
    format_datetime,
    get_now_localtz_pendulum,
)
from cardinal_pythonlib.excel import convert_for_openpyxl
from cardinal_pythonlib.logs import BraceStyleAdapter
from numpy import float64
from pendulum.datetime import DateTime
from semantic_version import Version
from sqlalchemy.engine.result import ResultProxy

from camcops_server.cc_modules.cc_constants import DateFormat

ODS_VIA_PYEXCEL = True  # significantly faster
XLSX_VIA_PYEXCEL = True

if ODS_VIA_PYEXCEL:
    import pyexcel_ods3  # e.g. pip install pyexcel-ods3==0.5.3
    ODSWriter = ODSSheet = None
else:
    from odswriter import ODSWriter, Sheet as ODSSheet  # noqa
    pyexcel_ods3 = None

if XLSX_VIA_PYEXCEL:
    import pyexcel_xlsx  # e.g. pip install pyexcel-xlsx==0.5.7
    openpyxl = XLWorkbook = XLWorksheet = None
else:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook as XLWorkbook
    from openpyxl.worksheet.worksheet import Worksheet as XLWorksheet
    pyexcel_xlsx = None

log = BraceStyleAdapter(logging.getLogger(__name__))


# =============================================================================
# Helper function (may move to cardinal_pythonlib at some point
# =============================================================================

def convert_for_pyexcel_ods3(x: Any) -> Any:
    """
    Converts known "unusual" data types to formats suitable for
    ``pyexcel-ods3``. Specifically, handles:

    - :class:`pendulum.datetime.DateTime`
    - :class:`datetime.datetime`
    - :class:`semantic_version.Version`
    - ``None``
    - :class:`numpy.float64`

    Args:
        x: a data value

    Returns:
        the same thing, or a more suitable value!
    """
    if isinstance(x, (DateTime, datetime.datetime)):
        return x.strftime(DateFormat.ISO8601)
    elif x is None:
        return ""
    elif isinstance(x, Version):
        return str(x)
    elif isinstance(x, float64):
        return float(x)
    else:
        return x


# =============================================================================
# TSV output holding structures
# =============================================================================

class TsvPage(object):
    """
    Represents a single TSV "spreadsheet".
    """
    def __init__(self, name: str,
                 rows: List[Union[Dict[str, Any], OrderedDict]]) -> None:
        """
        Args:
            name: name for the whole sheet
            rows: list of rows, where each row is a dictionary mapping
                column name to value
        """
        assert name, "Missing name"
        self.name = name
        self.rows = rows
        self.headings = []  # type: List[str]
        for row in rows:
            self._add_headings_if_absent(row.keys())

    def __str__(self) -> str:
        return f"TsvPage: name={self.name}\n{self.get_tsv()}"

    @classmethod
    def from_headings_rows(cls, name: str, headings: List[str],
                           rows: List[Sequence[Any]]) -> "TsvPage":
        """
        Creates a TsvPage object using a list of headings and the row data
        as a list of lists.
        """
        page = cls(name=name, rows=[])
        n_cols = len(headings)
        page.headings = headings
        for row in rows:
            assert len(row) == n_cols
            page.rows.append(dict(zip(headings, row)))
        return page

    @classmethod
    def from_resultproxy(cls, name: str, rp: ResultProxy) -> "TsvPage":
        """
        Creates a TsvPage object from an SQLAlchemy ResultProxy.

        Args:
            rp:
                A :class:` sqlalchemy.engine.result.ResultProxy`.
            name:
                Name for this sheet.
        """
        column_names = rp.keys()
        rows = rp.fetchall()
        return cls.from_headings_rows(
            name=name, headings=column_names, rows=rows)

    @property
    def empty(self) -> bool:
        """
        Do we have zero rows?
        """
        return len(self.rows) == 0

    def _add_headings_if_absent(self, headings: Iterable[str]) -> None:
        """
        Add any headings we've not yet seen to our list of headings.
        """
        for h in headings:
            if h not in self.headings:
                self.headings.append(h)

    def add_or_set_value(self, heading: str, value: Any) -> None:
        """
        If we contain only a single row, this function will set the value
        for a given column (``heading``) to ``value``.

        Raises:
            :exc:`AssertionError` if we don't have exactly 1 row
        """
        assert len(self.rows) == 1, "add_value can only be used if #rows == 1"
        self._add_headings_if_absent([heading])
        self.rows[0][heading] = value

    def add_or_set_column(self, heading: str, values: List[Any]) -> None:
        """
        Set the column labelled ``heading`` so it contains the values specified
        in ``values``. The length of ``values`` must equal the number of rows
        that we already contain.

        Raises:
            :exc:`AssertionError` if the number of values doesn't match
            the number of existing rows
        """
        assert len(values) == len(self.rows), "#values != #existing rows"
        self._add_headings_if_absent([heading])
        for i, row in enumerate(self.rows):
            row[heading] = values[i]

    def add_or_set_columns_from_page(self, other: "TsvPage") -> None:
        """
        This function presupposes that ``self`` and ``other`` are two pages
        ("spreadsheets") with *matching* rows.

        It updates values or creates columns in ``self`` such that the values
        from all columns in ``other`` are written to the corresponding rows of
        ``self``.

        Raises:
            :exc:`AssertionError` if the two pages (sheets) don't have
            the same number of rows.
        """
        assert len(self.rows) == len(other.rows), "Mismatched #rows"
        self._add_headings_if_absent(other.headings)
        for i, row in enumerate(self.rows):
            for k, v in other.rows[i].items():
                row[k] = v

    def add_rows_from_page(self, other: "TsvPage") -> None:
        """
        Add all rows from ``other`` to ``self``.
        """
        self._add_headings_if_absent(other.headings)
        self.rows.extend(other.rows)

    def sort_headings(self) -> None:
        """
        Sort our headings internally.
        """
        self.headings.sort()

    @property
    def plainrows(self) -> List[List[Any]]:
        """
        Returns a list of rows, where each row is a list of values.
        Does not include a "header" row.

        Compare :attr:`rows`, which is a list of dictionaries.
        """
        rows = []
        for row in self.rows:
            rows.append([row.get(h) for h in self.headings])
        return rows

    def spreadsheetrows(self, converter: Callable[[Any], Any]) \
            -> List[List[Any]]:
        """
        Like :meth:`plainrows`, but (a) ensures every cell is converted to a
        value that can be sent to a spreadsheet converted (e.g. ODS, XLSX), and
        (b) includes a header row.
        """
        rows = [self.headings.copy()]
        for row in self.rows:
            rows.append([converter(row.get(h))
                         for h in self.headings])
        return rows

    def get_tsv(self, dialect: str = "excel-tab") -> str:
        r"""
        Returns the entire page (sheet) as TSV: one header row and then
        lots of data rows.

        For the dialect, see
        https://docs.python.org/3/library/csv.html#csv.excel_tab.

        For CSV files, see RGC 4180: https://tools.ietf.org/html/rfc4180.

        For TSV files, see
        https://www.iana.org/assignments/media-types/text/tab-separated-values.

        Test code:

        .. code-block:: python

            import io
            import csv
            from typing import List

            def test(row: List[str], dialect: str = "excel-tab") -> str:
                f = io.StringIO()
                writer = csv.writer(f, dialect=dialect)
                writer.writerow(row)
                return f.getvalue()

            test(["hello", "world"])
            test(["hello\ttab", "world"])  # actual tab within double quotes
            test(["hello\nnewline", "world"])  # actual newline within double quotes
            test(['hello"doublequote', "world"])  # doubled double quote within double quotes

        """  # noqa
        f = io.StringIO()
        writer = csv.writer(f, dialect=dialect)
        writer.writerow(self.headings)
        for row in self.rows:
            writer.writerow([row.get(h) for h in self.headings])
        return f.getvalue()

    def write_to_openpyxl_xlsx_worksheet(self, ws: "XLWorksheet") -> None:
        """
        Writes data from this page to an existing ``openpyxl`` XLSX worksheet.
        """
        ws.append(self.headings)
        for row in self.rows:
            ws.append([convert_for_openpyxl(row.get(h))
                       for h in self.headings])

    def write_to_odswriter_ods_worksheet(self, ws: "ODSSheet") -> None:
        """
        Writes data from this page to an existing ``odswriter`` ODS sheet.
        """
        # noinspection PyUnresolvedReferences
        ws.writerow(self.headings)
        for row in self.rows:
            # noinspection PyUnresolvedReferences
            ws.writerow([row.get(h) for h in self.headings])

    def r_object_name(self) -> str:
        """
        Name of the object when imported into R.
        The main thing: no leading underscores.
        """
        n = self.name
        n = n[1:] if n.startswith("_") else n
        return f"camcops_{n}"  # less chance of conflict within R

    def r_data_table_definition(self) -> str:
        """
        Returns a string to define this object as a ``data.table`` in R.

        See also:

        - https://stackoverflow.com/questions/32103639/read-csv-file-in-r-with-double-quotes
        """  # noqa
        object_name = self.r_object_name()
        csv_text = self.get_tsv(dialect="excel")
        csv_text = csv_text.replace('"', r'\"')
        definition = (
            f'data.table::fread(sep=",", header=TRUE, text="{csv_text}"\n)'
        )
        return f"{object_name} <- {definition}"


class TsvCollection(object):
    """
    A collection of :class:`camcops_server.cc_modules.cc_tsv.TsvPage` pages
    (spreadsheets), like an Excel workbook.
    """
    def __init__(self) -> None:
        self.pages = []  # type: List[TsvPage]

    def __str__(self) -> str:
        return (
            "TsvCollection:\n" +
            "\n\n".join(page.get_tsv() for page in self.pages)
        )

    # -------------------------------------------------------------------------
    # Pages
    # -------------------------------------------------------------------------

    def page_with_name(self, page_name: str) -> Optional[TsvPage]:
        """
        Returns the page with the specific name, or ``None`` if no such
        page exists.
        """
        return next((page for page in self.pages if page.name == page_name),
                    None)

    def add_page(self, page: TsvPage) -> None:
        """
        Adds a new page to our collection. If the new page has the same name
        as an existing page, rows from the new page are added to the existing
        page. Does nothing if the new page is empty.
        """
        if page.empty:
            return
        existing_page = self.page_with_name(page.name)
        if existing_page:
            # Blend with existing page
            existing_page.add_rows_from_page(page)
        else:
            # New page
            self.pages.append(page)

    def add_pages(self, pages: List[TsvPage]) -> None:
        """
        Adds all ``pages`` to our collection, via :func:`add_page`.
        """
        for page in pages:
            self.add_page(page)

    def sort_headings_within_all_pages(self) -> None:
        """
        Sort headings within each of our pages.
        """
        for page in self.pages:
            page.sort_headings()

    def sort_pages(self) -> None:
        """
        Sort our pages by their page name.
        """
        self.pages.sort(key=lambda p: p.name)

    def get_page_names(self) -> List[str]:
        """
        Return a list of the names of all our pages.
        """
        return [p.name for p in self.pages]

    # -------------------------------------------------------------------------
    # TSV
    # -------------------------------------------------------------------------

    def get_tsv_file(self, page_name: str) -> str:
        """
        Returns a TSV file for a named page.

        Raises:
            :exc:`AssertionError` if the named page does not exist

        """
        page = self.page_with_name(page_name)
        assert page is not None, f"No such page with name {page_name}"
        return page.get_tsv()

    # -------------------------------------------------------------------------
    # ZIP of TSVs
    # -------------------------------------------------------------------------

    def write_zip(self,
                  file: Union[str, BinaryIO],
                  encoding: str = "utf-8",
                  compression: int = zipfile.ZIP_DEFLATED) -> None:
        """
        Writes data to a file, as a ZIP file of TSV files.

        Args:
            file: filename or file-like object
            encoding: encoding to use when writing the TSV files
            compression: compression method to use

        Choice of compression method: see

        - https://docs.python.org/3/library/zipfile.html
        - https://pkware.cachefly.net/webdocs/casestudies/APPNOTE.TXT
        - http://en.wikipedia.org/wiki/Zip_(file_format)#Compression_methods

        Note also that ``openpyxl`` uses ``ZIP_DEFLATED``, which seems to be
        the most portable if not the best compression.
        """
        if isinstance(file, str):  # it's a filename
            with open(file, "wb") as binaryfile:
                return self.write_zip(binaryfile, encoding)  # recurse once
        with zipfile.ZipFile(file, mode="w", compression=compression) as z:
            # Write to ZIP.
            # If there are no valid task instances, there'll be no TSV;
            # that's OK.
            for filename_stem in self.get_page_names():
                tsv_filename = filename_stem + ".tsv"
                tsv_contents = self.get_tsv_file(page_name=filename_stem)
                z.writestr(tsv_filename, tsv_contents.encode(encoding))

    def as_zip(self, encoding: str = "utf-8") -> bytes:
        """
        Returns the TSV collection as a ZIP file containing TSV files.

        Args:
            encoding: encoding to use when writing the TSV files
        """
        with io.BytesIO() as memfile:
            self.write_zip(memfile, encoding)
            zip_contents = memfile.getvalue()
        return zip_contents

    # -------------------------------------------------------------------------
    # XLSX, ODS
    # -------------------------------------------------------------------------

    def write_xlsx(self, file: Union[str, BinaryIO]) -> None:
        """
        Write the contents in XLSX (Excel) format to a file.

        Args:
            file: filename or file-like object
        """
        if XLSX_VIA_PYEXCEL:  # use pyexcel_xlsx
            data = self._get_pyexcel_data(convert_for_openpyxl)
            pyexcel_xlsx.save_data(file, data)
        else:  # use openpyxl
            # Marginal performance gain with write_only. Does not automatically
            # add a blank sheet
            wb = XLWorkbook(write_only=True)
            valid_name_dict = self.get_pages_with_valid_sheet_names()
            for page, title in valid_name_dict.items():
                ws = wb.create_sheet(title=title)
                page.write_to_openpyxl_xlsx_worksheet(ws)
            wb.save(file)

    def as_xlsx(self) -> bytes:
        """
        Returns the TSV collection as an XLSX (Excel) file.
        """
        with io.BytesIO() as memfile:
            self.write_xlsx(memfile)
            contents = memfile.getvalue()
        return contents

    @staticmethod
    def get_sheet_title(page: TsvPage) -> str:
        r"""
        Returns a worksheet name for a :class:`TsvPage`.

        See ``openpyxl/workbook/child.py``.

        - Excel prohibits ``\``, ``*``, ``?``, ``:``, ``/``, ``[``, ``]``
        - LibreOffice also prohibits ``'`` as first or last character but let's
          just replace that globally.
        """
        title = re.sub(r"[\\*?:/\[\]']", "_", page.name)

        if len(title) > 31:
            title = f"{title[:28]}..."

        return title

    def _get_pyexcel_data(self, converter: Callable[[Any], Any]) \
            -> Dict[str, List[List[Any]]]:
        """
        Returns data in the format expected by ``pyexcel``, which is an ordered
        dictionary mapping sheet names to a list of rows, where each row is a
        list of cell values.
        """
        data = OrderedDict()
        for page in self.pages:
            data[self.get_sheet_title(page)] = page.spreadsheetrows(converter)
        return data

    def write_ods(self, file: Union[str, BinaryIO]) -> None:
        """
        Writes an ODS (OpenOffice spreadsheet document) to a file.

        Args:
            file: filename or file-like object
        """
        if ODS_VIA_PYEXCEL:  # use pyexcel_ods3
            data = self._get_pyexcel_data(convert_for_pyexcel_ods3)
            pyexcel_ods3.save_data(file, data)
        else:  # use odswriter
            if isinstance(file, str):  # it's a filename
                with open(file, "wb") as binaryfile:
                    return self.write_ods(binaryfile)  # recurse once
            # noinspection PyCallingNonCallable
            with ODSWriter(file) as odsfile:
                valid_name_dict = self.get_pages_with_valid_sheet_names()
                for page, title in valid_name_dict.items():
                    sheet = odsfile.new_sheet(name=title)
                    page.write_to_odswriter_ods_worksheet(sheet)

    def as_ods(self) -> bytes:
        """
        Returns the TSV collection as an ODS (OpenOffice spreadsheet document)
        file.
        """
        with io.BytesIO() as memfile:
            self.write_ods(memfile)
            contents = memfile.getvalue()
        return contents

    def get_pages_with_valid_sheet_names(self) -> Dict[TsvPage, str]:
        """
        Returns an ordered mapping from :class:`TsvPage` objects to their
        sheet names.
        """
        name_dict = OrderedDict()

        for page in self.pages:
            name_dict[page] = self.get_sheet_title(page)

        self.make_sheet_names_unique(name_dict)

        return name_dict

    @staticmethod
    def make_sheet_names_unique(name_dict: Dict[TsvPage, str]) -> None:
        """
        Modifies (in place) a mapping from :class:`TsvPage` to worksheet names,
        such that all page names are unique.

        - See also :func:`avoid_duplicate_name` in
          ``openpxl/workbook/child.py``
        - We keep the 31 character restriction
        """
        unique_names = []  # type: List[str]

        for page, name in name_dict.items():
            attempt = 0

            while name.lower() in unique_names:
                attempt += 1

                if attempt > 1000:
                    # algorithm failure, better to let Excel deal with the
                    # consequences than get stuck in a loop
                    log.debug(
                        f"Failed to generate a unique sheet name from {name}"
                    )
                    break

                match = re.search(r'\d+$', name)
                count = 0
                if match is not None:
                    count = int(match.group())

                new_suffix = str(count + 1)
                name = name[:-len(new_suffix)] + new_suffix
            name_dict[page] = name
            unique_names.append(name.lower())

    # -------------------------------------------------------------------------
    # R
    # -------------------------------------------------------------------------

    def as_r(self) -> str:
        """
        Returns data as an R script.

        This could be more sophisticated, e.g. creating factors with
        appropriate levels (etc.).
        """
        now = format_datetime(get_now_localtz_pendulum(),
                              DateFormat.ISO8601_HUMANIZED_TO_SECONDS_TZ)
        table_definition_str = "\n\n".join(
            page.r_data_table_definition()
            for page in self.pages
        )
        script = f"""#!/usr/bin/env Rscript

# R script generated by CamCOPS at {now}

# =============================================================================
# Libraries
# =============================================================================

library(data.table)

# =============================================================================
# Data
# =============================================================================

{table_definition_str}

"""
        return script

    def write_r(self, filename: str, encoding: str = "utf-8") -> None:
        """
        Write the contents in R format to a file.

        Args:
            filename: filename or file-like object
            encoding: encoding to use
        """
        with open(filename, "wt", encoding=encoding) as f:
            f.write(self.as_r())


# =============================================================================
# Unit tests
# =============================================================================

class TsvCollectionTests(TestCase):
    def test_xlsx_created_from_zero_rows(self) -> None:
        page = TsvPage(name="test", rows=[])
        coll = TsvCollection()
        coll.add_page(page)

        output = coll.as_xlsx()

        # https://en.wikipedia.org/wiki/List_of_file_signatures
        self.assertEqual(output[0], 0x50)
        self.assertEqual(output[1], 0x4B)
        self.assertEqual(output[2], 0x03)
        self.assertEqual(output[3], 0x04)

    def test_xlsx_worksheet_names_are_page_names(self) -> None:
        page1 = TsvPage(name="name 1",
                        rows=[{"test data 1": "row 1"}])
        page2 = TsvPage(name="name 2",
                        rows=[{"test data 2": "row 1"}])
        page3 = TsvPage(name="name 3",
                        rows=[{"test data 3": "row 1"}])
        coll = TsvCollection()

        coll.add_pages([page1, page2, page3])

        data = coll.as_xlsx()
        buffer = io.BytesIO(data)
        expected_sheetnames = ["name 1", "name 2", "name 3"]
        if openpyxl:
            wb = openpyxl.load_workbook(buffer)  # type: XLWorkbook
            self.assertEqual(wb.sheetnames, expected_sheetnames)
        else:
            wb = pyexcel_xlsx.get_data(buffer)  # type: Dict[str, Any]
            sheetnames = list(wb.keys())
            self.assertEqual(sheetnames, expected_sheetnames)

    def test_xlsx_page_name_exactly_31_chars_not_truncated(self) -> None:
        page = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "abcdefghijklmnopqrstuvwxyz78901"
        )

    def test_xlsx_page_name_over_31_chars_truncated(self) -> None:
        page = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901234",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "abcdefghijklmnopqrstuvwxyz78..."
        )

    def test_xlsx_invalid_chars_in_page_name_replaced(self) -> None:
        page = TsvPage(name="[a]b\\c:d/e*f?g'h",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "_a_b_c_d_e_f_g_h"
        )

    def test_ods_page_name_sanitised(self) -> None:
        # noinspection PyUnresolvedReferences
        from xml.dom.minidom import parseString
        page = TsvPage(name="What perinatal service have you accessed?",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()
        coll.add_pages([page])

        data = coll.as_ods()

        zf = zipfile.ZipFile(io.BytesIO(data), "r")
        content = zf.read('content.xml')
        doc = parseString(content)
        sheets = doc.getElementsByTagName('table:table')
        self.assertEqual(sheets[0].getAttribute("table:name"),
                         "What perinatal service have ...")

    def test_worksheet_names_are_not_duplicated(self) -> None:
        page1 = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901234",
                        rows=[{"test data 1": "row 1"}])
        page2 = TsvPage(name="ABCDEFGHIJKLMNOPQRSTUVWXYZ789012345",
                        rows=[{"test data 2": "row 1"}])
        page3 = TsvPage(name="abcdefghijklmnopqrstuvwxyz7890123456",
                        rows=[{"test data 3": "row 1"}])
        coll = TsvCollection()

        coll.add_pages([page1, page2, page3])

        valid_sheet_names = coll.get_pages_with_valid_sheet_names()

        names = [v for k, v in valid_sheet_names.items()]

        self.assertIn("abcdefghijklmnopqrstuvwxyz78...", names)
        self.assertIn("ABCDEFGHIJKLMNOPQRSTUVWXYZ78..1", names)
        self.assertIn("abcdefghijklmnopqrstuvwxyz78..2", names)


def _make_benchmarking_collection(nsheets: int = 100,
                                  nrows: int = 200,
                                  ncols: int = 30,
                                  mindata: int = 0,
                                  maxdata: int = 1000000) -> TsvCollection:
    log.info(f"Creating TsvCollection with nsheets={nsheets}, nrows={nrows}, "
             f"ncols={ncols}...")
    coll = TsvCollection()
    for sheetnum in range(1, nsheets + 1):
        rows = [
            {
                f"c{colnum}": str(random.randint(mindata, maxdata))
                for colnum in range(1, ncols + 1)
            } for _ in range(1, nrows + 1)
        ]
        page = TsvPage(name=f"sheet{sheetnum}", rows=rows)
        coll.add_page(page)
    log.info("... done.")
    return coll


def file_size(filename: str) -> int:
    """
    Returns a file's size in bytes.
    """
    return os.stat(filename).st_size


def benchmark_save(xlsx_filename: str = "test.xlsx",
                   ods_filename: str = "test.ods",
                   tsv_zip_filename: str = "test.zip",
                   r_filename: str = "test.R") -> None:
    """
    Use with:

    .. code-block:: python

        from cardinal_pythonlib.logs import main_only_quicksetup_rootlogger
        from camcops_server.cc_modules.cc_tsv import benchmark_save
        main_only_quicksetup_rootlogger()
        benchmark_save()

    Args:
        xlsx_filename: XLSX file to create
        ods_filename: ODS file to create
        tsv_zip_filename: TSV ZIP file to create
        r_filename: R script to create

    Problem in Nov 2019 is that ODS is extremely slow. Rough timings:

    - TSV ZIP: about 4.1 Mb, about 0.2 s. Good.
    - XLSX (via openpyxl): about 4.6 Mb, 16 seconds.
    - XLSX (via pyexcel_xlsx): about 4.6 Mb, 16 seconds.
    - ODS (via odswriter): about 53 Mb, 56 seconds.
    - ODS (via pyexcel_ods3): about 2.8 Mb, 29 seconds.
    """
    coll = _make_benchmarking_collection()

    log.info("Writing TSV ZIP...")
    coll.write_zip(tsv_zip_filename)
    log.info(f"... done. File size {file_size(tsv_zip_filename)}")

    log.info("Writing XLSX...")
    coll.write_xlsx(xlsx_filename)
    log.info(f"... done. File size {file_size(xlsx_filename)}")

    log.info("Writing ODS...")
    coll.write_ods(ods_filename)
    log.info(f"... done. File size {file_size(ods_filename)}")

    log.info("Writing R...")
    coll.write_r(r_filename)
    log.info(f"... done. File size {file_size(r_filename)}")
