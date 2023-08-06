r"""
.. module:: tools.py
    :synopsis: Module pour gérer les compressions, diagnostics, calcul des tailles occupées, sécuriser les données etc.

.. moduleauthor:: Michel TRUONG <michel.truong@gmail.com>

.. topic:: Description générale

    Module pour gérer les compressions, diagnostics, calcul des tailles occupées, sécuriser les données etc.

"""
import csv
import datetime
import fnmatch
import hashlib
import json
import os
import re
import secrets
import shutil
import subprocess
import sys
import tempfile
import time
import zlib
from base64 import urlsafe_b64encode as b64e, urlsafe_b64decode as b64d
from collections import namedtuple
from pathlib import Path
from subprocess import Popen, PIPE

import PySimpleGUI as sg
import numpy as np
import pandas as pd
import yaml
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from openpyxl import load_workbook
from pandasgui import show

from corelibs import _corelibs as _c, config, lazy as lz, log as c_log, cleanse as cls, data

try:
    import enlighten
except ImportError:
    raise Exception(_c._print_import_exception("enlighten"))

BACKEND = default_backend()
ITERATIONS = 100_000
CIPHER_FILE_EXTENSION = ".clk"

CORELIBS_FOLDER = ".corelibs"

CORELIBS_TEMPLATE_DIR = "templates"
CORELIBS_CACHE_DIR = "cache"

corelibs_path = lz._get_corelibs_path()
corelibs_templates_root = lz.get_abspath(lz.get_abspath(corelibs_path, ".corelibs"), CORELIBS_TEMPLATE_DIR)
user_corelibs_root = lz.get_abspath(lz.get_home(), CORELIBS_FOLDER)
user_corelibs_cache_root = lz.get_abspath(user_corelibs_root, CORELIBS_CACHE_DIR)
user_corelibs_templates_root = lz.get_abspath(user_corelibs_root, CORELIBS_TEMPLATE_DIR)

SCAN_DIR_TEMPLATE_FILE = "scan_dir.xlsx"

log = _c.log

try:
    import coloredlogs as cl
except ImportError:
    raise Exception(_c._print_import_exception("coloredlogs"))


def _sliceit(iterable, tup):
    return iterable[tup[0]:tup[1]].strip()


def _win_stat(file_path):
    _owner = "unknown"
    WinStat = namedtuple("WinStat", ["owner", "group"])

    try:
        cmd = ["cmd", "/c", "dir", file_path, "/q"]
        session = Popen(cmd, stdin=PIPE, stdout=PIPE, stderr=PIPE)
        result = session.communicate()[0].decode(config.DEFAULT_DOS_CMD_CP_ENCODING)  # <=> DOS latin-1
        cmd_line = result.splitlines()[5]

        stat_index = WinStat(owner=(35, 59), group=(0, 0))

        _owner = _sliceit(cmd_line, stat_index.owner).split("\\")[1]  # 1 because COMPUTER_NAME\Owner_Name
    except IndexError:
        pass
    except PermissionError:
        pass

    stat = WinStat(
        owner=_owner,
        group=""  # windows group... GPRESULT /R... LOL =D
    )
    return stat


def _get_dir_total_size(root):
    size = 0

    try:
        for path, dirs, files in os.walk(root):
            for f in files:
                size += os.path.getsize(os.path.join(path, f))

    except FileNotFoundError as e:
        log.warning(f"-[{config.PACKAGE_NAME}]- Fichier ou répertoire inexistant • {e}")
        size += 0
        pass
    except PermissionError as e:
        log.critical(f"-[{config.PACKAGE_NAME}]- Habilitations insuffisantes ou chemin incorrect • {e}")
        size += 0
        pass
    except OSError as e:
        log.error(f"-[{config.PACKAGE_NAME}]- • {e}")
        size += 0
        pass

    return size


def get_file_properties(file_path, pretty_byte_size=True, ignore_errors=config.DEFAULT_IGNORE_ERROR):
    r"""
    .. admonition:: Description

        | Récupère les informations d'un fichier ou d'un répertoire

    :param file_path: chemin absolu du fichier/répertoire

    :param pretty_byte_size: afficher la taille du fichier/répertoire de manière lisible pour un humain

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `True`

    :param ignore_errors: forcer l'exécution lorsqu'une erreur est levée

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False` (cf. `DEFAULT_IGNORE_ERROR` dans :ref:`reference-label-config`)

    :return:
        :magenta:`tuple nommé` avec comme attributs :
            * :magenta:`st_mode` (mode de protection en binaire)
            * :magenta:`st_ino` (n° inode)
            * :magenta:`st_dev` (n° machine)
            * :magenta:`st_nlink` (nb de liens)
            * :magenta:`st_uid` (propriétaire)
            * :magenta:`st_gid` (groupe id sous Windows ou groupe sous Unix)
            * :magenta:`st_size` (taille du fichier/répertoire)
            * :magenta:`st_atime` (date dernier accès)
            * :magenta:`st_mtime` (date dernière modification)
            * :magenta:`st_ctime` (date de création sous Windows et date dernier modif/accès sous Unix)

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\get_file_properties.py
        :language: python

    """
    FileProperties = namedtuple(
        "FileProperties", [
            "st_mode",  # protection bits
            "st_ino",  # inode number
            "st_dev",  # device
            "st_nlink",  # number of hard links
            "st_uid",  # user id of owner
            "st_gid",  # group id of owner
            "st_size",  # size of file, in bytes
            "st_atime",  # time of most recent access
            "st_mtime",  # time of most recent content modification
            "st_ctime"  # platform dependent; time of most recent metadata change on Unix, or the time of creation on
            # Windows
        ]
    )

    if not lz.is_file_exists(file_path, ignore_errors=ignore_errors):
        if not ignore_errors:
            log.error(f"-[{config.PACKAGE_NAME}]- Le fichier \"{file_path}\" n'existe pas")
            sys.exit(1)

    err = False
    try:
        file_properties = Path(file_path).stat()
    except FileNotFoundError as e:
        log.warning(f"-[{config.PACKAGE_NAME}]- Fichier ou répertoire inexistant • {e}")
        err = True
    except PermissionError as e:
        log.critical(f"-[{config.PACKAGE_NAME}]- Habilitations insuffisantes ou chemin incorrect • {e}")
        err = True
    except OSError as e:
        log.error(f"-[{config.PACKAGE_NAME}]- • {e}")
        err = True

    if err:
        ByteSize = namedtuple("ByteSize", [
            "byte", "kilobyte", "megabyte", "gigabyte", "terabyte"
        ])

        return FileProperties(
            st_mode="PermissionError",
            st_ino="PermissionError",
            st_dev="PermissionError",
            st_nlink="PermissionError",
            st_uid="unknown",
            st_gid="unknown",
            st_size=ByteSize(
                byte=0,
                kilobyte=0,
                megabyte=0,
                gigabyte=0,
                terabyte=0
            ),
            st_atime=0,
            st_mtime=0,
            st_ctime=0
        )

    if lz.is_file_exists(file_path, is_dir=True, ignore_errors=ignore_errors):
        bytes_size = lz.get_bytes_size_formats(_get_dir_total_size(file_path))
    else:
        bytes_size = lz.get_bytes_size_formats(file_properties.st_size)

    unix_file_properties = Path(file_path)
    try:
        owner = unix_file_properties.owner()
        group = unix_file_properties.group()
    except NotImplementedError:
        win_stat = _win_stat(file_path)
        owner = win_stat.owner
        group = win_stat.group

    return FileProperties(
        st_mode=file_properties.st_mode,
        st_ino=file_properties.st_ino,
        st_dev=file_properties.st_dev,
        st_nlink=file_properties.st_nlink,
        st_uid=owner,
        st_gid=group if group else file_properties.st_gid,
        st_size=lz.get_bytes_size_4_human(bytes_size) if pretty_byte_size else bytes_size,
        st_atime=lz.epoch_2_datetime(file_properties.st_atime),
        st_mtime=lz.epoch_2_datetime(file_properties.st_mtime),
        st_ctime=lz.epoch_2_datetime(file_properties.st_ctime)
    )


def get_fingerprint(obj,
                    algorithm="sha256",
                    eval_as_string=False,
                    chunk=config.DEFAULT_BYTE_CHUNK_SIZE,
                    ignore_errors=config.DEFAULT_IGNORE_ERROR):
    r"""
    .. admonition:: Description

        | Calculer l'empreinte digitale (signature numérique) d'une chaîne de caractères ou d'un fichier passé en
            argument

    :param obj:
        | chaîne de caractères
        | ou
        | chemin absolu du fichier

    :param algorithm: applique l'algorithme de hashage

        | :magenta:`valeurs possibles`: `'blake2b', 'blake2s', 'md5', 'sha1', 'sha224', 'sha256', 'sha384', 'sha512',
            'sha3_256', 'sha3_224', 'sha3_384' ou 'sha3_512'`
        | :magenta:`valeur par défaut`: `sha256`

    :param eval_as_string: permet de forcer l'évaluation comme étant une chaine de caractère (si l'objet passé en
        argument a un nom contenant des caractères \\ ou / et n'est pas en réalité un chemin de fichiers)

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False`

    :param chunk: indique le buffer de lecture.

        | :magenta:`valeur par défaut`: `DEFAULT_BYTE_CHUNK_SIZE` (cf. :ref:`reference-label-config`)

    :param ignore_errors: forcer l'exécution lorsqu'une erreur est levée

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False` (cf. `DEFAULT_IGNORE_ERROR` dans :ref:`reference-label-config`)

    :return:
        :magenta:`empreinte digital unique`

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\get_fingerprint.py
        :language: python

    """
    _hash = ""

    _algorithm = {
        "blake2b": hashlib.blake2b, "blake2s": hashlib.blake2s,
        "md5": hashlib.md5,
        "sha1": hashlib.sha1,
        "sha224": hashlib.sha224, "sha256": hashlib.sha256, "sha384": hashlib.sha384, "sha512": hashlib.sha512,
        "sha3_256": hashlib.sha3_256, "sha3_224": hashlib.sha3_224, "sha3_384": hashlib.sha3_384,
        "sha3_512": hashlib.sha3_512
    }
    available_algorithm = [key for key, val in _algorithm.items()]

    now = datetime.datetime.now()
    epoch = lz.datetime_2_epoch(now.strftime("%d/%m/%Y %H:%M:%S.%f"), time_format="%d/%m/%Y %H:%M:%S.%f")
    err_return = f"PERMISSION_DENIED_ERROR_{epoch}"
    if lz.is_file_exists(obj, is_dir=True, ignore_errors=True):
        if not ignore_errors:
            log.error(f"-[{config.PACKAGE_NAME}]- Le hashage ne peut pas se faire sur un répertoire \"{obj}\"")
            sys.exit(1)
        else:
            return err_return

    is_file = False
    if "/" in obj or "\\" in obj:
        if not lz.is_file_exists(obj, ignore_errors=ignore_errors):
            _msg = f"-[{config.PACKAGE_NAME}]- Le fichier \"{obj}\" n'existe pas"
            if eval_as_string:
                log.warning(_msg)
                log.warning(f"-[{config.PACKAGE_NAME}]- La chaine \"{obj}\" sera évaluée comme demandée")
            else:
                if not ignore_errors:
                    log.error(_msg)
                    sys.exit(1)
                else:
                    return err_return
        else:
            is_file = True

    if algorithm not in _algorithm:
        if not ignore_errors:
            log.error(f"-[{config.PACKAGE_NAME}]- L'algorithme de hashage \"{algorithm}\" passé en argument n'existe "
                      f"pas. Les algorithmes disponibles sont {available_algorithm}.")
            sys.exit(1)
        else:
            return err_return

    if is_file:
        try:
            _hash = _algorithm[algorithm]()
            with open(obj, "rb") as f:
                for byte_block in iter(lambda: f.read(chunk), b""):  # 1024Kb
                    _hash.update(byte_block)

        except FileNotFoundError as e:
            log.warning(f"-[{config.PACKAGE_NAME}]- Fichier ou répertoire inexistant • {e}")
        except PermissionError as e:
            log.critical(f"-[{config.PACKAGE_NAME}]- Habilitations insuffisantes ou chemin incorrect • {e}")
            time.sleep(0.7)
            return err_return
        except OSError as e:
            log.error(f"-[{config.PACKAGE_NAME}]- • {e}")
            time.sleep(0.7)
            return err_return
    else:
        _hash = _algorithm[algorithm](str(obj).encode("utf-8"))

    if _hash != "":
        return _hash.hexdigest()

    return _hash


def _get_total_byte_size(files_2_zip):
    bytes = 0
    for f in files_2_zip:
        fp = get_file_properties(f, pretty_byte_size=False)
        bytes += fp.st_size.byte

    return bytes


def _get_best_dict_option_size(total_byte_size):
    dict_options_index = (64, 1000000, 2000000, 3000000, 4000000, 6000000, 8000000, 12000000, 16000000, 24000000,
                          32000000, 48000000, 64000000, 96000000, 128000000, 192000000, 256000000, 384000000, 512000000,
                          768000000, 1024000000, 1536000000)
    dict_options = {
        64: "-md64k",
        1000000: "-md1m",
        2000000: "-md2m",
        3000000: "-md3m",
        4000000: "-md4m",
        6000000: "-md6m",
        8000000: "-md8m",
        12000000: "-md12m",
        16000000: "-md16m",
        24000000: "-md24m",
        32000000: "-md32m",
        48000000: "-md48m",
        64000000: "-md64m",
        96000000: "-md96m",
        128000000: "-md128m",
        192000000: "-md192m",
        256000000: "-md256m",
        384000000: "-md384m",
        512000000: "-md512m",
        768000000: "-md768m",
        1024000000: "-md1024m",
        1536000000: "-md1536m"
    }
    dic_ratio = 1 / 4

    return dict_options[lz.get_closest_value_in_list(int(total_byte_size * dic_ratio), dict_options_index)]


def _validate_zip_yaml_file(yaml_file, ignore_errors=config.DEFAULT_IGNORE_ERROR):
    schema_archive = lz._get_corelibs_abs_path("yaml") + r"\schema_archives.yaml"

    if not lz.is_validated_schema(yaml_file, schema_archive, is_dict_schema=False):
        if not ignore_errors:
            log.error(f"-[{config.PACKAGE_NAME}]- Fichier de configuration \"{yaml_file}\" non conforme au schéma")
            sys.exit(1)

    with open(yaml_file, "r", encoding="utf-8") as in_yaml_file:
        try:
            return yaml.safe_load(in_yaml_file)
        except yaml.YAMLError as e:
            if not ignore_errors:
                log.error(f"-[{config.PACKAGE_NAME}]- Problème avec le fichier de configuration \"{yaml_file}\" ({e})")
                sys.exit(1)


def _is_archived_file(archive_dir_path, archive_name, ignore_errors=config.DEFAULT_IGNORE_ERROR, action="zip"):
    if not lz.is_file_exists(archive_dir_path, is_dir=True):
        if not ignore_errors:
            if action == "zip":
                log.error(f"-[{config.PACKAGE_NAME}]- Le répertoire \"{archive_dir_path}\" pour enregistrer l'archive "
                          f"\"{archive_name}\" n'existe pas.")
            else:
                log.error(f"-[{config.PACKAGE_NAME}]- L'archive \"{archive_dir_path}\\{archive_name}\" n'existe pas à "
                          f"l'emplacement indiqué.")
            sys.exit(1)


def _get_7z_exe():
    if lz.is_platform("Windows"):
        return '"' + lz._get_corelibs_abs_path("bin") + r"\x64\7za.exe" + '"'

    return '"' + lz._get_corelibs_abs_path("bin") + r"\linux\7z" + '"'


def _zip(yaml_file, method="7z", delete_sources_files=True, ignore_errors=config.DEFAULT_IGNORE_ERROR):
    if method not in ["7z"]:
        if not ignore_errors:
            log.error(f"-[{config.PACKAGE_NAME}]- Méthode de compression \"{method}\" inconnue. Méthodes possibles "
                      f"\"7z\"")
            sys.exit(1)

    archive_yaml = _validate_zip_yaml_file(yaml_file, ignore_errors)

    archive_dir_path = archive_yaml["archive"]["root"]
    archive_name = archive_yaml["archive"]["name"]
    _is_archived_file(archive_dir_path, archive_name, ignore_errors)

    files_2_zip = []
    if isinstance(archive_yaml["files"], (list, tuple)):
        for file in archive_yaml["files"]:
            if "/" in file or "\\" in file:
                _file = file
            else:
                _file = lz.get_abspath(archive_dir_path, file)

            if not lz.is_file_exists(_file):
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le fichier \"{_file}\" n'existe pas.")
                    sys.exit(1)

            files_2_zip.append(_file)
    else:
        if "/" in archive_yaml["files"] or "\\" in archive_yaml["files"]:
            _file = archive_yaml["files"]
        else:
            _file = lz.get_abspath(archive_dir_path, archive_yaml["files"])

        files_2_zip.append(_file)

    total_bytes = _get_total_byte_size(files_2_zip)
    dict_option_size = _get_best_dict_option_size(total_bytes)

    exe = _get_7z_exe()

    target = lz.get_abspath(archive_dir_path, archive_name)

    source_file = ""
    for file in files_2_zip:
        source_file += '"' + file + '" '

    method_option = "-t7z -mmt2"
    command = exe + " a " + method_option + " \"" + target + "\" " + source_file + " -mx9 " + dict_option_size

    command += " -sdel" if delete_sources_files else ""

    try:
        sp = subprocess.Popen(
            command,
            bufsize=0,
            shell=True,
            stdout=sys.stdout,  # subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True
        )
        exit_code = sp.wait()

        return exit_code
    except subprocess.CalledProcessError as e:
        if not ignore_errors:
            log.critical(f"-[{config.PACKAGE_NAME}]- Appel subprocess problème... ({e})")
            sys.exit(1)


def _unzip(archive_name, archive_path, files_2_unzip, ignore_errors=config.DEFAULT_IGNORE_ERROR):
    global exit_code
    # print(archive_name, archive_path, files_2_unzip, ignore_errors)

    exe = _get_7z_exe()

    for file in files_2_unzip:
        command = exe + " x " + "\"" + lz.get_abspath(archive_path, archive_name) + "\""  # -aoa -r
        _base = lz.get_dir_n_basename(file)
        if _base.dir_path:
            command += " \"" + _base.base_name + "\" -o\"" + _base.dir_path + "\""
        else:
            command += " \"" + _base.base_name + "\" -o\"" + archive_path + "\""

        command += " -r -y"
        try:
            sp = subprocess.Popen(
                command,
                bufsize=0,
                shell=True,
                stdout=sys.stdout,  # subprocess.PIPE,
                stderr=subprocess.STDOUT,
                universal_newlines=True
            )
            exit_code = sp.wait()

            if exit_code > 0:
                return exit_code
        except subprocess.CalledProcessError as e:
            if not ignore_errors:
                log.critical(f"-[{config.PACKAGE_NAME}]- Appel subprocess problème... ({e})")
                sys.exit(1)

    return exit_code


class Archive:
    _compress_level = {
        "fastest": "-mx1",
        "fast": "-mx3",
        "normal": "",
        "maximum": "-mx7",
        "ultra": "-mx9"
    }

    _gzip_compress_level = {
        "fastest": "-mx1",
        "normal": "",
        "maximum": "-mx7",
        "ultra": "-mx9"
    }

    _ppmd_dict_size = {
        "1 Mo": "-md1m",
        "2 Mo": "-md2m",
        "3 Mo": "-md3m",
        "4 Mo": "-md4m",
        "6 Mo": "-md6m",
        "8 Mo": "-md8m",
        "12 Mo": "-md12m",
        "16 Mo": "-md16m",
        "24 Mo": "-md24m",
        "32 Mo": "-md32m",
        "48 Mo": "-md48m",
        "64 Mo": "-md64m",
        "96 Mo": "-md96m",
        "128 Mo": "-md128m",
        "192 Mo": "-md192m",
        "256 Mo": "-md256m",
        "384 Mo": "-md384m",
        "512 Mo": "-md512m",
        "768 Mo": "-md768m",
        "1024 Mo": "-md1024m"
    }

    _lzma_dict_size = _ppmd_dict_size.copy()
    _lzma_dict_size.update({
        "1536 Mo": "-md1536m"
    })

    _bzip2_dict_size = {
        "100 Ko": "-md100k",
        "200 Ko": "-md200k",
        "300 Ko": "-md300k",
        "400 Ko": "-md400k",
        "500 Ko": "-md500k",
        "600 Ko": "-md600k",
        "700 Ko": "-md700k",
        "800 Ko": "-md800k",
        "900 Ko": "-md900k"
    }

    _ppmd_word_size = {
        "8": "-mfb8",
        "12": "-mfb12",
        "16": "-mfb16",
        "24": "-mfb24",
        "32": "-mfb32"
    }

    _lzma_word_size = _zip_word_size = _gzip_word_size = _deflate64_word_size = _ppmd_word_size.copy()
    _lzma_word_size.update({
        "48": "-mfb48",
        "64": "-mfb64",
        "96": "-mfb96",
        "128": "-mfb128",
        "192": "-mfb192",
        "256": "-mfb256",
        "273": "-mfb273"
    })

    _gzip_word_size.update({
        "48": "-mfb48",
        "64": "-mfb64",
        "96": "-mfb96",
        "128": "-mfb128",
        "192": "-mfb192",
        "256": "-mfb256",
        "258": "-mfb258"
    })

    _zip_word_size.update({
        "48": "-mfb48",
        "64": "-mfb64",
        "96": "-mfb96",
        "128": "-mfb128",
        "192": "-mfb192",
        "256": "-mfb256",
        "258": "-mfb258"
    })

    _deflate64_word_size.update({
        "48": "-mfb48",
        "64": "-mfb64",
        "96": "-mfb96",
        "128": "-mfb128",
        "192": "-mfb192",
        "256": "-mfb256",
        "257": "-mfb257"
    })

    _bzip2_solid_block_size = _ppmd_solid_block_size = _lzma_solid_block_size = {
        "1 Mo": "-ms1m",
        "2 Mo": "-ms2m",
        "4 Mo": "-ms4m",
        "8 Mo": "-ms8m",
        "16 Mo": "-ms16m",
        "32 Mo": "-ms32m",
        "64 Mo": "-ms64m",
        "128 Mo": "-ms128m",
        "256 Mo": "-ms256m",
        "512 Mo": "-ms512m",
        "1 Go": "-ms1g",
        "2 Go": "-ms2g",
        "4 Go": "-ms4g",
        "8 Go": "-ms8g",
        "16 Go": "-ms16g",
        "32 Go": "-ms32g",
        "64 Go": "-ms64g"
    }

    _archive_command_options = {
        "7z": {
            "command": "",  # -t7z - default
            "compress_level": _compress_level,
            "algorithm": {
                "BZip2": {
                    "command": "-m0=BZip2",
                    "dict_size": _bzip2_dict_size,
                    "word_size": "",
                    "solid_block_size": _bzip2_solid_block_size
                },
                "LZMA": {
                    "command": "-m0=LZMA",
                    "dict_size": _lzma_dict_size,
                    "word_size": _lzma_word_size,
                    "solid_block_size": _lzma_solid_block_size
                },
                "LZMA2": {
                    "command": "",
                    "dict_size": _lzma_dict_size,
                    "word_size": _lzma_word_size,
                    "solid_block_size": _lzma_solid_block_size
                },
                "PPMd": {
                    "command": "-m0=PPMd",
                    "dict_size": _ppmd_dict_size,
                    "word_size": _ppmd_word_size,
                    "solid_block_size": _ppmd_solid_block_size
                }
            },
            "extension": ".7z"
        },
        "bzip2": {
            "command": "-tbzip2",
            "compress_level": _compress_level,
            "algorithm": {
                "BZip2": {
                    "command": "",
                    "dict_size": _bzip2_dict_size,
                    "word_size": "",
                    "solid_block_size": ""
                }
            },
            "extension": ".bz2"
        },
        "gzip": {
            "command": "-tgzip",
            "compress_level": _gzip_compress_level,
            "algorithm": {
                "Deflate": {
                    "command": "",
                    "dict_size": "",  # 32 KB
                    "word_size": _gzip_word_size,
                    "solid_block_size": ""
                }
            },
            "extension": ".gz"
        },
        "tar": {
            "command": "-ttar",
            "compress_level": "",
            "algorithm": {
                "tar": {
                    "command": "",
                    "dict_size": "",
                    "word_size": "",
                    "solid_block_size": ""
                }
            },
            "extension": ".tar"
        },
        "wim": {
            "command": "-twim",
            "compress_level": "",
            "algorithm": {
                "wim": {
                    "command": "",
                    "dict_size": "",
                    "word_size": "",
                    "solid_block_size": ""
                }
            },
            "extension": ".wim"
        },
        "xz": {
            "command": "-txz",
            "compress_level": _compress_level,
            "algorithm": {
                "LZMA2": {
                    "command": "",
                    "dict_size": _lzma_dict_size,
                    "word_size": _lzma_word_size,
                    "solid_block_size": ""
                }
            },
            "extension": ".xz"
        },
        "zip": {
            "command": "-tzip",
            "compress_level": _compress_level,
            "algorithm": {
                "Deflate": {
                    "command": "",
                    "dict_size": "",  # 32 KB
                    "word_size": _zip_word_size,
                    "solid_block_size": ""
                },
                "Deflate64": {
                    "command": "-mm=Deflate64",
                    "dict_size": "",  # 64 KB
                    "word_size": _deflate64_word_size,
                    "solid_block_size": ""
                },
                "BZip2": {
                    "command": "-mm=BZip2",
                    "dict_size": _bzip2_dict_size,
                    "word_size": "",
                    "solid_block_size": ""
                },
                "LZMA": {
                    "command": "-mm=LZMA",
                    "dict_size": _lzma_dict_size,
                    "word_size": _lzma_word_size,
                    "solid_block_size": ""
                },
                "PPMd": {
                    "command": "-mm=PPMd",
                    "dict_size": {
                        "1 Mo": "-md1m",
                        "2 Mo": "-md2m",
                        "3 Mo": "-md3m",
                        "4 Mo": "-md4m",
                        "6 Mo": "-md6m",
                        "8 Mo": "-md8m",
                        "12 Mo": "-md12m",
                        "16 Mo": "-md16m",
                        "24 Mo": "-md24m",
                        "32 Mo": "-md32m",
                        "48 Mo": "-md48m",
                        "64 Mo": "-md64m",
                        "96 Mo": "-md96m",
                        "128 Mo": "-md128m",
                        "192 Mo": "-md192m",
                        "256 Mo": "-md256m"
                    },
                    "word_size": {
                        "2": "-mfb2",
                        "3": "-mfb3",
                        "4": "-mfb4",
                        "5": "-mfb5",
                        "6": "-mfb6",
                        "7": "-mfb7",
                        "8": "-mfb8",
                        "10": "-mfb10",
                        "11": "-mfb11",
                        "12": "-mfb12",
                        "13": "-mfb13",
                        "14": "-mfb14",
                        "15": "-mfb15",
                        "16": "-mfb16"
                    },
                    "solid_block_size": ""
                }
            },
            "extension": ".zip"
        }
    }

    @c_log.status_bar()
    def zip(self,
            yaml_file=None,
            archive_name=None,
            files_2_zip=None,
            delete_sources_files=False,
            ignore_errors=config.DEFAULT_IGNORE_ERROR):
        r"""
        .. admonition:: Description

            | Permet de compresser les données

        :param yaml_file: fichier de configuration pour compresser ; si utilisé en argument, les 2 arguments
            :red:`archive_name` et :red:`files_2_zip` ne seront pas pris en compte.

            | Un fichier de configuration ayant pour nom ":red:`archive_name.YAML`" sera généré à la racine de
                :red:`archive_name` et peut être utilisé en argument pour :

                * compresser en mode mise à jour
                * décompresser (cf. :func:`corelibs.tools.Archive.unzip()`)

        :param archive_name: nom souhaité pour l'archive **avec** son chemin absolu pour le stockage
            (i.e. "D:\\OneDrive\\Documents\\_TEST_\\NOM_ARCHIVE")

        :param files_2_zip: le ou les fichiers à compresser

            | si fichier simple, alors définir avec une chaîne de caractères
            | si liste de fichiers alors définir les différentes chaînes de caractères dans un tuple, séparé par des
                virgules (i.e. ("fichier 1", "fichier 2", ..., "fichier n"))
            | les noms des fichiers peuvent être précédés par un chemin ou non :

                * sans le chemin en préfixe, alors l'emplacement du fichier à ziper est celui du fichier archive
                * sinon le chemin en préfixe sera pris en priorité sur le chemin de l'archive

        :param delete_sources_files: supprimer les fichiers sources une fois la compression finie (i.e. terminée sans
            erreur ou sans avoir été arrêtée par un autre processus)

            | :magenta:`valeurs possibles`: `False/True`
            | :magenta:`valeur par défaut`: `False`

        :param ignore_errors: forcer l'exécution lorsqu'une erreur est levée

            | :magenta:`valeurs possibles`: `False/True`
            | :magenta:`valeur par défaut`: `False` (cf. `DEFAULT_IGNORE_ERROR` dans :ref:`reference-label-config`)

        :return:
            | :magenta:`code retour` :

                * 0 si OK
                * <> 0 si KO

        .. note::

            fichier s'entend au sens Unix du terme, i.e. fichier régulier ou répertoire

        :green:`Exemple` :

        .. literalinclude:: ..\\..\\tests\\tools\\archive.py
            :language: python

        :green:`Exemple Fichier YAML généré` :

        .. literalinclude:: ..\\..\\tests\\tools\\TEST.yaml
            :language: yaml

        """
        if yaml_file is None and archive_name is None and files_2_zip is None:
            if not ignore_errors:
                log.warning(f"-[{config.PACKAGE_NAME}]- Il n'y a rien à compresser...")
                sys.exit(1)

        if yaml_file:
            if yaml_file == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le nom du fichier de configuration ne peut pas être vide")
                    sys.exit(1)

            _conf_file_info = lz.get_file_extension(yaml_file)
            if ".yaml" not in _conf_file_info.file_extension and ".yml" not in _conf_file_info.file_extension:
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Fichier de configuration \"{yaml_file}\" non reconnu pour "
                              f"compresser les données")
                    sys.exit(1)

            return _zip(yaml_file, delete_sources_files=delete_sources_files)
        else:
            _yaml_zip = {}

            if archive_name is None or archive_name == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le nom de l'archive ne peut pas être vide")
                    sys.exit(1)

            if files_2_zip is None or files_2_zip == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le nom de(s) fichier(s) à compresser ne peut pas être vide")
                    sys.exit(1)

            archive_dir_base = lz.get_dir_n_basename(archive_name)
            if archive_dir_base.dir_path == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le chemin de stockage du fichier archive \"{archive_name}\" "
                              f"est absent")
                    sys.exit(1)

            zipe_file_name = lz.get_file_extension(archive_dir_base.base_name).file_name

            _yaml_zip["archive"] = {
                "root": archive_dir_base.dir_path,
                "name": zipe_file_name + self._archive_command_options["7z"]["extension"]
            }

            _hr_len = 123
            _header_msg = "# Généré le " \
                          + cls.is_datetime(lz.get_timestamp(timestamp_format="NOW"),
                                            in_format="%Y-%m-%d %H:%M:%S",
                                            out_format="%A %d %B %Y à %H:%M:%S") \
                          + " "
            _yaml_zip["files"] = files_2_zip
            _yaml_zip_path = lz.get_abspath(archive_dir_base.dir_path, zipe_file_name + ".yaml")
            with open(_yaml_zip_path, "w", encoding="utf-8") as out_yaml_file:
                out_yaml_file.write("#" * _hr_len + "\n")
                out_yaml_file.write(_header_msg + "#" * (_hr_len - len(_header_msg)) + "\n")
                out_yaml_file.write("#" * _hr_len + "\n")
                # https://stackoverflow.com/questions/9169025/how-can-i-add-a-python-tuple-to-a-yaml-file-using-pyyaml
                yaml.safe_dump(_yaml_zip, out_yaml_file, default_flow_style=False, allow_unicode=True)
                out_yaml_file.write("#" * _hr_len + "\n")

            return _zip(_yaml_zip_path, delete_sources_files=delete_sources_files)

    @c_log.status_bar()
    def unzip(self,
              yaml_file=None,
              archive_name=None,
              files_2_unzip=None,
              ignore_errors=config.DEFAULT_IGNORE_ERROR):
        r"""
        .. admonition:: Description

            | Permet de décompresser les données

        :param yaml_file: fichier de configuration pour décompresser ; si utilisé en argument, les 2 arguments
            :red:`archive_name` et :red:`files_2_unzip` ne seront pas pris en compte. (ce fichier peut être créé
            manuellement ou être celui généré par la méthode :func:`corelibs.tools.Archive.zip()`)

        :param archive_name: nom du fichier archive (avec son ou ses extensions) contenant les fichiers à décompresser
            **avec** son chemin absolu (i.e. "D:\\OneDrive\\Documents\\_TEST_\\NOM_ARCHIVE.7z")

        :param files_2_unzip: le ou les fichiers à décompresser

            | si fichier simple, alors définir avec une chaîne de caractères
            | si liste de fichiers alors définir les différentes chaînes de caractères dans un tuple, séparé par des
                virgules (i.e. ("fichier 1", "fichier 2", ..., "fichier n"))
            | les noms des fichiers peuvent être précédés par un chemin ou non :

                * sans le chemin en préfixe, alors l'emplacement du fichier à décompresser est celui du fichier archive
                * sinon le chemin en préfixe sera pris en priorité sur le chemin de l'archive pour la décompression

        :param ignore_errors: forcer l'exécution lorsqu'une erreur est levée

            | :magenta:`valeurs possibles`: `False/True`
            | :magenta:`valeur par défaut`: `False` (cf. `DEFAULT_IGNORE_ERROR` dans :ref:`reference-label-config`)

        :return:
            | :magenta:`code retour` :

                * 0 si OK
                * <> 0 si KO

        :green:`Exemple` :

        .. literalinclude:: ..\\..\\tests\\tools\\unarchive.py
            :language: python

        """
        if yaml_file is None and archive_name is None and files_2_unzip is None:
            if not ignore_errors:
                log.warning(f"-[{config.PACKAGE_NAME}]- Il n'y a rien à décompresser...")
                sys.exit(1)

        if yaml_file:
            if yaml_file == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le nom du fichier de configuration ne peut pas être vide")
                    sys.exit(1)

            _conf_file_info = lz.get_file_extension(yaml_file)
            if ".yaml" not in _conf_file_info.file_extension and ".yml" not in _conf_file_info.file_extension:
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Fichier de configuration \"{yaml_file}\" non reconnu pour "
                              f"décompresser les données")
                    sys.exit(1)

            archive_yaml = _validate_zip_yaml_file(yaml_file, ignore_errors)

            archive_dir_path = archive_yaml["archive"]["root"]
            archive_name = archive_yaml["archive"]["name"]
            _is_archived_file(archive_dir_path, archive_name, ignore_errors, action="unzip")

            archive_files = yaml.safe_load(open(yaml_file, "r", encoding="utf-8"))["files"]

            _files_2_unzip = []
            if isinstance(archive_files, str):
                _files_2_unzip.append(archive_files)
            else:
                _files_2_unzip = archive_files

            return _unzip(archive_name, archive_dir_path, _files_2_unzip, ignore_errors)
        else:
            if archive_name is None or archive_name == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le nom de l'archive ne peut pas être vide")
                    sys.exit(1)

            if files_2_unzip is None or files_2_unzip == "":
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le nom de(s) fichier(s) à décompresser ne peut pas être vide")
                    sys.exit(1)

            if not lz.is_file_exists(archive_name):
                if not ignore_errors:
                    log.error(f"-[{config.PACKAGE_NAME}]- L'archive \"{archive_name}\" n'existe pas à l'emplacement "
                              f"indiqué.")
                    sys.exit(1)

            archive_dir_base = lz.get_dir_n_basename(archive_name)

            _files_2_unzip = []
            if isinstance(files_2_unzip, str):
                _files_2_unzip.append(files_2_unzip)
            else:
                _files_2_unzip = files_2_unzip

            return _unzip(archive_dir_base.base_name, archive_dir_base.dir_path, _files_2_unzip, ignore_errors)

    def __init__(self):
        pass

    @property
    def options(self):  # todo...
        return self._archive_command_options


def _count_lines(file, chunk=config.DEFAULT_BYTE_CHUNK_SIZE):
    while True:
        f = file.read(chunk)
        if not f:
            break
        yield f


def get_total_lines_in_file(file, chunk=config.DEFAULT_BYTE_CHUNK_SIZE):
    r"""
    .. admonition:: Description

        | Permet de compter le nombre de lignes dans un fichier plat

    :param file: indique l'emplacement du fichier avec son chemin absolu

    :param chunk: indique le buffer de lecture.

        | :magenta:`valeur par défaut`: `DEFAULT_BYTE_CHUNK_SIZE` (cf. :ref:`reference-label-config`)

    :return:
        | :magenta:`total lignes lues`

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\get_total_lines_in_file.py
        :language: python

    """
    if not lz.is_file_exists(file):
        log.error(f"-[{config.PACKAGE_NAME}]- L'archive \"{file}\" n'existe pas à l'emplacement indiqué.")
        sys.exit(1)

    total_lines = 0
    with open(file, "r", encoding="utf-8", errors="ignore") as input:
        total_lines += sum(bl.count("\n") for bl in _count_lines(input, chunk))

    return total_lines + 1


def get_total_lines_in_folder(dir_2_scan, files_pattern, to_exclude=None):
    r"""
    .. admonition:: Description

        | Permet de compter le total de nombres de lignes de tous les fichiers plats dans un dossier, avec possibilités
            de filtrage

    :param dir_2_scan: indique l'emplacement du répertoire à scanner en chemin absolu

    :param files_pattern: indique le schéma du scan.

    :param to_exclude: indique les exclusions, qu'elles soient des sous dossiers et/ou des fichiers. Les exclusions
        peuvent être des schémas d'exclusions.

    :return:
        :magenta:`tuple nommé` avec comme attributs :
            * :magenta:`total_files` (total de fichiers lus)
            * :magenta:`total_lines` (total de lignes lues)

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\get_total_lines_in_folder.py
        :language: python

    """
    total_lines = total_files = 0

    if not lz.is_file_exists(dir_2_scan, is_dir=True):
        log.error(f"-[{config.PACKAGE_NAME}]- Le dossier à scanner \"{dir_2_scan}\" n'existe pas")
        sys.exit(1)

    _files_pattern = []
    if isinstance(files_pattern, str):
        _files_pattern.append(files_pattern)
        files_pattern = _files_pattern

    for p in files_pattern:
        if "*" not in p:
            log.error(f"-[{config.PACKAGE_NAME}]- Le schéma des fichiers \"{files_pattern}\" est erronné, doit "
                      f"contenir des wildcards")
            sys.exit(1)

    _to_exclude = []
    if to_exclude is not None and isinstance(to_exclude, str):
        _to_exclude.append(to_exclude)
        to_exclude = _to_exclude

    if to_exclude is None:
        to_exclude = []

    # transform glob patterns to regular expressions
    files_pattern = r"|".join([fnmatch.translate(x) for x in files_pattern])
    to_exclude = r"|".join([fnmatch.translate(x) for x in to_exclude]) or r"$."
    to_exclude = to_exclude

    for root, dirs, files in os.walk(dir_2_scan):
        # exclude dirs
        dirs[:] = [os.path.join(root, d) for d in dirs]
        dirs[:] = [d for d in dirs if not re.match(to_exclude, d)]

        # exclude/include files
        files = [os.path.join(root, f) for f in files]
        files = [f for f in files if not re.match(to_exclude, f)]
        files = [f for f in files if re.match(files_pattern, f, re.IGNORECASE)]

        total_files += len(files)
        for f in files:
            total_lines += get_total_lines_in_file(f)

    Wcl = namedtuple("Wcl", ["total_files", "total_lines"])
    return Wcl(
        total_files=total_files,
        total_lines=total_lines
    )


def _derive_key(password, salt, iterations=ITERATIONS):
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=iterations,
        backend=BACKEND
    )

    return b64e(kdf.derive(password))


def _cipher(obj, password, iterations=ITERATIONS):
    salt = secrets.token_bytes(16)
    key = _derive_key(password.encode(), salt, iterations)

    _ = b64e(
        b'%b%b%b' % (
            salt,
            iterations.to_bytes(4, "big"),
            b64d(Fernet(key).encrypt(obj)),
        )
    )

    return zlib.compress(_, zlib.Z_BEST_COMPRESSION)


def cipher(obj, password, delete_source_file=True):
    r"""
    .. admonition:: Description

        | Permet de chiffrer une chaîne de caractère ou un fichier

    .. warning::

        | Ce chiffrement est adapté pour protéger des fichiers plats contenant du code ou partie de codes sensibles, ou
            un fichier de petite taille.

        | Le mot de passe tient compte de la casse.

    :param obj: indique la chaine de caractère ou le chemin absolu du fichier à chiffrer

    :param password: indique le mot de passe à appliquer lors du chiffrement

    :param delete_source_file: indique si le fichier source doit être supprimé après le chiffrement

            | :magenta:`valeurs possibles`: `False/True`
            | :magenta:`valeur par défaut`: `True`

    :return:
        | :magenta:`True`
        | ou
        | :magenta:`chaîne chiffrée` (si :red:`obj` est une chaîne de caractère)

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\cipher.py
        :language: python

    """
    if not isinstance(obj, str):
        log.error(f"-[{config.PACKAGE_NAME}]- L'argument \"{obj}\" n'est pas une chaine de caractères")
        sys.exit(1)

    if lz.is_file_exists(obj, is_dir=True, ignore_errors=True):
        log.error(f"-[{config.PACKAGE_NAME}]- Le chiffrement ne peut pas se faire sur un dossier \"{obj}\"")
        sys.exit(1)

    if lz.is_file_exists(obj, ignore_errors=True):
        with open(obj, "rb") as f_in, open(obj + CIPHER_FILE_EXTENSION, "wb") as f_out:
            f_out.write(_cipher(f_in.read(), password))

        if delete_source_file:
            lz.delete_files(obj, remove_empty_dir=False, verbose=config.DEFAULT_VERBOSE)

        return True

    if not re.search("\n", obj, re.M) and ("/" in obj or "\\" in obj):
        log.error(f"-[{config.PACKAGE_NAME}]- Le fichier \"{obj}\" à chiffre n'existe pas")
        sys.exit(1)

    return _cipher(str(obj).encode(), password)


def _decipher(obj, password):
    decoded = b64d(zlib.decompress(obj))
    salt, iter, token = decoded[:16], decoded[16:20], b64e(decoded[20:])
    iterations = int.from_bytes(iter, "big")
    key = _derive_key(password.encode(), salt, iterations)

    try:
        return Fernet(key).decrypt(token)
    except InvalidToken:
        log.error(f"-[{config.PACKAGE_NAME}]- Le mot de passe \"{password}\" est incorrect")
        sys.exit(1)


def decipher(obj, password, delete_source_file=True):
    r"""
    .. admonition:: Description

        | Permet de déchiffrer une chaîne de caractère ou un fichier

    .. warning::

        | Le mot de passe tient compte de la casse.

    :param obj: indique la chaine de caractère ou le chemin absolu du fichier à déchiffrer

    :param password: indique le mot de passe à appliquer lors du déchiffrement

    :param delete_source_file: indique si le fichier source doit être supprimé après le déchiffrement

            | :magenta:`valeurs possibles`: `False/True`
            | :magenta:`valeur par défaut`: `True`

    :return:
        | :magenta:`True`
        | ou
        | :magenta:`chaîne déchiffrée` (si :red:`obj` est une chaîne binaire)

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\decipher.py
        :language: python

    """
    if not isinstance(obj, bytes):
        if isinstance(obj, str):
            if lz.is_file_exists(obj, is_dir=True, ignore_errors=True):
                log.error(f"-[{config.PACKAGE_NAME}]- Le déchiffrement ne peut pas se faire sur un dossier \"{obj}\"")
                sys.exit(1)

            if lz.is_file_exists(obj, ignore_errors=True):
                obj_ext = lz.get_file_extension(obj, split_extensions=True)

                if "".join(obj_ext.file_extension[-1:]) != CIPHER_FILE_EXTENSION:
                    log.error(f"-[{config.PACKAGE_NAME}]- Le fichier \"{obj}\" à déchiffrer n'a pas la bonne extension "
                              f"\"{CIPHER_FILE_EXTENSION}\"")
                    sys.exit(1)

                with open(obj, "rb") as f_in, open(
                        lz.get_abspath(
                            lz.get_dir_n_basename(obj).dir_path,
                            obj_ext.file_name + (
                                    "".join(obj_ext.file_extension[:-1])
                                    if "".join(obj_ext.file_extension[:-1])
                                    else "".join(obj_ext.file_extension[-1:])
                            )
                        ), "wb") as f_out:
                    f_out.write(_decipher(f_in.read(), password))

                if delete_source_file:
                    lz.delete_files(obj, remove_empty_dir=False, verbose=config.DEFAULT_VERBOSE)

                return True
            else:
                log.error(f"-[{config.PACKAGE_NAME}]- Le fichier \"{obj}\" à déchiffre n'existe pas")
                sys.exit(1)
        else:
            log.error(f"-[{config.PACKAGE_NAME}]- L'argument \"{obj}\" n'est pas une chaine binaire")
            sys.exit(1)

    return _decipher(obj, password).decode()


pre_scan_files_info = pre_scan_total_duration = None
total_directories = total_files = total_directories_size = total_files_size = total_size = 0


def _pre_scan(dir_path, encoding=config.DEFAULT_ENCODING_FILE, skip_directories_properties=True, std_print=False):
    global pre_scan_files_info, pre_scan_total_duration, \
        total_directories, total_files, total_directories_size, total_files_size, total_size, scan_columns_headers

    _cl = status_bar = sub_status_bar_desc = None
    if std_print is False:
        _cl = c_log.ColorLog(name="Pré-scan", output_2_log_file=False, log_level=20)
        status_bar = c_log.StatusBars(title=f"Pré-scan dossier \"{dir_path}\"")
        sub_status_bar_desc = status_bar.init_sub_process(color="magenta")

    temp_file = tempfile.NamedTemporaryFile(
        mode="w",
        suffix=config.PACKAGE_SUFFIX_TMP_NAME,
        delete=False,
        encoding=encoding
    )

    header = ";".join(scan_columns_headers)
    temp_file.write(header + "\n")

    for root, dirs, files in os.walk(os.path.expanduser(dir_path)):
        dirs.sort()
        total_directories += 1

        tech_id = get_fingerprint(
            str(root).replace("\\", ":").replace("/", ":").lower(), eval_as_string=True, algorithm="md5"
        )

        if skip_directories_properties is False:
            _d = get_file_properties(root, pretty_byte_size=False, ignore_errors=True)
            _sd = _d.st_size

            _psd = lz.get_bytes_size_4_human(lz.get_bytes_size_formats(_sd.byte))

            temp_file.write(
                f'"{tech_id}";"D";"{root}";"{_d.st_uid}";"{_d.st_atime}";"{_d.st_mtime}";'
                f'"{_d.st_ctime}";"{_psd}";"{_sd.byte}";"{_sd.kilobyte}";"{_sd.megabyte}";"{_sd.gigabyte}";;;\n'
            )

            d_info = f"•[ Dossier détecté : \"{root}\" • Poids : {_psd} ]•"
            # total_directories_size += _sd.byte
        else:
            temp_file.write(
                f'"{tech_id}";"D";"{root}";;;;;;;;;;;;\n'
            )

            d_info = f"•[ Dossier détecté : \"{root}\" ]•"

        if std_print is False:
            if lz.is_jupyter() or lz.is_interactive_python():
                _cl.info(d_info)
            else:
                status_bar.update_sub_process(sub_status_bar_desc, d_info)
        else:
            print(" D " + d_info[21:-3].replace('"', ""))

        for file in sorted(files):
            total_files += 1

            _file = root + "\\" + file
            _ = get_file_properties(_file, pretty_byte_size=False, ignore_errors=True)
            _s = _.st_size

            _ps = lz.get_bytes_size_4_human(lz.get_bytes_size_formats(_s.byte))

            tech_id = get_fingerprint(
                str(_file).replace("\\", ":").replace("/", ":").lower(), eval_as_string=True, algorithm="md5"
            )
            temp_file.write(
                f'"{tech_id}";"-";"{_file}";"{_.st_uid}";"{_.st_atime}";"{_.st_mtime}";"{_.st_ctime}";'
                f'"{_ps}";"{_s.byte}";"{_s.kilobyte}";"{_s.megabyte}";"{_s.gigabyte}";;;\n'
            )

            f_info = f"•[ Fichier détecté : \"{_file}\" • Poids : {_ps} ]•"

            if std_print is False:
                if lz.is_jupyter() or lz.is_interactive_python():
                    _cl.info(f_info)
                else:
                    status_bar.update_sub_process(sub_status_bar_desc, f_info)
            else:
                print(" - ", f_info[21:-3].replace('"', ""))

            total_files_size += _s.byte

        total_size = total_directories_size = total_files_size

    total_directories -= 1
    temp_file.close()

    if std_print is False:
        f_info = f"•[ Fin Pré-scan • {total_files} Fichiers, {total_directories} Dossiers " \
                 f"• Poids total : {lz.get_bytes_size_4_human(lz.get_bytes_size_formats(total_size))} ]•"
        if lz.is_jupyter() or lz.is_interactive_python():
            _cl.info(f_info)
        else:
            status_bar.update_sub_process(sub_status_bar_desc, f_info)

        pre_scan_total_duration = status_bar.terminate()

    scan_properties = {
        "dir_path": dir_path,
        "files": {
            "total": total_files,
            "size": lz.get_bytes_size_4_human(lz.get_bytes_size_formats(total_files_size))
        },
        "dir": {
            "total": total_directories,
            "size": lz.get_bytes_size_4_human(lz.get_bytes_size_formats(total_directories_size))
        },
        "dir_n_files": {
            "total": total_files + total_directories,
            "size": lz.get_bytes_size_4_human(lz.get_bytes_size_formats(total_size))
        }
    }
    scan_out_file_name = get_fingerprint(
        str(dir_path).replace("\\", "_").replace("/", "_").lower(), eval_as_string=True
    )
    with open(lz.get_abspath(user_corelibs_cache_root, scan_out_file_name + ".json"), "w") as f_out:
        json.dump(scan_properties, f_out)

    return temp_file.name


_current_file_sb = _previous_file_sb = None


def _status_current_file(manager, file, init=False, position=2):
    global _current_file_sb, _previous_file_sb

    if init:
        if position == 2:
            _current_file_sb = manager.status_bar(
                status_format="{phases}",
                phases=file,
                position=position,
                fill=' ',
                justify=enlighten.Justify.CENTER
            )
        else:
            _previous_file_sb = manager.status_bar(
                status_format="{phases}",
                phases=file,
                position=position,
                fill=' ',
                justify=enlighten.Justify.CENTER
            )

    if position == 2:
        _current_file_sb.update(
            status_format="{phases}",
            phases=file,
        )
    else:
        _previous_file_sb.update(
            status_format="{phases}",
            phases=file,
        )


_progress_bar = None


def _finger_print_progress_bar(manager, initials, init=False, count_done=None, count_remain=None, total_elapsed=None):
    global _progress_bar

    bar_format = u"{desc}{desc_pad}{percentage:3.0f}% | {bar} | " + \
                 u"{count_done:{len_total}d}/{count_total_2_process:{len_total}d} • " + \
                 u"(Traités/Total - Reste:{count_remain:{len_total}d}) • " + \
                 u"Temps écoulé {total_elapsed}"

    if init:
        _progress_bar = manager.counter(total=initials, desc="Avancement :", color="green",
                                        count_total_2_process=total_files_2_finger_print,
                                        count_done=current_file_iter,
                                        count_remain=total_files_2_finger_print - current_file_iter,
                                        total_elapsed=0,
                                        bar_format=bar_format, position=4)
    else:
        _progress_bar.update(count_done=count_done, count_remain=count_remain, total_elapsed=total_elapsed)

    _progress_bar.close()


total_files_2_finger_print = current_file_iter = 0
step_duration = total_duration = None


def _csv_2_df(file_path, encoding=config.DEFAULT_ENCODING_FILE):
    return pd.read_csv(
        file_path,
        sep=";",
        engine="python",
        quotechar='"',
        index_col=False,
        encoding=encoding
    ).fillna("")


def _get_finger_print(df, std_print=False, gui_instance=None):
    global total_files_2_finger_print, current_file_iter, total_duration, step_duration

    df_files = df[(df["File Type"] == "-") & (df["Fingerprint"] == "")]
    total_files_2_finger_print = len(df_files.index)
    len_total = len(str(total_files_2_finger_print))

    _cl = status_bar = progress_bar_desc = sub_status_bar_desc = None
    if std_print is False:
        _cl = c_log.ColorLog(name="Calcul Empreinte Digital", output_2_log_file=False, log_level=20)
        status_bar = c_log.StatusBars(title="Calcul Empreinte Digitale")
        progress_bar_desc = status_bar.init_progress_bar(total=total_files_2_finger_print, display_status=False)
        sub_status_bar_desc = status_bar.init_sub_process(color="magenta")

    files_2_process = [(i, fn, fs) for i, fn, fs in zip(df_files.index, df_files["File Name"], df_files["File Size"])]

    if total_files_2_finger_print > 0:
        if gui_instance is not None and isinstance(gui_instance, sg.PySimpleGUI.Window):
            gui_instance["-SCAN_FRAME_LAYOUT-"].update(visible=True)
            gui_instance["-TEXT_SCAN_INFO-"].update(visible=True)
            gui_instance["-PROGRESS_BAR-"].update(visible=True, current_count=0)

        for i, f in enumerate(files_2_process):
            current_file_iter = current_file_iter + 1

            if gui_instance is not None and isinstance(gui_instance, sg.PySimpleGUI.Window):
                gui_instance["-SCAN_FRAME_LAYOUT-"]\
                    .update(f"Calcul Empreinte Digitale "
                            f"[ "
                            f"{(i + 1): >{len_total}} / {total_files_2_finger_print} "
                            f"- Reste {(total_files_2_finger_print - i - 1): >{len_total}} "
                            f" • {int(round((i + 1) * 100 / total_files_2_finger_print, 0))}%"
                            f" ]")
                gui_instance["-TEXT_SCAN_INFO-"].update(
                    f"Fichier en cours de traitement : {f[1]} ({f[2]})"
                )
                gui_instance["-PROGRESS_BAR-"].UpdateBar(current_file_iter, total_files_2_finger_print)

            if std_print is False:
                if lz.is_jupyter() or lz.is_interactive_python():
                    _cl.info("{2}/{3} •[ Fichier en cours de traitement : \"{0}\" • Poids : {1} ]•"
                             .format(f[1], f[2],
                                     str(current_file_iter).rjust(len_total, " "),
                                     str(total_files_2_finger_print).rjust(len_total, " ")
                                     ))
                else:
                    status_bar.update_sub_process(sub_status_bar_desc, "•[ Fichier en cours de traitement : \"{0}\" • "
                                                                       "Poids : {1} ]•".format(f[1], f[2]))
                    step_duration = status_bar.update_progress_bar(
                        progress_bar_desc,
                        status_text=f"{(i + 1): >{len_total}}/{total_files_2_finger_print} "
                                    f"- Reste {(total_files_2_finger_print - i - 1): >{len_total}}",
                    )

            finger_print = get_fingerprint(f[1], ignore_errors=True)
            df_files.at[f[0], "Fingerprint"] = finger_print

        if std_print is False:
            total_duration = status_bar.terminate()

        return df_files
    else:
        msg = "•[ Tous les fichiers sont déjà présents dans le cache, aucune action ne sera appliquée... ]•"

        if std_print is False:
            if lz.is_jupyter() or lz.is_interactive_python():
                _cl.info(msg)
            else:
                status_bar.update_sub_process(sub_status_bar_desc, msg)

            total_duration = status_bar.terminate()

        return None


def _set_duplicated_indicator(df):
    _df = df[["Fingerprint", "Is Duplicated"]]
    _df = _df[(_df["Fingerprint"] != "")]
    _df = _df[_df.duplicated()]
    _df["Is Duplicated"] = "Yes"
    df = df[["Technical ID", "Fingerprint"]]
    df_no_dup = _df.drop_duplicates()

    return pd.merge(df, df_no_dup, how="left", on=["Fingerprint"], copy=False).fillna("")


scan_columns_headers = ["Technical ID", "File Type", "File Name", "Owner", "Last Access", "Last Modification",
                        "OS Time", "File Size", "Bytes", "KiloBytes", "MegaBytes", "GigaBytes", "Fingerprint",
                        "Is Duplicated"]


def _consolidate_df(ref_df, dup_df):
    global scan_columns_headers

    ref_df_with_key = ref_df[["Technical ID"]]
    ref_df_dir = ref_df[(ref_df["File Type"] == "D")]
    ref_df_with_finger = ref_df[(ref_df["Fingerprint"] != "")]
    ref_df_without_finger = ref_df[(ref_df["File Type"] == "-") & (ref_df["Fingerprint"] == "")]

    ref_df_with_finger = ref_df_with_finger.drop(["Fingerprint", "Is Duplicated"], axis=1)
    ref_df_with_finger = ref_df_with_finger.merge(dup_df, how="left", on=["Technical ID"], copy=False).fillna("")

    df_consolidated = ref_df_with_finger \
        .append(ref_df_without_finger, ignore_index=True) \
        .append(ref_df_dir, ignore_index=True)

    _df = ref_df_with_key.merge(df_consolidated, how="left", on=["Technical ID"], copy=False).fillna("")

    return _df[scan_columns_headers]


def _render_on_excel(dir_path, df):
    excel_template = lz.get_abspath(user_corelibs_templates_root, SCAN_DIR_TEMPLATE_FILE)
    if not lz.is_file_exists(excel_template, ignore_errors=True):
        lz.copy(lz.get_abspath(corelibs_templates_root, SCAN_DIR_TEMPLATE_FILE), excel_template)

    scan_out_file_name = get_fingerprint(
        str(dir_path).replace("\\", "_").replace("/", "_").lower(), eval_as_string=True
    )
    _cache_file = lz.get_abspath(user_corelibs_cache_root, scan_out_file_name)
    scan_properties_json = _cache_file + ".json"
    if not lz.is_file_exists(scan_properties_json, ignore_errors=True):
        scan_properties = {
            "dir_path": dir_path,
            "files": {
                "total": "Info Supprimée",
                "size": "Info Supprimée"
            },
            "dir": {
                "total": "Info Supprimée",
                "size": "Info Supprimée"
            },
            "dir_n_files": {
                "total": "Info Supprimée",
                "size": "Info Supprimée"
            }
        }
    else:
        with open(scan_properties_json, "r") as f_in:
            scan_properties = json.load(f_in)

    wb = load_workbook(excel_template)
    if wb.sheetnames[0] == "data":
        wb.active = 0

    try:
        _has_finger_df = df[df["Fingerprint"] != ""]
        total_finger = len(_has_finger_df.index)
        _ = df[df["Is Duplicated"] != ""]
        dup_files_df = _.groupby(["Fingerprint"]).size()
        total_dup_files = len(dup_files_df.index)
        total_overall_dup_files = len(_.index)
    except KeyError:
        total_finger = total_dup_files = total_overall_dup_files = 0

    ws = wb.active
    ws["B1"] = scan_properties["dir_path"]
    _cache_file_properties = get_file_properties(_cache_file, pretty_byte_size=False, ignore_errors=True)
    ws["B2"] = _cache_file_properties.st_ctime

    ws["C4"] = scan_properties["files"]["total"]
    if total_finger == 0:
        ws["C5"] = "Non calculé"
    else:
        if total_dup_files > 0:
            total_overall_dup_files = "{0:,}".format(total_overall_dup_files).replace(",", " ")
            total_dup_files = "{0:,}".format(total_dup_files).replace(",", " ")
            ws["C5"] = f"{total_overall_dup_files} fichier(s) au total pour {total_dup_files} fichier(s) distinct(s)"
        else:
            ws["C5"] = "Aucun fichier détecté"
    ws["C6"] = scan_properties["dir"]["total"]
    ws["C7"] = scan_properties["dir_n_files"]["total"]
    ws["C8"] = scan_properties["dir_n_files"]["size"]

    xls_row = 13

    try:
        for t_i, f_t, f_n, o, l_a, l_m, ot, f_s, b, k_b, m_b, g_b, f_p, i_d \
                in zip(df["Technical ID"], df["File Type"], df["File Name"], df["Owner"], df["Last Access"],
                       df["Last Modification"], df["OS Time"], df["File Size"], df["Bytes"], df["KiloBytes"],
                       df["MegaBytes"],
                       df["GigaBytes"], df["Fingerprint"], df["Is Duplicated"]):
            ws[f"A{xls_row}"] = t_i
            ws[f"B{xls_row}"] = f_t
            ws[f"C{xls_row}"] = f_n
            ws[f"D{xls_row}"] = o
            ws[f"E{xls_row}"] = l_a
            ws[f"F{xls_row}"] = l_m
            ws[f"G{xls_row}"] = ot
            ws[f"H{xls_row}"] = f_s
            ws[f"I{xls_row}"] = b
            ws[f"J{xls_row}"] = k_b
            ws[f"K{xls_row}"] = m_b
            ws[f"L{xls_row}"] = g_b
            ws[f"M{xls_row}"] = f_p
            ws[f"N{xls_row}"] = "Oui" if str(i_d).lower() == "yes" else ""

            xls_row += 1
    except KeyError:
        for t_i, f_t, f_n, o, l_a, l_m, ot, f_s, b, k_b, m_b, g_b \
                in zip(df["Technical ID"], df["File Type"], df["File Name"], df["Owner"], df["Last Access"],
                       df["Last Modification"], df["OS Time"], df["File Size"], df["Bytes"], df["KiloBytes"],
                       df["MegaBytes"], df["GigaBytes"]):
            ws[f"A{xls_row}"] = t_i
            ws[f"B{xls_row}"] = f_t
            ws[f"C{xls_row}"] = f_n
            ws[f"D{xls_row}"] = o
            ws[f"E{xls_row}"] = l_a
            ws[f"F{xls_row}"] = l_m
            ws[f"G{xls_row}"] = ot
            ws[f"H{xls_row}"] = f_s
            ws[f"I{xls_row}"] = b
            ws[f"J{xls_row}"] = k_b
            ws[f"K{xls_row}"] = m_b
            ws[f"L{xls_row}"] = g_b

            xls_row += 1
    except ValueError as e:
        log.critical(f"-[{config.PACKAGE_NAME}]- {e}")

    _cache_file_xls = _cache_file + ".xlsx"
    try:
        wb.save(_cache_file_xls)
        os.startfile(_cache_file_xls)
    except PermissionError as e:
        log.critical(f"-[{config.PACKAGE_NAME}]- Permission insuffisante, fichier {scan_out_file_name}.xlsx ouvert en "
                     f"lecture")


def _get_last_access_from_cache_file(df, df_2_merge):
    _df = df.drop(["Fingerprint", "Is Duplicated"], axis=1)
    _df = _df.merge(df_2_merge, how="left", on=["Technical ID"], copy=False).fillna("")
    _df["Last Access"] = np.where(_df["Last Access_y"] == "", _df["Last Access_x"], _df["Last Access_y"])
    _df = _df[scan_columns_headers]

    return _df


def _col_as_str(df):
    all_columns = list(df)
    df[all_columns] = df[all_columns].astype(str)

    return df


EXCEL_LIMIT = 1048000


def scan_dir(dir_path,
             duplicated_files_indicator=False,
             skip_directories_properties=True,
             skip_pre_scan=False,
             ref_scan_file=None,
             caching=True,
             render="Excel",
             encoding="utf-8",
             std_print=False, gui_instance=None, force_excel_2_refresh=False):
    r"""
    .. admonition:: Description

        | Permet de scanner un disque ou répertoire afin d'analyser son contenu, et en particulier d'y détecter des
            fichiers en doublon. Le scan s'appuie sur un caching pour réduire les temps de calcul et fonctionnera donc
            en mode delta à partir du 2ème scan.

    :param dir_path: indique chemin à analyser

    :param duplicated_files_indicator: indique s'il faut identifier les fichiers en doublon

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False`

    :param skip_directories_properties: indique s'il faut éviter le calcul des propriétés dossiers (pouvant être
        coûteux en temps d'exécution)

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `True`

    :param skip_pre_scan: indique s'il faut éviter le préscan. Lorsqu'il est à True, alors le :red:`ref_scan_file` sera
        pris en compte. Si ce dernier est à `None`, le cache sera utilisé (si existe), sinon une alerte sera levée. Et
        il faudrait donc préciser l'emplacement d'un fichier contenant les données scannées.

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False`

    :param ref_scan_file: indique le fichier référentiel du préscan. Si la valeur est à `None` alors le cache du préscan
        sera pris en compte si existe. Sinon le fichier préscan fournit en entrée doit avoir les conditions suivantes :

        * être un fichier CSV, séparé par ;
        * avoir l'entête suivante "Technical ID";"File Type";"File Name";"Owner";"Last Access";"Last Modification";"OS Time";"File Size";"Bytes";"KiloBytes";"MegaBytes";"GigaBytes";"Fingerprint";"Is Duplicated"
        * assurer l'unicité du champ "Technical ID"

    :param caching: indique s'il faut utiliser ou non le caching

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `True`

    :param render: indique le rendu final

        | :magenta:`valeurs possibles`: `Excel`, `Web` ou `App`
        | :magenta:`valeur par défaut`: `Excel`

    :param encoding: indique l'encodage de lecture/écriture

    :param std_print: indique s'il faut utiliser la sortie standard pour le printing

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False`

    :param gui_instance: indique l'instance gui (pour l'interface visuelle, non utile en instructions codes)

    :param force_excel_2_refresh: indique s'il faut forcer le rafraichissement des fichiers caches Excel

        | :magenta:`valeurs possibles`: `False/True`
        | :magenta:`valeur par défaut`: `False`

    :return:
        | :magenta:`rien...`

    .. warning::

        | Le rendu "Web" n'est possible que sous Jupyter ou iPYthon

    .. note::

        | A partir du 2ème scan, si le calcul des doublons est souhaité, les fichiers du préscan ne se trouvant pas dans
            le cache seront calculés. Si les fichiers du préscan se trouvent dans le cache, alors le calcul se
            rafraichit dans les cas suivants :

        * la taille du fichier a été modifiée
        * la date de modification a été changée

    :green:`Exemple` :

    .. literalinclude:: ..\\..\\tests\\tools\\scan_dir.py
        :language: python

    .. figure:: ..\\ss\\scan_dir.png

        Scan via code

    .. figure:: ..\\ss\\scan_dir_gui.png

        Scan via interface

    """
    global user_corelibs_cache_root, scan_columns_headers, \
        total_directories, total_files, total_directories_size, total_files_size, total_size

    total_directories = total_files = total_directories_size = total_files_size = total_size = 0

    # dir_path = dir_path.replace("\\\\", "\\") <=> induit problème avec les chemins réseaux

    if not lz.is_file_exists(dir_path, is_dir=True, ignore_errors=True):
        log.error(f"-[{config.PACKAGE_NAME}]- Le chemin \"{dir_path}\" n'existe pas")
        sys.exit(1)

    if str(render).lower() not in ("excel", "web", "app"):
        log.error(f"-[{config.PACKAGE_NAME}]- Argument render non valide, valeurs possibles "
                  f"\"{('Excel', 'Web', 'App')}\"")
        sys.exit(1)

    scan_out_file_name = get_fingerprint(
        str(dir_path).replace("\\", "_").replace("/", "_").lower(), eval_as_string=True
    )

    _cache_file = lz.get_abspath(user_corelibs_cache_root, scan_out_file_name)
    _history_file = lz.get_abspath(user_corelibs_cache_root, "history.json")
    if not lz.is_file_exists(_history_file, ignore_errors=True):
        history = {}
    else:
        with open(_history_file, "r") as f_in:
            try:
                history = json.load(f_in)
            except json.decoder.JSONDecodeError:
                history = {}

    if dir_path not in history.values():
        history.update({len(history): dir_path})

    with open(_history_file, "w") as f_out:
        json.dump(history, f_out)

    _tkam_suffix = "_TKAM"
    if skip_pre_scan is False:
        shutil.move(
            _pre_scan(dir_path,
                      encoding=encoding,
                      skip_directories_properties=skip_directories_properties,
                      std_print=std_print),
            _cache_file + _tkam_suffix
        )

        temp_file = _cache_file + _tkam_suffix
    else:
        if ref_scan_file is None:
            if lz.is_file_exists(_cache_file + ".xlsx") \
                    and render.lower() == "excel" and force_excel_2_refresh is False:
                os.startfile(_cache_file + ".xlsx")
                sys.exit(0)
            else:
                if lz.is_file_exists(_cache_file + _tkam_suffix):
                    temp_file = _cache_file + _tkam_suffix
                else:
                    log.error(
                        f"-[{config.PACKAGE_NAME}]- Il n'existe pas de fichier cache pour le répertoire \"{dir_path}\" "
                        f"qui est scanné pour la première fois. Si le pré-scan n'est pas souhaité, le fichier scan de "
                        f"référence (ref_scan_file) ne doit pas être vide"
                    )
                    sys.exit(1)
        else:
            if not lz.is_file_exists(ref_scan_file):
                log.error(f"-[{config.PACKAGE_NAME}]- Le fichier scan de référence \"{ref_scan_file}\" n'existe pas")
                sys.exit(1)

            has_same_header = data.has_same_header(_cache_file, ref_scan_file)
            if has_same_header.result is False:
                log.error(f"-[{config.PACKAGE_NAME}]- Les entêtes du fichier cache avec le fichier de référence "
                          f"\"{ref_scan_file}\" diffèrent\n"
                          f"\t{has_same_header.cause}. "
                          f"Format entête attendu \"{scan_columns_headers}\"")
                sys.exit(1)

            temp_file = ref_scan_file

    _temp_file_df = _csv_2_df(temp_file, encoding=encoding)

    _cache_file_df = None
    if caching and lz.is_file_exists(_cache_file):
        _cache_file_df = _csv_2_df(_cache_file, encoding=encoding)
        _cache_file_df = _cache_file_df[_cache_file_df["File Type"] == "-"]
        _cache_file_df = _temp_file_df[["Technical ID", "File Name", "Last Modification", "Bytes"]].merge(
            _cache_file_df[["Technical ID", "Last Modification", "Bytes", "Fingerprint", "Is Duplicated",
                            "Last Access"]], how="left", on=["Technical ID"], copy=False).fillna("")
        _cache_file_df["Fingerprint"] = np.where(
            (_cache_file_df["Bytes_x"] != _cache_file_df["Bytes_y"])
            | (_cache_file_df["Last Modification_x"] != _cache_file_df["Last Modification_y"]),
            "", _cache_file_df["Fingerprint"])
        _cache_file_df = _cache_file_df[["Technical ID", "Fingerprint", "Is Duplicated", "Last Access"]]
        _temp_file_df = _get_last_access_from_cache_file(_temp_file_df, _cache_file_df)

    if duplicated_files_indicator:
        _df_wip_finger = _get_finger_print(_temp_file_df, std_print=std_print, gui_instance=gui_instance)
        if _df_wip_finger is not None:
            _df_wip_finger = _df_wip_finger[["Technical ID", "Fingerprint"]]
            _temp_file_df = _temp_file_df.merge(_df_wip_finger, how="left", on=["Technical ID"], copy=False).fillna("")
            _temp_file_df["Fingerprint"] = np.where(_temp_file_df["Fingerprint_x"] == "",
                                                    _temp_file_df["Fingerprint_y"], _temp_file_df["Fingerprint_x"])
            _temp_file_df = _temp_file_df[scan_columns_headers]
            dup_def = _set_duplicated_indicator(_temp_file_df)
            final_df = _consolidate_df(_temp_file_df, dup_def)
        else:
            final_df = _temp_file_df
    else:
        final_df = _csv_2_df(_cache_file + _tkam_suffix, encoding=encoding)
        if caching and _cache_file_df is not None:
            final_df = _get_last_access_from_cache_file(final_df, _cache_file_df)

    if caching:
        final_df.to_csv(_cache_file, sep=";", quotechar='"', index=False, encoding=encoding, quoting=csv.QUOTE_ALL)

    if render.lower() == "web":
        final_df = _col_as_str(final_df)
        data.preview(final_df)
    elif render.lower() == "app":
        final_df = _col_as_str(final_df)
        show(final_df)
    else:
        if len(final_df.index) > EXCEL_LIMIT:
            final_df = _col_as_str(final_df)
            show(final_df)
        else:
            _render_on_excel(dir_path, final_df)
