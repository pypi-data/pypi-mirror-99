import openpyxl
from copy import copy
from openpyxl.styles import PatternFill, Border, Alignment, Font
"""
Documentation
"""
class Variable:
    """
    Documentation
    """
    def __init__(self, name, row, column, value):
        """
        Documentation
        """
        self.__name = name
        self.__row = row
        self.__column = column
        self.__value = value

    @property
    def name(self):
        """
        Documentation
        """
        return self.__name

    @name.setter
    def name(self, name):
        """
        Documentation
        """
        self.__name = name
        return None

    @property
    def row(self):
        """
        Documentation
        """
        return self.__row

    @row.setter
    def row(self, row):
        """
        Documentation
        """
        self.__row = row
        return None

    @property
    def column(self):
        """
        Documentation
        """
        return self.__column

    @column.setter
    def column(self, column):
        """
        Documentation
        """
        self.__column = column
        return None

    @property
    def value(self):
        """
        Documentation
        """
        return self.__value

    @value.setter
    def value(self, value):
        """
        Documentation
        """
        self.__value = value
        return None


class Style:
    """
    Documentation
    """
    def __init__(self, name, row, column, data_type, font, alignment, fill, border):
        """
        Documentation
        """
        self.__name = name
        self.__row = row
        self.__column = column
        self.__data_type = data_type
        self.__font = copy(font)
        self.__alignment = copy(alignment)
        self.__fill = copy(fill)
        self.__border = copy(border)

    @property
    def name(self):
        """
        Documentation
        """
        return self.__name

    @name.setter
    def name(self, name):
        """
        Documentation
        """
        self.__name = name
        return None

    @property
    def row(self):
        """
        Documentation
        """
        return self.__row

    @row.setter
    def row(self, row):
        """
        Documentation
        """
        self.__row = row
        return None

    @property
    def column(self):
        """
        Documentation
        """
        return self.__column

    @column.setter
    def column(self, column):
        """
        Documentation
        """
        self.__column = column
        return None

    @property
    def data_type(self):
        """
        Documentation
        """
        return self.__data_type

    @property
    def font(self):
        """
        Documentation
        """
        return self.__font

    @property
    def alignment(self):
        """
        Documentation
        """
        return self.__alignment

    @property
    def fill(self):
        """
        Documentation
        """
        return self.__fill

    @property
    def border(self):
        """
        Documentation
        """
        return self.__border

    def rename(self):
        """
        Documentation
        """
        self.__name =  chr(self.__column+64)+str(self.__row)
        return None

class Model:
    """
    Documentation
    """
    def __init__(self, filename, worksheetname):
        """
        Documentation
        """
        self.__workbook = openpyxl.load_workbook(filename)
        self.__worksheet = self.__workbook[worksheetname]

        numbers = []
        for range in self.__worksheet.merged_cells.ranges:
            s = str(range).split(':')[0]
            number = int(s[1:len(s)])
            numbers.append(number)
        self.__rows = max(numbers)

        s = str(self.__worksheet.merged_cells.ranges[0]).split(':')[1]
        while s.isalnum():
            s = s[0:len(s)-1]
            if s.isalpha(): break
        self.__columns = ord(s) - 64

        self.__variables = []
        self.__styles = []
        for row in self.__worksheet.iter_rows(min_row=1, max_col=self.__columns, max_row=self.__rows):
            for cell in row:
                if cell.value != None:
                    value = str(cell.value)
                    if value[0]=='$':
                        self.__variables.append(Variable(value[1:len(value)], cell.row, cell.column, None))
                    else:
                        self.__variables.append(Variable('Constanta', cell.row, cell.column, cell.value))
                self.__styles.append(Style(cell.coordinate, cell.row, cell.column, cell.data_type, cell.font, cell.alignment, cell.fill, cell.border))

        self.__columns_width = []
        for row in self.__worksheet.iter_rows(min_row=1, max_col=self.__columns, max_row=1):
            for cell in row:
                s = cell.coordinate
                while s.isalnum():
                    s = s[0:len(s)-1]
                    if s.isalpha(): break
                value = self.__worksheet.column_dimensions[s].width
                self.__columns_width.append(Variable(s, cell.row, cell.column, value))

        self.__rows_height = []
        for row in self.__worksheet.iter_rows(min_row=1, max_col=1, max_row=self.__rows):
            for cell in row:
                s = cell.coordinate
                number = int(s[1:len(s)])
                value = self.__worksheet.row_dimensions[number].height
                self.__rows_height.append(Variable(number, cell.row, cell.column, value))

        self.__mergedcells = []
        for range in self.__worksheet.merged_cells.ranges:
            self.__mergedcells.append(Variable(str(range), None, None, range))

    @property
    def rows(self):
        """
        Documentation
        """
        return self.__rows

    @rows.setter
    def rows(self, rows):
        """
        Documentation
        """
        self.__rows = rows
        return None

    @property
    def columns(self):
        """
        Documentation
        """
        return self.__columns

    @columns.setter
    def columns(self, columns):
        """
        Documentation
        """
        self.__columns = columns
        return None

    @property
    def worksheets(self):
        """
        Documentation
        """
        return self.__workbook.worksheets

    @property
    def variablenames(self):
        """
        Documentation
        """
        variablenames = []
        for variable in self.__variables:
            if variable.name != 'Constanta':
                variablenames.append(variable.name)
        return variablenames

    @property
    def constantvalues(self):
        """
        Documentation
        """
        constantvalues = []
        for variable in self.__variables:
            if variable.name == 'Constanta':
                constantvalues.append(variable.value)
        return constantvalues

    @property
    def mergedcells(self):
        """
        Documentation
        """
        ranges = []
        for range in self.__mergedcells:
            ranges.append(range.name)
        return ranges

    def remerge(self, sizetables):
        """
        Documentation
        """
        for sizetable in sizetables:
            for range in self.__mergedcells:
                incell = [range.name.split(':')[0],range.name.split(':')[1]]
                outcell = []
                for cell in incell:
                    cell_column = cell[0]
                    cell_row = int(cell[1:len(cell)])
                    if cell_row > sizetable[0]:
                        cell_row = cell_row + sizetable[1]
                    outcell.append(cell_column+str(cell_row))
                range.name = outcell[0]+':'+str(outcell[1])
                range.value = outcell[0]+':'+str(outcell[1])
        return None

    @property
    def variables(self):
        """
        Documentation
        """
        return self.__variables


    def variable(self, name):
        """
        Documentation
        """
        result = None
        for variable in self.__variables:
            if variable.name == name:
                result = variable
                break
        return result

    @property
    def styles(self):
        """
        Documentation
        """
        return self.__styles

    @property
    def columns_width(self):
        """
        Documentation
        """
        columns_width = []
        for column in self.__columns_width:
            columns_width.append([column.name, column.value])
        return columns_width

    @property
    def rows_height(self):
        """
        Documentation
        """
        rows_height = []
        for row in self.__rows_height:
            rows_height.append([row.name, row.value])
        return rows_height

    def sizetables(self, data):
        """
        Documentation
        """
        dict = {}
        positions = {}
        tables = {}
        for name in self.variablenames:
            if 'tbl' in name:
                positions.update({name:self.variable(name).row})

        for name in self.variablenames:
            if 'tbl' in name:
                n = 0
                for tbl in data['table']:
                    for key in tbl:
                        if name == key:
                            n += 1
                dict.update({name:[positions[name],n]})

        for key in dict:
            s = str(dict[key][0])+':'+str(dict[key][1])
            tables.update({s:key})
        result = []
        for key in tables:
            result.append([int(key.split(':')[0]),int(key.split(':')[1])])
        return result

    def update(self, sizetables):
        """
        Documentation
        """
        self.remerge(sizetables)
        for sizetable in sizetables:
            self.rows = self.rows + sizetable[1]
            for variable in self.variables:
                if variable.row > sizetable[0]:
                    variable.row = variable.row + sizetable[1]
            for style in self.styles:
                if style.row > sizetable[0]:
                    style.row = style.row + sizetable[1]
        styles = self.styles
        for sizetable in sizetables:
            for style in styles:
                if style.row == sizetable[0]:
                    for n in range(sizetable[1]-1):
                        self.__styles.append(Style(style.name,style.row+n+1,style.column,style.data_type, style.font, style.alignment,style.fill, style.border))
        for style in self.styles:
            style.rename()
        return None

    def import_data(self, data):
        """
        Documentation
        """

        for variable in self.variables:
            if variable.name in data:
                variable.value = data[variable.name]
            elif 'table' in data:
                for table in data['table']:
                    if variable.name in table and variable.value == None:
                        variable.value = table[variable.name]
                        table.popitem()
        variables = self.variables
        for variable in variables:
            if 'table' in data:
                n=1
                for table in data['table']:
                    if variable.name in table:
                        n=n+1
                        self.__variables.append(Variable(variable.name+str(n),variable.row+n,variable.column,table[variable.name]))
        return None

    def past_in_worksheet(self,worksheet):
        """
        Documentation
        """
        for column in self.columns_width:
            worksheet.column_dimensions[column[0]].width = column[1]
        for row in self.rows_height:
            worksheet.row_dimensions[row[0]].height = row[1]
        for variable in self.variables:
            worksheet.cell(variable.row, variable.column, variable.value)
        for style in self.styles:
            worksheet[style.name].data_type = copy(style.data_type)
            worksheet[style.name].font = copy(style.font)
            worksheet[style.name].fill = copy(style.fill)
            worksheet[style.name].border = copy(style.border)
            worksheet[style.name].alignment = copy(style.alignment)
        for merge in self.mergedcells:
            worksheet.merge_cells(merge)
        return None
