from divinegift import main

import os
from datetime import datetime, timedelta, date
try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    raise ImportError("openpyxl isn't installed. Run: pip install -U openpyxl")
try:
    import xlrd
except ImportError:
    raise ImportError("xlrd isn't installed. Run: pip install -U xlrd")
try:
    import xlsxwriter
except ImportError:
    raise ImportError("xlsxwriter isn't installed. Run: pip install -U xlsxwriter")
try:
    import xlwt
except ImportError:
    raise ImportError("xlwt isn't installed. Run: pip install -U xlwt")


def create_excel(list_: list, fname: str, excel_header: dict = None, worksheet: str = 'Sheet', start_row: int = 0):
    """
    Создает Excel-файл в формате xlsx
    :param list_: Входной список
    :param fname: Имя файла с путем
    :param excel_header: Словарь заголовков Excel с шириной столбцов
    :param worksheet: Имя вкладки
    :return:
    """
    # Если указанной папки не существует - создаем ее (включая родительские, если нужно)
    fd, fn = os.path.split(fname)
    if fd and not os.path.exists(fd):
        os.makedirs(fd)
    # Создаем книгу
    if fname.split('.')[-1] == 'xlsx':
        wb = xlsxwriter.Workbook(fname)
        ws = wb.add_worksheet(worksheet)
        bold = wb.add_format({'bold': 1})
        bold.set_bottom()
        int_f = wb.add_format({'num_format': 'General'})
        float_f = wb.add_format({'num_format': 'General'})
        date_f = wb.add_format({'num_format': 'dd.mm.yyyy'})
        datetime_f = wb.add_format({'num_format': 'dd.mm.yyyy hh:mm:ss'})
    else:
        wb = xlwt.Workbook()
        ws = wb.add_sheet(worksheet)
        int_f = xlwt.easyxf(num_format_str='General')
        float_f = xlwt.easyxf(num_format_str='General')
        date_f = xlwt.easyxf(num_format_str='dd.mm.yyyy')
        datetime_f = xlwt.easyxf(num_format_str='dd.mm.yyyy hh:mm:ss')
    # Получаем список заголовков в экселе из настроек
    if excel_header:
        header = list(excel_header.keys())
        # Получаем ширину столбцов из настроек
        try:
            column_width = [x[0] for x in list(excel_header.values())]
        except TypeError:
            column_width = list(excel_header.values())
        # Заполняем шапку
        for c, v in enumerate(header):
            try:
                ws.write(0, c, v, bold)
            except:
                ws.write(0, c, v)
            ws.set_column(c, c, column_width[c])
        start_row += 1
    # Вставляем значения
    for rn, row in enumerate(list_):
        try:
            col = list(row.values())
        except AttributeError:
            col = row
        for c, v in enumerate(col):
            # Если значение типа datetime - преобразовываем в строку вида dd.MM.yyyy HH:mm:ss
            if isinstance(v, datetime):
                st = datetime_f
                # v = datetime.strftime(v, '%d.%m.%Y %H:%M:%S')
            # Если значение типа date - преобразовываем в строку вида dd.MM.yyyy
            elif isinstance(v, date):
                st = date_f
                # v = date.strftime(v, '%d.%m.%Y')
            elif isinstance(v, int):
                st = int_f
            elif isinstance(v, float):
                st = float_f
            else:
                st = None
            if st is not None:
                ws.write(rn + start_row, c, v, st)
            else:
                ws.write(rn + start_row, c, v)
            # ws.write(rn + start_row, c, str(v) if v else '')
    try:
        wb.close()
    except:
        wb.save(fname)


def read_excel(filename: str, excel_header, sheet_name: str = None, 
               int_columns: list = None, date_columns: list = None, start_row: int = 1):
    # Считываем Excel-файл
    excel_arr = []
    if isinstance(excel_header, list):
        columns = excel_header
    else:
        columns = [x[1] for x in list(excel_header.values())]

    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name] if sheet_name else wb.active
        records = [[x.value for x in row] for row in ws.iter_rows(min_row=start_row)]
    else:
        wb = xlrd.open_workbook(filename)
        ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
        records = [ws.row_values(i) for i in range(start_row - 1, ws.nrows)]

    int_columns = int_columns if int_columns else []
    date_columns = date_columns if date_columns else []

    # По каждой строчке
    for i, r in enumerate(records):
        # Сцепляем заголовки (английские) из настроек с данными из Excel-файла
        tmp_row = dict(zip(columns, r))

        # Костыль от преобразований данных экселем после ручных правок
        for c in columns:
            # Костыль для столбцов, которые должны быть int
            if tmp_row.get(c) is not None:
                if c in int_columns and not isinstance(tmp_row.get(c), int):
                    tmp_row[c] = int(tmp_row[c])
                # Костыль для столбцов, которые должны быть datetime
                if c in date_columns and not isinstance(tmp_row.get(c), datetime):
                    try:
                        tmp_row[c] = main.parse_date(tmp_row[c])
                    except TypeError:
                        # Костыль, если дата в экселе не текст, а число
                        tmp_row[c] = datetime(1899, 12, 31) + timedelta(days=tmp_row[c] - 1)
                    except ValueError:
                        tmp_row[c] = None
                # Костыль для столбцов, которые должны быть str
                if c not in date_columns and c not in int_columns and not isinstance(tmp_row.get(c), str):
                    tmp_row[c] = str(tmp_row.get(c))
        excel_arr.append(tmp_row)

    return excel_arr    


if __name__ == '__main__':
    pass
