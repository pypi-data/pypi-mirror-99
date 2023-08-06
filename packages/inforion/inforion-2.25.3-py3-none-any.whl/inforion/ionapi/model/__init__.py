import json
import logging
import time

import numpy as np
import pandas as pd
import progressbar
import requests
import xlsxwriter
from oauthlib.oauth2 import BackendApplicationClient
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series
from pandas import compat
from requests.adapters import HTTPAdapter
from requests.auth import HTTPBasicAuth
from requests_oauthlib import OAuth2Session

# from requests.packages.urllib3.util.retry import Retry
from urllib3.util import Retry

import inforion.helper.api_fields_helper as apifieldhelp
import inforion.helper.filehandling as filehandling
import inforion.ionapi.controller as controller

# import grequests

# from logger import get_logger
# import sendresults, saveresults
# from inforion.ionapi.model import

DEFAULT_TIMEOUT = 50  # seconds
MaxChunk = 100


def execute(
    url,
    headers,
    program,
    methode,
    dataframe,
    outputfile=None,
    start=0,
    end=None,
    on_progress=None,
):

    df = dataframe

    df = df.replace(np.nan, "", regex=True)
    df = df.astype(str)

    # Coverting data type
    # numeric_fields = apifieldhelp.get_numeric_fields_list_from_db(program)
    # for field in numeric_fields:
    #    if field in df:
    #        try:
    #            df[field] = pd.to_numeric(df[field])
    #        except Exception as ex:
    #            raise Exception("Error converting field '{0}' to numeric".format(field))
    # df = df.replace(np.nan, "", regex=True)

    data = {"program": program, "cono": 409}
    data1 = {}
    a = []

    chunk = MaxChunk
    if end is not None:
        df = df[start:end].copy(deep=False)
        df = df.reset_index(drop=True)
        # print (df.head(10))

    # else:
    total_rows = df.shape[0]
    total_rows = int(total_rows)

    methode = methode.split(",")
    methode_count = len(methode)

    logging.info("Number of rows " + str(total_rows))

    with progressbar.ProgressBar(max_value=total_rows) as bar:
        for index, row in df.iterrows():

            bar.update(index)

            if on_progress:
                on_progress(total_rows, index + 1)

            row = row.to_json()
            row = json.loads(row)

            for i in methode:
                data1["transaction"] = i
                data1["record"] = row
                a.append(data1.copy())

            if chunk == 0:
                data["transactions"] = a

                r = controller.sendresults(url, headers, data)
                df, data, chunk = controller.saveresults(
                    r, df, program, index, chunk, MaxChunk, methode_count
                )
                data1 = {}
                a = []

            else:
                chunk = chunk - 1

        data["transactions"] = a

        r = controller.sendresults(url, headers, data)
        index = index + 1
        df, data, chunk = controller.saveresults(
            r, df, methode, index, chunk, MaxChunk, methode_count
        )

    # df = df.replace(np.nan, '', regex=True)
    # df = df.astype(str)

    if outputfile is not None:
        print("Save to file: " + outputfile)
        filehandling.savetodisk(outputfile, df)

        df_results = getSuccessGraphDataframe(df)
        createGraph(outputfile, df_results)

    return df


def executeSnd(
    url, headers, program, methode, dataframe, outputfile=None, start=0, end=None
):

    df = dataframe

    data = {"program": program, "cono": 409}

    methode = methode.split(",")
    len(methode)
    data1 = {}

    chunk = MaxChunk
    if end is not None:
        df = df[start:end].copy(deep=False)
        df = df.reset_index(drop=True)

    # else:
    total_rows = df.shape[0]
    total_rows = int(total_rows)

    logging.info("Number of rows " + str(total_rows))

    a = []
    with progressbar.ProgressBar(max_value=total_rows) as bar:
        for index, row in df.iterrows():

            bar.update(index)

            row = row.to_json()
            row = json.loads(row)

            for i in methode:
                data1["transaction"] = i
                data1["record"] = row
                a.append(data1.copy())

        data["transactions"] = a

        index = index + 1

    print(data)

    r = controller.sendresults(url, headers, data)

    df, data, chunk = controller.saveresults(r, df, methode, index, chunk)

    df = df.replace(np.nan, "", regex=True)
    df = df.astype(str)

    if outputfile is not None:
        logging.info("Save to file: " + outputfile)
        filehandling.savetodisk(outputfile, df)

    return df

    logging.info("Still in Beta")

    df = dataframe

    data = {"program": program, "cono": 409}

    methode = methode.split(",")
    len(methode)
    data1 = {}
    if end is not None:
        df = df[start:end].copy(deep=False)
        df = df.reset_index(drop=True)
        # print (df.head(10))

    # else:
    total_rows = df.shape[0]
    total_rows = int(total_rows)

    logging.info("Number of rows " + str(total_rows))

    a = []
    with progressbar.ProgressBar(max_value=total_rows) as bar:
        for index, row in df.iterrows():

            bar.update(index)

            row = row.to_json()
            row = json.loads(row)

            for i in methode:
                data1["transaction"] = i
                data1["record"] = row
                a.append(data1.copy())

        data["transactions"] = a

        print(data)

        r = controller.sendresults(url, headers, data, stream=True)
        index = index + 1
        # df,data,chunk = saveresults(r,df,methode,index,chunk)

    logging.info(r)

    df = df.replace(np.nan, "", regex=True)
    df = df.astype(str)

    if outputfile is not None:
        print("Save to file: " + outputfile)
        filehandling.savetodisk(outputfile, df)

    return df


def getSuccessGraphDataframe(df):
    results = {}
    for _, row in df.iterrows():
        msgs = row["MESSAGE"].split("^_^")
        for msg in msgs:
            if msg:
                
                if ":" in msg:
                    key = msg[0 : msg.index(":")]
                    val = msg[msg.index(":") + 1 :]
                else:
                    key = 'Unknown'
                    val = msg

                if key not in results:
                    results[key] = {}
                    results[key]["success"] = 0
                    results[key]["fail"] = 0

                if "OK" in val or "ist bereits vorhanden" in val:
                    results[key]["success"] = results[key]["success"] + 1
                else:
                    results[key]["fail"] = results[key]["fail"] + 1
            

    success = []
    fail = []
    prog = []
    for key in results:
        prog.append(key)
        success.append(results[key]["success"])
        fail.append(results[key]["fail"])

    df = pd.DataFrame({"success": success, "fail": fail}, index=prog)
    return df


def createGraph(excel_file, df):
    sheet_name = "Data"
    sheet_name_graphs = "Graphs"

    writer = pd.ExcelWriter(excel_file, engine="openpyxl")
    book = load_workbook(excel_file)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    ws_graph = writer.book.create_sheet(sheet_name_graphs)
    df.to_excel(writer, sheet_name=sheet_name)

    ws = writer.sheets[sheet_name]

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "ETL Results"
    chart1.y_axis.title = "Count"
    chart1.x_axis.title = "Programm"

    data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.height = 12
    chart1.width = 20

    ws_graph.add_chart(chart1, "A10")
    writer.save()
