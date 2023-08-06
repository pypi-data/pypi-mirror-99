#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author : mocobk
# @Email : mailmzb@qq.com
# @Time : 2021/2/19 18:20
import json
import pickle
import time
from collections import OrderedDict
from pathlib import Path
from urllib.parse import urljoin

import click
import requests
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.reader import excel
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

excel.warnings.simplefilter('ignore')

ROOT_DIR = Path(__file__).parent
TEMPLATE_XLS = ROOT_DIR.joinpath('template', 'case.xlsm')
TAPD_BASE_URL = 'http://9.134.187.9'

BVT_TASK_TEMPLATE = ROOT_DIR.joinpath('template', 'bvt_task.html')


def secure_filename(filename: str):
    chars_map = {'\\': '_',
                 '/': '_',
                 ':': '_',
                 '*': '_',
                 '?': '',
                 '"': '',
                 '<': '_',
                 '>': '_',
                 '|': '_'}

    trans_table = filename.maketrans(chars_map)
    return filename.translate(trans_table)


class Config:
    CONFIG_FILE = ROOT_DIR.joinpath('config.pkl')

    def get(self):
        if self.CONFIG_FILE.exists():
            return pickle.load(self.CONFIG_FILE.open('rb'))
        else:
            return {}

    def put(self, config):
        pickle.dump(config, self.CONFIG_FILE.open('wb'))


class ApiException(Exception):
    pass


class Tapd:
    def __init__(self, iteration_id=None, base_url=None):
        self.base_url = base_url or TAPD_BASE_URL
        self.iteration_id = iteration_id

    def set_iteration_id(self, iteration_id):
        self.iteration_id = iteration_id

    def get_tapd_iterations(self):
        url = urljoin(self.base_url, '/tapd/origin/iterations')
        response = requests.get(url, params={'status': 'open', 'fields': 'id,name'})
        data = [(item['Iteration']['name'], item['Iteration']['id']) for item in response.json().get('data', [])]
        return OrderedDict(data[:8])

    def get_tapd_iteration_developers(self):
        url = urljoin(self.base_url, '/tapd/origin/tasks')
        response = requests.get(
            url, params={'iteration_id': self.iteration_id, 'fields': 'owner', 'order': 'owner asc', 'limit': 200}
        )
        data = [item['Task']['owner'] for item in response.json().get('data', []) if item['Task']['owner']]
        return set(data)

    def get_tapd_iteration_stories(self):
        url = urljoin(self.base_url, '/tapd/origin/stories')
        response = requests.get(url, params={'iteration_id': self.iteration_id, 'fields': 'id,name'})
        data = [(item['Story']['name'], item['Story']['id']) for item in response.json().get('data', [])]
        return OrderedDict(data)

    def get_tapd_case_categories(self):
        url = urljoin(self.base_url, '/tapd/origin/tcase_categories')
        response = requests.get(url, params={'parent_id': 0, 'fields': 'id,name', 'order': 'name asc'})
        data = [(item['TcaseCategory']['name'], item['TcaseCategory']['id']) for item in
                response.json().get('data', [])]
        return OrderedDict(data)

    def get_tapd_test_plans(self):
        """custom_field_2 为迭代 ID"""
        url = urljoin(self.base_url, '/tapd/origin/test_plans')
        response = requests.get(url,
                                params={'status': 'open', 'custom_field_2': self.iteration_id, 'fields': 'id,name'})
        data = [(item['TestPlan']['name'], item['TestPlan']['id']) for item in
                response.json().get('data', [])]
        return OrderedDict(data)

    def post_tapd_task(self, **kwargs):
        """
        创建或更新任务
        name, description, creator, owner, story_id, iteration_id
        """
        url = urljoin(self.base_url, '/tapd/origin/tasks')
        kwargs.setdefault('iteration_id', self.iteration_id)
        kwargs.setdefault('priority', 4)
        kwargs.setdefault('custom_field_one', '冒烟')
        response = requests.post(url, data=kwargs)
        json_data = response.json()
        if not json_data.get('data'):
            raise ApiException('用例上传失败, 原因：', json_data.get('info', 'no data'))

        task_id = json_data.get('data', {}).get('Task', {}).get('id')
        return task_id

    def post_tapd_case(self, **kwargs):
        """
        创建或更新用例
        id, category_id, name, precondition, steps, expectation, priority, creator, type
        custom_field_1 一级模块
        custom_field_2 二级模块
        custom_field_5 迭代名称
        """
        url = urljoin(self.base_url, '/tapd/origin/tcases')
        kwargs.setdefault('priority', '中')
        kwargs.setdefault('type', '功能测试')
        response = requests.post(url, data=kwargs)
        json_data = response.json()
        if not json_data.get('data'):
            raise ApiException('用例上传失败, 原因：', json_data.get('info', 'no data'))

        case_id = json_data.get('data', {}).get('Tcase', {}).get('id')
        return case_id

    def post_tapd_test_plans_and_tcase_relation(self, **kwargs):
        """
        创建测试计划和测试用例关联关系
        teet_plan_id, tcase_ids, creator
        """
        url = urljoin(self.base_url, '/tapd/origin/test_plans/create_tcase_relation')
        response = requests.post(url, data=kwargs)
        json_data = response.json()
        if json_data.get('status') != 1:
            raise ApiException('创建测试计划和测试用例关联关系失败, 原因：', json_data.get('info', 'status != 1'))


class Excel:
    def __init__(self, filename, tapd: Tapd = None):
        self.filename = filename
        self.wb = load_workbook(filename, keep_vba=True)
        self.case_ws = self.wb.worksheets[0]
        self.validate_ws = None
        self.case_title_column_letters = self.get_title_column_letter_map(self.case_ws, 2)
        self.tapd = tapd

    @staticmethod
    def get_title_column_letter_map(sheet: Worksheet, title_row=1):
        column_letter_map = {}
        for (cell,) in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=title_row, max_row=title_row):
            if cell.value is None:
                continue
            column_letter_map[cell.value] = cell.column_letter
        return column_letter_map

    def update_case_title_column_letters(self):
        self.case_title_column_letters = self.get_title_column_letter_map(self.case_ws, 2)

    def set_sheet_name(self, name):
        self.case_ws.title = name

    def set_iteration_meta_info(self, data: dict):
        """在用例表中存储额外的元信息, 本想放在sheet表标题的，但标题有长度限制 31"""
        self.case_ws['A1'] = json.dumps(data, ensure_ascii=False)

    def get_iteration_meta_info(self):
        """在用例表中存储额外的元信息, 本想放在sheet表标题的，但标题有长度限制 31"""
        meta_info = self.case_ws['A1'].value
        try:
            return json.loads(meta_info)
        except (json.JSONDecodeError, TypeError):
            return None

    @staticmethod
    def append_cols(sheet, column_data):
        column_index = cur_max_column = sheet.max_column
        if sheet.cell(1, cur_max_column).value is not None:
            column_index += 1
        for index, value in enumerate(column_data, start=1):
            sheet.cell(index, column_index, value)

    def set_data_validation(self, cell, formula_data, show_error_message=True):
        """
        :param cell: 'A1:A10'
        :param formula_data: ['高', '中', '低']
        :param show_error_message: 填错是否弹窗提示
        :return:
        """
        self.append_cols(self.validate_ws, formula_data)
        column_letter = self.validate_ws.cell(1, self.validate_ws.max_column).column_letter
        dv = DataValidation(
            type="list",
            formula1=f"{quote_sheetname(self.validate_ws.title)}!${column_letter}$2:${column_letter}${len(formula_data)}",
            allow_blank=True,
            showErrorMessage=show_error_message)
        dv.error = '请使用下拉选择可用的值'
        dv.errorTitle = '输入的值有误'
        dv.add(cell)
        self.case_ws.add_data_validation(dv)

    def _set_categories_data_validation(self):
        """设置 一级模块（用例目录）的数据验证"""
        column_letter = self.case_title_column_letters['一级模块']
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('一级模块', *self.tapd.get_tapd_case_categories().keys()))

    def _set_developer_data_validation(self):
        """设置 开发人员 的数据验证"""
        column_letter = self.case_title_column_letters['开发人员']
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('开发人员', *self.tapd.get_tapd_iteration_developers()), show_error_message=False)

    def _set_stories_data_validation(self):
        """设置 需求 的数据验证"""
        column_letter = self.case_title_column_letters['需求']
        cell = f'{column_letter}3:{column_letter}1048576'
        self.set_data_validation(cell, ('需求', *self.tapd.get_tapd_iteration_stories().keys()))
        self.append_cols(self.validate_ws, ('需求ID', *self.tapd.get_tapd_iteration_stories().values()))

    def set_iteration_data_validation(self):
        data_validation_sheet_name = '数据验证（勿删）'
        if data_validation_sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[data_validation_sheet_name])

        self.validate_ws = self.wb.create_sheet('数据验证（勿删）')
        self._set_categories_data_validation()
        self._set_developer_data_validation()
        self._set_stories_data_validation()

    def save(self, filename):
        self.wb.save(filename)


class Cases:
    def __init__(self, excel: Excel):
        self.excel = excel
        self.case_ws = self.excel.case_ws
        self.iteration_meta_info = self.excel.get_iteration_meta_info()
        self.title_column_letters = self.excel.case_title_column_letters

    def is_valid_case(self, *required_cell):
        """
        判断是否是有效的用例
        功能用例：`一级模块`	`用例名称`
        冒烟用例：`一级模块`	`用例名称`	`用例等级（必须是 高）`	`开发人员`	`需求`
        """
        for cell in required_cell:
            if not self.case_ws[cell].value:
                return False
        else:
            return True

    def validate_data(self):
        """
        功能用例：`一级模块`	`用例名称` 必填
        """
        results = []
        cases_count = 0
        for row_index in range(3, self.case_ws.max_row + 1):
            title_cell = f'{self.title_column_letters["用例名称"]}{row_index}'
            custom_field_1_cell = f'{self.title_column_letters["一级模块"]}{row_index}'
            if self.case_ws[title_cell].value:
                if not self.is_valid_case(custom_field_1_cell):
                    results.append(f'第 {row_index} 行用例【一级模块】/【用例名称】不能为空')
                else:
                    cases_count += 1
        return results, cases_count

    def upload(self, creator):
        # 如果 case_id 列不存在（用例未上传过），则插入 case_id 列至最后，否则代表用例有上传过，不处理
        if 'case_id' not in self.title_column_letters:
            self.title_column_letters['case_id'] = self.case_ws.cell(2, self.case_ws.max_column + 1,
                                                                     'case_id').column_letter

        iteration_name = self.iteration_meta_info.get('iteration_name')
        categories_map = self.excel.tapd.get_tapd_case_categories()
        test_plans_map = self.excel.tapd.get_tapd_test_plans()

        def html_format(content: str):
            return '<br />'.join(content.strip().splitlines())

        # 使用进度条
        with click.progressbar(
                iterable=range(3, self.case_ws.max_row + 1),  # 遍历行号
                label="上传进度",
                bar_template="%(label)s | %(bar)s | %(info)s",
                fill_char=click.style("█", fg="green"),
                empty_char=" ",
        ) as bar:
            time.sleep(0.5)  # 强制先暂停 0.5s, 避免太快进度条没刷出来就结束了，体验提升 50%
            for row_index in bar:
                """
                id, category_id, name, precondition, steps, expectation, priority, creator, type
                custom_field_1 一级模块
                custom_field_2 二级模块
                custom_field_5 迭代名称
                """
                custom_field_1_cell = f'{self.title_column_letters["一级模块"]}{row_index}'
                custom_field_2_cell = f'{self.title_column_letters["二级模块"]}{row_index}'
                title_cell = f'{self.title_column_letters["用例名称"]}{row_index}'
                precondition_cell = f'{self.title_column_letters["前置条件"]}{row_index}'
                steps_cell = f'{self.title_column_letters["用例步骤"]}{row_index}'
                expectation_cell = f'{self.title_column_letters["预期结果"]}{row_index}'
                priority_cell = f'{self.title_column_letters["用例等级"]}{row_index}'
                story_cell = f'{self.title_column_letters["需求"]}{row_index}'
                case_id_cell = f'{self.title_column_letters["case_id"]}{row_index}'

                # 只上传合法的用例
                if self.is_valid_case(custom_field_1_cell, title_cell):
                    case_id = self.excel.tapd.post_tapd_case(
                        id=self.case_ws[case_id_cell].value or None,
                        category_id=categories_map.get(self.case_ws[custom_field_1_cell].value, 'None Category'),
                        name=self.case_ws[title_cell].value,
                        precondition=html_format(self.case_ws[precondition_cell].value or ''),
                        steps=html_format(self.case_ws[steps_cell].value or ''),
                        expectation=html_format(self.case_ws[expectation_cell].value or ''),
                        priority=html_format(self.case_ws[priority_cell].value or ''),
                        creator=creator,
                        custom_field_1=self.case_ws[custom_field_1_cell].value or '',
                        custom_field_2=self.case_ws[custom_field_2_cell].value or '',
                        custom_field_5=iteration_name,
                    )

                    if case_id:
                        # 创建测试计划和测试用例关联关系, 已关联过的用例不重复关联
                        test_plan_id = test_plans_map.get(self.case_ws[story_cell].value)
                        if test_plan_id and not self.case_ws[case_id_cell].value:
                            self.excel.tapd.post_tapd_test_plans_and_tcase_relation(
                                test_plan_id=test_plan_id,
                                tcase_ids=case_id,
                                creator=creator,
                            )

                        # 写入用例 ID
                        self.case_ws[case_id_cell].value = case_id


class BVTCases(Cases):
    def validate_data(self):
        """
        冒烟用例：`一级模块`	`用例名称`	`用例等级（必须是 高）`	`开发人员`	`需求`
        """
        results = []
        cases_count = 0
        for row_index in range(3, self.case_ws.max_row + 1):
            priority_cell = f'{self.title_column_letters["用例等级"]}{row_index}'
            title_cell = f'{self.title_column_letters["用例名称"]}{row_index}'
            if self.case_ws[title_cell].value and self.case_ws[priority_cell].value == '高':  # 冒烟用例
                custom_field_1_cell = f'{self.title_column_letters["一级模块"]}{row_index}'
                developer_cell = f'{self.title_column_letters["开发人员"]}{row_index}'
                story_cell = f'{self.title_column_letters["需求"]}{row_index}'
                if not self.is_valid_case(custom_field_1_cell, developer_cell, story_cell):
                    results.append(f'第 {row_index} 行冒烟用例【一级模块】/【用例名称】/【开发人员】/【需求】不能为空')
                else:
                    cases_count += 1

        return results, cases_count

    def upload(self, creator):
        # 如果 task_id 列不存在（用例未上传过），则插入 task_id 列至最后，否则代表用例有上传过，不处理
        if 'task_id' not in self.title_column_letters:
            self.title_column_letters['task_id'] = self.case_ws.cell(2, self.case_ws.max_column + 1,
                                                                     'task_id').column_letter

        template = Template(BVT_TASK_TEMPLATE.read_text(encoding='utf-8'))
        stories_map = self.iteration_meta_info.get('iteration_stories', {})

        # 使用进度条
        with click.progressbar(
                iterable=range(3, self.case_ws.max_row + 1),  # 遍历行号
                label="上传进度",
                bar_template="%(label)s | %(bar)s | %(info)s",
                fill_char=click.style("█", fg="green"),
                empty_char=" ",
        ) as bar:
            time.sleep(0.5)  # 强制先暂停 0.5s, 避免太快进度条没刷出来就结束了，体验提升 50%
            for row_index in bar:
                priority_cell = f'{self.title_column_letters["用例等级"]}{row_index}'
                developer_cell = f'{self.title_column_letters["开发人员"]}{row_index}'
                if self.case_ws[developer_cell].value and self.case_ws[priority_cell].value == '高':
                    custom_field_1_cell = f'{self.title_column_letters["一级模块"]}{row_index}'
                    custom_field_2_cell = f'{self.title_column_letters["二级模块"]}{row_index}'
                    title_cell = f'{self.title_column_letters["用例名称"]}{row_index}'
                    precondition_cell = f'{self.title_column_letters["前置条件"]}{row_index}'
                    steps_cell = f'{self.title_column_letters["用例步骤"]}{row_index}'
                    expectation_cell = f'{self.title_column_letters["预期结果"]}{row_index}'
                    story_cell = f'{self.title_column_letters["需求"]}{row_index}'
                    task_id_cell = f'{self.title_column_letters["task_id"]}{row_index}'

                    # 只上传合法的冒烟用例
                    if self.is_valid_case(custom_field_1_cell, title_cell, developer_cell, story_cell):
                        task_id = self.excel.tapd.post_tapd_task(
                            id=self.case_ws[task_id_cell].value or None,
                            name=f'【冒烟用例】{self.case_ws[title_cell].value}',
                            creator=creator,
                            owner=self.case_ws[developer_cell].value,
                            story_id=stories_map.get(self.case_ws[story_cell].value),
                            iteration_id=self.iteration_meta_info['iteration_id'],
                            description=template.render(
                                custom_field_1=self.case_ws[custom_field_1_cell].value,
                                custom_field_2=self.case_ws[custom_field_2_cell].value or '',
                                precondition=self.case_ws[precondition_cell].value or '',
                                case_steps=self.case_ws[steps_cell].value or '',
                                expectation=self.case_ws[expectation_cell].value or ''
                            )
                        )
                        if task_id:
                            self.case_ws[task_id_cell].value = task_id
