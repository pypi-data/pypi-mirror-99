import os
import numpy as np
import datetime
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, asdict
from tqdm import trange
import sys

from openpyxl.cell.cell import Cell
from openpyxl.styles import Font

from steam_nb_api.resources.ResourceReader import ResourceReader

@dataclass
class LEDETInputs:
    T00: float = 0.0
    l_magnet: float = 0.0
    I00: float = 0.0
    GroupToCoilSection: np.ndarray = np.array([])
    polarities_inGroup: np.ndarray = np.array([])
    nT: np.ndarray = np.array([])
    nStrands_inGroup: np.ndarray = np.array([])
    l_mag_inGroup: np.ndarray = np.array([])
    ds_inGroup: np.ndarray = np.array([])
    f_SC_strand_inGroup: np.ndarray = np.array([])
    f_ro_eff_inGroup: np.ndarray = np.array([])
    Lp_f_inGroup: np.ndarray = np.array([])
    RRR_Cu_inGroup: np.ndarray = np.array([])
    SCtype_inGroup: np.ndarray = np.array([])
    STtype_inGroup: np.ndarray = np.array([])
    insulationType_inGroup: np.ndarray = np.array([])
    internalVoidsType_inGroup: np.ndarray = np.array([])
    externalVoidsType_inGroup: np.ndarray = np.array([])
    wBare_inGroup: np.ndarray = np.array([])
    hBare_inGroup: np.ndarray = np.array([])
    wIns_inGroup: np.ndarray = np.array([])
    hIns_inGroup: np.ndarray = np.array([])
    Lp_s_inGroup: np.ndarray = np.array([])
    R_c_inGroup: np.ndarray = np.array([])
    Tc0_NbTi_ht_inGroup: np.ndarray = np.array([])
    Bc2_NbTi_ht_inGroup: np.ndarray = np.array([])
    c1_Ic_NbTi_inGroup: np.ndarray = np.array([])
    c2_Ic_NbTi_inGroup: np.ndarray = np.array([])
    Tc0_Nb3Sn_inGroup: np.ndarray = np.array([])
    Bc2_Nb3Sn_inGroup: np.ndarray = np.array([])
    Jc_Nb3Sn0_inGroup: np.ndarray = np.array([])
    df_inGroup: np.ndarray = np.array([0])
    selectedFit_inGroup: np.ndarray = np.array([0])
    fitParameters_inGroup: np.ndarray = np.array([0])
    overwrite_f_internalVoids_inGroup: np.ndarray = np.array([])
    overwrite_f_externalVoids_inGroup: np.ndarray = np.array([])
    alphasDEG: np.ndarray = np.array([])
    rotation_block: np.ndarray = np.array([])
    mirror_block: np.ndarray = np.array([])
    mirrorY_block: np.ndarray = np.array([])
    el_order_half_turns: np.ndarray = np.array([])
    iContactAlongWidth_From: np.ndarray = np.array([])
    iContactAlongWidth_To: np.ndarray = np.array([])
    iContactAlongHeight_From: np.ndarray = np.array([])
    iContactAlongHeight_To: np.ndarray = np.array([])
    t_PC: float = 0.0
    t_PC_LUT: np.ndarray = np.array([])
    I_PC_LUT: np.ndarray = np.array([])
    R_circuit: float = 0.0
    R_crowbar: float = 0.0
    Ud_crowbar: float = 0.0
    tEE: float = 9999
    R_EE_triggered: float = 0.0
    tCLIQ: np.ndarray = np.array([9999])
    directionCurrentCLIQ: np.ndarray = np.array([0])
    nCLIQ: np.ndarray = np.array([0])
    U0: np.ndarray = np.array([0])
    C: np.ndarray = np.array([0])
    Rcapa: np.ndarray = np.array([0])
    tQH: np.ndarray = np.array([9999])
    U0_QH: np.ndarray = np.array([0])
    C_QH: np.ndarray = np.array([0])
    R_warm_QH: np.ndarray = np.array([0])
    w_QH: np.ndarray = np.array([0])
    h_QH: np.ndarray = np.array([0])
    s_ins_QH: np.ndarray = np.array([0])
    type_ins_QH: np.ndarray = np.array([0])
    s_ins_QH_He: np.ndarray = np.array([0])
    type_ins_QH_He: np.ndarray = np.array([0])
    l_QH: np.ndarray = np.array([0])
    f_QH: np.ndarray = np.array([0])
    iQH_toHalfTurn_From: np.ndarray = np.array([0])
    iQH_toHalfTurn_To: np.ndarray = np.array([0])
    tQuench: np.ndarray = np.array([])
    initialQuenchTemp: np.ndarray = np.array([])
    iStartQuench: np.ndarray = np.array([1])
    tStartQuench: np.ndarray = np.array([9999])
    lengthHotSpot_iStartQuench: np.ndarray = np.array([0])
    vQ_iStartQuench: np.ndarray = np.array([0])
    sim3D_uThreshold: float = 0.0
    sim3D_f_cooling_down: float = 0.0
    sim3D_f_cooling_up: float = 0.0
    sim3D_f_cooling_left: float = 0.0
    sim3D_f_cooling_right: float = 0.0
    sim3D_fExToIns: float = 0.0
    sim3D_fExUD: float = 0.0
    sim3D_fExLR: float = 0.0
    sim3D_min_ds_coarse: float = 0.0
    sim3D_min_ds_fine: float = 0.0
    sim3D_min_nodesPerStraightPart: float = 0.0
    sim3D_min_nodesPerEndsPart: float = 0.0
    sim3D_idxFinerMeshHalfTurn: float = 0.0
    sim3D_Tpulse_sPosition: float = 0.0
    sim3D_Tpulse_peakT: float = 0.0
    sim3D_Tpulse_width: float = 0.0
    sim3D_durationGIF: float = 0.0
    sim3D_flag_saveFigures: float = 0.0
    sim3D_flag_saveGIF: float = 0.0
    sim3D_flag_VisualizeGeometry3D: float = 0.0
    sim3D_flag_SaveGeometry3D: float = 0.0
    M_m: np.ndarray = np.array([])
    fL_I: np.ndarray = np.array([])
    fL_L: np.ndarray = np.array([])
    HalfTurnToInductanceBlock: np.ndarray = np.array([])
    M_InductanceBlock_m: np.ndarray = np.array([])
@dataclass
class LEDETOptions:
    time_vector_params: np.ndarray = np.array([])
    Iref: float = 0.0
    flagIron: float = 0.0
    flagSelfField: float = 0.0
    headerLines: float = 0.0
    columnsXY: np.ndarray = np.array([])
    columnsBxBy: np.ndarray = np.array([])
    flagPlotMTF: float = 0.0
    flag_typeWindings: float = 0.0
    flag_calculateInductanceMatrix: float = 0.0
    flag_useExternalInitialization: float = 0.0
    flag_initializeVar: float = 0.0
    flag_fastMode: float = 0.0
    flag_controlCurrent: float = 0.0
    flag_automaticRefinedTimeStepping: float = 0.0
    flag_IronSaturation: float = 0.0
    flag_InvertCurrentsAndFields: float = 0.0
    flag_ScaleDownSuperposedMagneticField: float = 0.0
    flag_HeCooling: float = 0.0
    fScaling_Pex: float = 0.0
    fScaling_Pex_AlongHeight: float = 0.0
    fScaling_MR: float = 0.0
    flag_scaleCoilResistance_StrandTwistPitch: float = 0.0
    flag_separateInsulationHeatCapacity: float = 0.0
    flag_persistentCurrents: float = 0.0
    flag_ISCL: float = 0.0
    fScaling_Mif: float = 0.0
    fScaling_Mis: float = 0.0
    flag_StopIFCCsAfterQuench: float = 0.0
    flag_StopISCCsAfterQuench: float = 0.0
    tau_increaseRif: float = 0.0
    tau_increaseRis: float = 0.0
    fScaling_RhoSS: float = 0.0
    maxVoltagePC: float = 0.0
    minCurrentDiode: float = 10
    flag_symmetricGroundingEE: float = 0.0
    flag_removeUc: float = 0.0
    BtX_background: float = 0.0
    BtY_background: float = 0.0
    flag_showFigures: float = 0.0
    flag_saveFigures: float = 0.0
    flag_saveMatFile: float = 0.0
    flag_saveTxtFiles: float = 0.0
    flag_generateReport: float = 0.0
    flag_hotSpotTemperatureInEachGroup: float = 0.0
    flag_3D: float = 0.0
@dataclass
class LEDETPlots:
    suffixPlot: str = ''
    typePlot: int = 0
    outputPlotSubfolderPlot: str = ''
    variableToPlotPlot: np.ndarray = np.array([])
    selectedStrandsPlot: np.ndarray = np.array([])
    selectedTimesPlot: np.ndarray = np.array([])
    labelColorBarPlot: np.ndarray = np.array([])
    minColorBarPlot: float = 0.0
    maxColorBarPlot: float = 0.0
    MinMaxXYPlot: np.ndarray = np.array([])
    flagSavePlot: int = 0
    flagColorPlot: int = 0
    flagInvisiblePlot: int = 0
@dataclass
class LEDETVariables:
    variableToSaveTxt: np.ndarray = np.array([])
    typeVariableToSaveTxt: np.ndarray = np.array([])
    variableToInitialize: np.ndarray = np.array([])
@dataclass
class Cable:
    A_CableInsulated: np.ndarray = np.array([])
    A_SC: np.ndarray = np.array([])
    f_SC: np.ndarray = np.array([])
    f_ST: np.ndarray = np.array([])
    SCtype: np.ndarray = np.array([])
    STtype: np.ndarray = np.array([])
    Tc0_NbTi: np.ndarray = np.array([])
    Bc20_NbTi: np.ndarray = np.array([])
    c1_Ic_NbTi: np.ndarray = np.array([])
    c2_Ic_NbTi: np.ndarray = np.array([])
    alpha_NbTi: float = 0.59
    Jc_Nb3Sn0: np.ndarray = np.array([])
    Tc0_Nb3Sn: np.ndarray = np.array([])
    Bc20_Nb3Sn: np.ndarray = np.array([])

def read_row(workSheet, Nrow, St = False):
    rowValues = np.array([])
    row = workSheet[Nrow]
    for cell in row:
        if not St:
            if isinstance(cell.value, str): continue
        rowValues = np.append(rowValues, cell.value)
    rowValues = rowValues[rowValues != None]
    return rowValues

def CompareLEDETParameters(FileA, FileB, Precision = 1E-5, showIndices = 0):
    Diff = 0

    P_a = ParametersLEDET()
    P_a.readLEDETExcel(FileA)
    P_b = ParametersLEDET()
    P_b.readLEDETExcel(FileB)

    print("Starting Comparison of A: ({}) and B: ({})".format(FileA, FileB))

    ## Check Inputs
    for attribute in P_a.Inputs.__annotations__:
        Block = 1
        CC = 0

        P_a_a = P_a.getAttribute("Inputs", attribute)
        P_b_a = P_b.getAttribute("Inputs", attribute)

        if isinstance(P_a_a, float) or isinstance(P_a_a, int):
            if P_a_a != P_b_a:
                if Block: print("Found difference in Parameter {}, A: {}, B: {}".format(attribute,P_a_a, P_b_a))
                Block = 0
                Diff = 1
        elif len(P_a_a) != len(P_b_a):
            Diff = 1
            if Block:
                Block = 0
                print('Parameter {} of A, {} has not the same length as Parameter of B, {}'.format(attribute, len(P_a_a), len(P_b_a)))
        else:
            Pos = []
            for k in range(len(P_a_a)):
                try:
                    if P_a_a[k] != P_b_a[k]:
                        if abs(P_a_a[k]-P_b_a[k])>Precision:
                            Diff = 1
                            if Block:
                                print("Found difference in Parameter {}".format(attribute))
                                Block = 0
                            Pos.append(k)
                except:
                    for j in range(P_a_a.shape[1]):
                        if P_a_a[k,j] != P_b_a[k,j]:
                            if abs(P_a_a[k,j] - P_b_a[k,j]) > Precision:
                                Diff = 1
                                if Block:
                                    print("Found difference in Parameter {}".format(attribute))
                                    Block = 0
                                Pos.append([k,j])

            if len(Pos)>0:
                if len(Pos)<10:
                    print("Different Positions: {}".format(Pos))
                else:
                    print("Many values are different (>10)")
                    if showIndices: print(Pos)


    ## Check Options
    for attribute in P_a.Options.__annotations__:
        Block = 1
        P_a_a = P_a.getAttribute("Options", attribute)
        P_b_a = P_b.getAttribute("Options", attribute)
        if isinstance(P_a_a, float) or isinstance(P_a_a, int):
            if P_a_a != P_b_a:
                if Block: print("Found difference in Parameter {}, A: {}, B: {}".format(attribute, P_a_a, P_b_a))
                Block = 0
                Diff = 1
        elif len(P_a_a) != len(P_b_a):
            Diff = 1
            if Block:
                Block = 0
                print(
                    'Parameter {} of A, {} has not the same length as Parameter of B, {}'.format(attribute, len(P_a_a),
                                                                                                 len(P_b_a)))
        else:
            Pos = []
            for k in range(len(P_a_a)):
                try:
                    if P_a_a[k] != P_b_a[k]:
                        if abs(P_a_a[k] - P_b_a[k]) > Precision:
                            Diff = 1
                            if Block:
                                print("Found difference in Parameter {}".format(attribute))
                                Block = 0
                            Pos.append(k)
                except:
                    for j in range(P_a_a.shape[1]):
                        if P_a_a[k, j] != P_b_a[k, j]:
                            if abs(P_a_a[k, j] - P_b_a[k, j]) > Precision:
                                Diff = 1
                                if Block:
                                    print("Found difference in Parameter {}".format(attribute))
                                    Block = 0
                                Pos.append([k, j])

            if len(Pos) > 0:
                if len(Pos) < 10:
                    print("Different Positions: {}".format(Pos))
                else:
                    print("Many values are different (>10)")

    if Diff==0:
        print("Files are equal.")

class ParametersLEDET:
    '''
        Class of LEDET parameters
    '''
    def setAttribute(self, LEDETclass, attribute, value):
        try:
            setattr(LEDETclass, attribute, value)
        except:
            setattr(getattr(self, LEDETclass), attribute, value)

    def getAttribute(self, LEDETclass, attribute):
        try:
            return getattr(LEDETclass, attribute)
        except:
            return getattr(getattr(self, LEDETclass), attribute)

    def fillAttribute(self, LEDETclass, attribute, value):
        imitate = self.getAttribute(LEDETclass, attribute)
        if isinstance(imitate, np.ndarray) and isinstance(value, np.ndarray):
            if imitate.shape != value.shape:
                imitate.resize(value.shape, refcheck=False)

        idx_v = np.where(value != 0)
        imitate[idx_v] = value[idx_v]
        try:
            setattr(LEDETclass, attribute, imitate)
        except:
            setattr(getattr(self, LEDETclass), attribute, imitate)

    def readLEDETExcel(self, file):
        ##File must be whole eos string
        workbookVariables = openpyxl.load_workbook(file, data_only=True)

        #Inputs
        worksheetInputs = workbookVariables['Inputs']
        lastAttribute = worksheetInputs.cell(1, 2).value
        for i in range(1, worksheetInputs.max_row+1):
            # self.variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            attribute = worksheetInputs.cell(i, 2).value
            try:
                if (attribute == None):
                    if worksheetInputs.cell(i, 3).value is not None:
                        values = read_row(worksheetInputs, i)
                        values = np.array([k for k in values if(str(k))])
                        current = self.getAttribute(self.Inputs, lastAttribute)
                        current = np.vstack((current, values))
                        self.setAttribute(self.Inputs, lastAttribute, current)
                    else:
                        continue
                elif (type(self.getAttribute(self.Inputs, attribute)) == np.ndarray):
                    lastAttribute = attribute
                    values = read_row(worksheetInputs, i)
                    values = np.array([k for k in values if(str(k))])
                    self.setAttribute(self.Inputs, attribute, values)
                else:
                    value = worksheetInputs.cell(i, 3).value
                    self.setAttribute(self.Inputs, attribute, value)
            except TypeError as e:
                if attribute in self.Inputs.__annotations__: raise e
                if attribute=='None' or attribute==None: continue
                print("Error with attribute: {}, continuing.".format(attribute))
        #Options
        worksheetOptions = workbookVariables['Options']
        for i in range(1, worksheetOptions.max_row+1):
            # self.variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            attribute = worksheetOptions.cell(i, 2).value
            try:
                if (type(self.getAttribute(self.Options, attribute)) == np.ndarray):
                    values = read_row(worksheetOptions, i)
                    values = np.array([k for k in values if(str(k))])
                    self.setAttribute(self.Options, attribute, values)
                else:
                    value = worksheetOptions.cell(i, 3).value
                    self.setAttribute(self.Options, attribute, value)
            except  TypeError as e:
                if attribute in self.Options.__annotations__: raise e
                if attribute == 'None' or attribute == None: continue
                print("Error with attribute: {}, continuing.".format(attribute))
        #Plots
        worksheetPlots = workbookVariables['Plots']
        for i in range(1, worksheetPlots.max_row+1):
            # self.variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            attribute = worksheetPlots.cell(i, 2).value
            try:
                if (type(self.getAttribute(self.Plots, attribute)) == np.ndarray):
                    values = read_row(worksheetPlots, i, St=True)[2:]
                    values = np.array([k for k in values if(str(k))])
                    self.setAttribute(self.Plots, attribute, values)
                else:
                    try:
                        value = worksheetPlots.cell(i, 3).value
                    except:
                        value = ''
                    self.setAttribute(self.Plots, attribute, value)
            except  TypeError as e:
                if attribute == 'None' or attribute == None: continue
                print("Error with attribute: {}, continuing.".format(attribute))
        # Variables
        try:
            worksheetVariables = workbookVariables['Variables']
            for i in range(1, worksheetVariables.max_row+1):
                # self.variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
                attribute = worksheetVariables.cell(i, 2).value
                try:
                    if (type(self.getAttribute(self.Variables, attribute)) == np.ndarray):
                        if attribute != 'typeVariableToSaveTxt':  values = read_row(worksheetVariables, i, St = True)[2:]
                        else:  values = read_row(worksheetVariables, i)
                        values = np.array([k for k in values if(str(k))])
                        self.setAttribute(self.Variables, attribute, values)
                    else:
                        value = worksheetVariables.cell(i, 3).value
                        self.setAttribute(self.Variables, attribute, value)
                except TypeError as e:
                    if attribute in self.Variables.__annotations__: raise e
                    if attribute == 'None' or attribute == None: continue
                    print("Error with attribute: {}, continuing.".format(attribute))
        except:
            pass
            print("Error while reading Variables. Please check!")

    def __init__(self):
        self.Inputs = LEDETInputs()
        self.Options = LEDETOptions()
        self.Plots = LEDETPlots()
        self.Variables = LEDETVariables()
        self.variablesInputs = {}
        self.variablesOptions = {}
        self.variablesPlots = {}
        self.variablesVariables = {}

        self.variableGroupInputs = asdict(self.Inputs)
        self.variableGroupOptions = asdict(self.Options)
        self.variableGroupPlots = asdict(self.Plots)
        self.variableGroupVariables = asdict(self.Variables)
        self.sectionTitles = {}

        # Load and set the default LEDET parameters
        self.fileDefaultParameters = os.path.join('ledet', 'variableNamesDescriptions.xlsx')
        self.loadDefaultParameters(self.fileDefaultParameters)

    def loadDefaultParameters(self, fileDefaultParameters: str):
        '''
            **Loads and sets the default LEDET parameters **

            Function to load and set the default LEDET parameters

            :param fileName: String defining the name of the file defining the default LEDET parameters
            :type fileName: str

            :return: None
        '''

        # Load default LEDET parameters
        # Read variable names and descriptions
        fullfileName = ResourceReader.getResourcePath(fileDefaultParameters)
        # print(fullfileName) # for debug
        workbookVariables = openpyxl.load_workbook(fullfileName)

        # Load "Inputs" sheet
        worksheetInputs = workbookVariables['Inputs']
        variablesInputs = {}
        previousVar = 'Start_I'
        for i in range(1, worksheetInputs.max_row+1):

            if str(worksheetInputs.cell(i, 2).value)=='None' and str(worksheetInputs.cell(i, 1).value)!=None:
                if str(worksheetInputs.cell(i, 1).value)=='-': continue
                if str(worksheetInputs.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetInputs.cell(i, 1).value)
                continue
            variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            previousVar = str(worksheetInputs.cell(i, 2).value)

        # Load "Options" sheet
        worksheetOptions = workbookVariables['Options']
        variablesOptions = {}
        previousVar = 'Start_O'
        for i in range(1, worksheetInputs.max_row+1):
            if str(worksheetOptions.cell(i, 2).value)=='None' and str(worksheetOptions.cell(i, 1).value)!=None:
                if str(worksheetOptions.cell(i, 1).value)=='-': continue
                if str(worksheetOptions.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetOptions.cell(i, 1).value)
                continue

            variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            previousVar = str(worksheetOptions.cell(i, 2).value)

        # Load "Plots" sheet
        worksheetPlots = workbookVariables['Plots']
        variablesPlots = {}
        previousVar = 'Start_P'
        for i in range(1, worksheetInputs.max_row+1):
            if str(worksheetPlots.cell(i, 2).value)=='None' and str(worksheetPlots.cell(i, 1).value)!=None:
                if str(worksheetPlots.cell(i, 1).value)=='-': continue
                if str(worksheetPlots.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetPlots.cell(i, 1).value)
                continue
            variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            previousVar = str(worksheetPlots.cell(i, 2).value)

        # Load "Variables" sheet
        worksheetVariables = workbookVariables['Variables']
        variablesVariables = {}
        previousVar = 'Start_V'
        for i in range(1, worksheetInputs.max_row+1):
            if str(worksheetVariables.cell(i, 2).value)=='None' and str(worksheetVariables.cell(i, 1).value)!=None:
                if str(worksheetVariables.cell(i, 1).value)=='-': continue
                if str(worksheetVariables.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetVariables.cell(i, 1).value)
                continue
            variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
            previousVar = str(worksheetVariables.cell(i, 2).value)

        # Set descriptions
        self.variablesInputs, self.variablesOptions, self.variablesPlots, self.variablesVariables = variablesInputs, variablesOptions, variablesPlots, variablesVariables

    def localsParser(self, locals):
        for attribute in locals:
            if attribute in self.Inputs.__annotations__:
                group = self.Inputs
            elif attribute in self.Options.__annotations__:
                group = self.Options
            elif attribute in self.Plots.__annotations__:
                group = self.Plots
            elif attribute in self.Variables.__annotations__:
                group = self.Variables
            else:
                continue
            tt = type(self.getAttribute(group, attribute))
            if tt == np.ndarray and isinstance(locals[attribute], list):
                self.setAttribute(group, attribute, np.array(locals[attribute]))
            elif tt == np.ndarray and not isinstance(locals[attribute], np.ndarray):
                self.setAttribute(group, attribute, np.array([locals[attribute]]))
            else:
                self.setAttribute(group, attribute, locals[attribute])

    def __cpCu_nist_mat(self, T):
        density = 8960
        cpCu_perMass = np.zeros(T.size)
        T[T < 4] = 4
        idxT1 = np.where(T < 300)
        idxT2 = np.where(T >= 300)
        dc_a = -1.91844
        dc_b = -0.15973
        dc_c = 8.61013
        dc_d = -18.996
        dc_e = 21.9661
        dc_f = -12.7328
        dc_g = 3.54322
        dc_h = -0.3797

        logT1 = np.log10(T[idxT1])
        tempVar = \
        dc_a + dc_b * (logT1)**1 + dc_c * (logT1)**2 + dc_d * (logT1)**3 + \
        dc_e * (logT1)**4 + dc_f * (logT1)**5 + dc_g * (logT1)** 6 + dc_h * (logT1)**7
        cpCu_perMass[idxT1] = 10**tempVar

        cpCu_perMass[idxT2]= 361.5 + 0.093 * T[idxT2]
        cpCu = density * cpCu_perMass
        return cpCu

    def __rhoCu_nist(self, T, B, f_MR = 1):
        RRR = np.repeat(self.Inputs.RRR_Cu_inGroup,self.Inputs.nT.astype(int))
        B = abs(B)

        idxLowB = np.where(B < 0.1)
        idxHighB = np.where(B >= 0.1)

        rho0 = 1.553e-8 / RRR
        rhoi = 1.171e-17 * (T** 4.49) / (1 + 4.48e-7 * (T** 3.35) * np.exp(-(50. / T)** 6.428))
        rhoiref = 0.4531 * rho0 * rhoi / (rho0 + rhoi)
        rhcu = rho0 + rhoi + rhoiref
        rhoCu = np.zeros(B.shape)
        rhoCu[idxLowB] = rhcu[idxLowB]

        lgs = 0.43429 * np.log(1.553E-8 * B[idxHighB] / rhcu[idxHighB])
        polys = -2.662 + lgs * (0.3168 + lgs * (0.6229 + lgs  * (-0.1839 + lgs * 0.01827)))
        corrs = (10.**polys)
        rhoCu[idxHighB] = (1. + corrs * f_MR) * rhcu[idxHighB]
        return rhoCu

    def __kCu_WiedemannFranz(self, rhoCu, T):
        L0 = 2.45E-8
        kCu = L0 * T / rhoCu
        return kCu

    def __kG10(self, T):
        kG10 = np.zeros(T.size)
        LimitValidity = 500
        idxT1 = np.where(T <= LimitValidity)
        idxT2 = np.where(T > LimitValidity)

        a, b, c, d, e, f, g, h = -4.1236, 13.788, -26.068, 26.272, -14.663, 4.4954, -0.6905, 0.0397
        logT = np.log10(T[idxT1]);
        logk = a + b * logT + c * logT**2 + d * logT**3 + e * logT**4 + f * logT**5 + g * logT**6 + h * logT**7
        kG10[idxT1] = 10**logk;

        logLimitValidity = np.log10(LimitValidity)
        logkLimitValidity = a + b * logLimitValidity + c * logLimitValidity**2 + d * logLimitValidity**3 + e * logLimitValidity**4 + \
        f * logLimitValidity**5 + g * logLimitValidity**6 + h * logLimitValidity**7;
        kG10[idxT2] = 10**logkLimitValidity
        return kG10

    def __cpG10(self, T):
        density, a0, a1, a2, a3, a4, a5, a6, a7 = 1900, -2.4083, 7.6006, -8.2982, 7.3301, -4.2386, 1.4294, -0.24396, 0.015236
        logT = np.log10(T)
        p = 10**(a7 * ((logT)**7) + a6 * ((logT)**6) + a5 * ((logT)**5) + a4 * ((logT)**4) + a3 * ((logT)**3) + a2 * (
                    (logT)**2) + a1 * ((logT)) + a0)
        cpG10 = density * p
        return cpG10

    def __kKapton(self, T):
        kKapton = np.zeros(T.size)
        idxLow = np.where(T < 4.3)
        if idxLow:
            kKapton[idxLow[0]] = 0.010703 - 0.00161 * (4.3 - T[idxLow[0]])
        idxHigh = np.where(T >= 4.3)
        if idxHigh:
            a, b, c, d, e, f, g, h = 5.73101, -39.5199, 79.9313, -83.8572, 50.9157, -17.9835, 3.42413, -0.27133
            logT = np.log10(T)
            logk = a + b * logT + c * logT**2 + d * logT**3 + e * logT**4 + f * logT**5 + g * logT**6 + h * logT**7
            kKapton[idxHigh[0]] = 10**logk[idxHigh[0]]
        return kKapton

    def __cpKapton(self, T):
        density, a0, a1, a2, a3, a4, a5, a6, a7 = 1420, -1.3684, 0.65892, 2.8719, 0.42651, -3.0088, 1.9558, -0.51998, 0.051574
        logT = np.log10(T)
        p = 10**(a7 * ((logT)**7) + a6 * ((logT)**6) + a5 * ((logT)**5) + a4 * ((logT)**4) + a3 * (
                    (logT)**3) + a2 * ((logT)**2) + a1 * ((logT)) + a0)
        cpKapton = density * p
        return cpKapton

    def __cpNbTi_cudi_mat(self, T, B):
        Tc0 = 9.2
        Bc20 = 14.5
        alpha = .59
        B[B>= Bc20] = Bc20-10E-4

        Tc = Tc0 * (1 - B / Bc20)**alpha
        cpNbTi = np.zeros(T.size)

        idxT1 = np.where(T <= Tc)
        idxT2 = np.where((T > Tc) & (T <= 20.0))
        idxT3 = np.where((T > 20) & (T <= 50))
        idxT4 = np.where((T > 50) & (T <= 175))
        idxT5 = np.where((T > 175) & (T <= 500))
        idxT6 = np.where((T > 500) & (T <= 1000))
        idxT7 = np.where(T > 1000)

        p1 = [0.00000E+00,    4.91000E+01,   0.00000E+00,   6.40000E+01,  0.00000E+00]
        p2 = [0.00000E+00,   1.62400E+01,   0.00000E+00,  9.28000E+02,   0.00000E+00]
        p3 = [-2.17700E-01,   1.19838E+01,   5.53710E+02, - 7.84610E+03,  4.13830E+04]
        p4 = [-4.82000E-03,  2.97600E+00, -7.16300E+02,  8.30220E+04,  -1.53000E+06]
        p5 = [-6.29000E-05, 9.29600E-02, -5.16600E+01,  1.37060E+04,  1.24000E+06]
        p6 = [0.00000E+00, 0.00000E+00,  -2.57000E-01,  9.55500E+02,  2.45000E+06]
        p7 = [0, 0, 0, 0, 3.14850E+06]

        cpNbTi[idxT1] = p1[0] * T[idxT1]**4 + p1[1] * T[idxT1]**3 + p1[2] * T[idxT1]**2 + p1[3] * T[idxT1] + p1[4]
        cpNbTi[idxT2] = p2[0] * T[idxT2]**4 + p2[1] * T[idxT2]**3 + p2[2] * T[idxT2]**2 + p2[3] * T[idxT2] + p2[4]
        cpNbTi[idxT3] = p3[0] * T[idxT3]**4 + p3[1] * T[idxT3]**3 + p3[2] * T[idxT3]**2 + p3[3] * T[idxT3] + p3[4]
        cpNbTi[idxT4] = p4[0] * T[idxT4]**4 + p4[1] * T[idxT4]**3 + p4[2] * T[idxT4]**2 + p4[3] * T[idxT4] + p4[4]
        cpNbTi[idxT5] = p5[0] * T[idxT5]**4 + p5[1] * T[idxT5]**3 + p5[2] * T[idxT5]**2 + p5[3] * T[idxT5] + p5[4]
        cpNbTi[idxT6] = p6[0] * T[idxT6]**4 + p6[1] * T[idxT6]**3 + p6[2] * T[idxT6]**2 + p6[3] * T[idxT6] + p6[4]
        cpNbTi[idxT7] = p7[0] * T[idxT7]**4 + p7[1] * T[idxT7]**3 + p7[2] * T[idxT7]**2 + p7[3] * T[idxT7] + p7[4]
        return cpNbTi

    def __cpNb3Sn_alternative_mat(self, T, B, Tc0_Nb3Sn, Bc20_Nb3Sn):
        B[B < .001] = 0.001
        cpNb3Sn = np.zeros(T.shape)
        alpha = .59
        Tc = Tc0_Nb3Sn * (1 - B / Bc20_Nb3Sn)** alpha
        density = 8950.0 # [kg / m ^ 3]

        idxT0 = np.where(T <= Tc)
        idxT1 = np.where((T > Tc) & (T <= 20))
        idxT2 = np.where((T > 20) & (T <= 400))
        idxT3 = np.where(T > 400)


        betaT = 1.241E-3 # [J / K ^ 4 / kg]
        gammaT = .138 # [J / K ^ 2 / kg]

        if len(B) > 1 and len(Tc0_Nb3Sn) > 1:
            cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn[idxT0]** 2) * T[idxT0]** 3 + gammaT* B[idxT0] / Bc20_Nb3Sn[idxT0] * T[idxT0]
        elif len(B) > 1:
            cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn** 2) * T[idxT0]**3 + gammaT * B[idxT0] / Bc20_Nb3Sn * T[idxT0]
        elif len(Tc0_Nb3Sn) > 1:
            cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn[idxT0]** 2) * T[idxT0]**3 + gammaT * B / Bc20_Nb3Sn[idxT0] * T[idxT0]
        elif len(B) == 1 and len(Tc0_Nb3Sn) == 1:
            cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn**2) * T[idxT0]**3 + gammaT * B / Bc20_Nb3Sn * T[idxT0]

        cpNb3Sn[idxT1] = betaT * T[idxT1]**3 + gammaT * T[idxT1]
        polyFit_20K_400K = [0.1662252, -0.6827738, -6.3977, 57.48133, -186.90995, 305.01434, -247.44839, 79.78547]
        logT = np.log10(T[idxT2])
        logCp2 = np.polyval(polyFit_20K_400K, logT)
        cpNb3Sn[idxT2] = 10** logCp2

        log400K = np.log10(400)
        logCp400K = np.polyval(polyFit_20K_400K, log400K)
        cpNb3Sn[idxT3] = 10**logCp400K
        cpNb3Sn = cpNb3Sn * density
        return cpNb3Sn

    def __Jc_Nb3Sn_Summer(self, T, B, Jc_Nb3Sn0, Tc0_Nb3Sn, Bc20_Nb3Sn):
        if type(T)== int or type(T)== float:
            T = np.repeat(T, len(Jc_Nb3Sn0)).astype(float)

        B[abs(B) < .001] = 0.001
        T[T < 0.001] = 0.001
        f_T_T0 = T / Tc0_Nb3Sn
        f_T_T0[f_T_T0 > 1] = 1
        Bc2 = Bc20_Nb3Sn * (1 - f_T_T0**2) * (1 - 0.31 * f_T_T0**2 * (1 - 1.77 * np.log(f_T_T0)))
        f_B_Bc2 = B / Bc2
        f_B_Bc2[f_B_Bc2 > 1] = 1
        Jc_T_B = Jc_Nb3Sn0 / np.sqrt(B) * (1 - f_B_Bc2)**2 * (1 - f_T_T0** 2)**2
        return Jc_T_B

    def __Tc_Tcs_Nb3Sn_approx(self, J, B, Jc_Nb3Sn0, Tc0_Nb3Sn, Bc20_Nb3Sn):
        J = abs(J)
        B = abs(B)

        f_B_Bc2 = B / Bc20_Nb3Sn
        f_B_Bc2[f_B_Bc2 > 1] = 1
        Tc = Tc0_Nb3Sn * (1 - f_B_Bc2)**.59

        Jc0 = self.__Jc_Nb3Sn_Summer(0, B, Jc_Nb3Sn0, Tc0_Nb3Sn, Bc20_Nb3Sn)
        f_J_Jc0 = J/ Jc0
        f_J_Jc0[f_J_Jc0 > 1] = 1

        Tcs = (1 - f_J_Jc0) * Tc
        return [Tc, Tcs]

    def _obtainThermalConnections(self):
        # Calculate group to which each half-turn belongs
        nHalfTurnsDefined = len(self.Inputs.HalfTurnToInductanceBlock)
        indexTstart = np.hstack([1, 1 + np.cumsum(self.Inputs.nT[:-1])]).astype(int)
        indexTstop = np.cumsum(self.Inputs.nT).astype(int)
        HalfTurnToGroup = np.zeros((1, nHalfTurnsDefined), dtype=int)
        HalfTurnToGroup = HalfTurnToGroup[0]
        for g in range(1, len(self.Inputs.nT) + 1):
            HalfTurnToGroup[indexTstart[g - 1] - 1:indexTstop[g - 1]] = g

        # Obtain all thermal connections of each turn and store them in dictionaries for width and height
        # First width
        th_con_w = {}
        for i in range(1, len(self.Inputs.HalfTurnToInductanceBlock) + 1):
            con_list = []
            iWidthFrom = np.where(self.Inputs.iContactAlongWidth_From == i)
            if iWidthFrom: con_list = con_list + self.Inputs.iContactAlongWidth_To[iWidthFrom[0]].astype(int).tolist()
            iWidthTo = np.where(self.Inputs.iContactAlongWidth_To == i)
            if iWidthTo: con_list = con_list + self.Inputs.iContactAlongWidth_From[iWidthTo[0]].astype(int).tolist()
            th_con_w[str(i)] = con_list

        # Then height
        th_con_h = {}
        for i in range(1, len(self.Inputs.HalfTurnToInductanceBlock) + 1):
            con_list = []
            iHeightFrom = np.where(self.Inputs.iContactAlongHeight_From == i)
            if iHeightFrom: con_list = con_list + self.Inputs.iContactAlongHeight_To[iHeightFrom[0]].astype(
                int).tolist()
            iHeightTo = np.where(self.Inputs.iContactAlongHeight_To == i)
            if iHeightTo: con_list = con_list + self.Inputs.iContactAlongHeight_From[iHeightTo[0]].astype(int).tolist()
            th_con_h[str(i)] = con_list
        return [HalfTurnToGroup, th_con_w, th_con_h]

    def __calculateTransversalDelay(self, cp, kIns, Tc, Tcs, T_bath):
        [HalfTurnToGroup, th_con_w, th_con_h] = self._obtainThermalConnections()
        # Use dictionaries to calculate the transversal quench delay into each direction based on respective properties
        delta_t_w = {}
        delta_t_h = {}
        for i in range(1,len(self.Inputs.HalfTurnToInductanceBlock)+1):
            con = th_con_h[str(i)]
            delta_t_h_temp = []
            for k in range(len(con)):
                idx_con1 = HalfTurnToGroup[k-1]-1
                idx_con2 = HalfTurnToGroup[con[k]-1]-1
                T_temp = 1
                delta_t = cp[idx_con2] / kIns[idx_con2] * (
                            self.Inputs.wBare_inGroup[idx_con2] + 2 * self.Inputs.wIns_inGroup[idx_con2]) \
                          * (self.Inputs.wIns_inGroup[idx_con2] + self.Inputs.wIns_inGroup[idx_con1]) * T_temp
                delta_t_h_temp.append(delta_t)
            delta_t_h[str(i)] = delta_t_h_temp

            con = th_con_w[str(i)]
            delta_t_w_temp = []
            for k in range(len(con)):
                idx_con1 = HalfTurnToGroup[k - 1] - 1
                idx_con2 = HalfTurnToGroup[con[k]-1]-1
                T_temp = (Tcs[idx_con2]-T_bath)/(Tc[idx_con1]-(Tcs[idx_con2]+T_bath)/2)
                delta_t = cp[idx_con2] / kIns[idx_con2] * (
                            self.Inputs.hBare_inGroup[idx_con2] + 2 * self.Inputs.hIns_inGroup[idx_con2]) \
                          * (self.Inputs.hIns_inGroup[idx_con2] + self.Inputs.hIns_inGroup[idx_con1]) * T_temp
                delta_t_w_temp.append(delta_t)
            delta_t_w[str(i)] = delta_t_w_temp

        return [th_con_h, delta_t_h, th_con_w, delta_t_w]

    def _quenchPropagationVelocity(self, I, B, T_bath, cable):
        # Calculate Quench propagation velocity
        L0 = 2.44E-08
        A_CableBare = cable.A_CableInsulated * (cable.f_SC + cable.f_ST)
        f_SC_inStrand = cable.f_SC / (cable.f_SC + cable.f_ST)
        f_ST_inStrand = cable.f_ST / (cable.f_SC + cable.f_ST)
        I = abs(I)
        J_op = I / A_CableBare

        idxNbTi = np.where(np.repeat(self.Inputs.SCtype_inGroup,self.Inputs.nT.astype(int)) == 1)[0]
        idxNb3Sn = np.where(np.repeat(self.Inputs.SCtype_inGroup,self.Inputs.nT.astype(int)) == 2)[0]
        idxCu_ST = np.where(np.repeat(self.Inputs.STtype_inGroup,self.Inputs.nT.astype(int)) == 1)[0]

        Tc = np.zeros(B.shape)
        Tcs = np.zeros(B.shape)
        if len(idxNbTi)>0:
            Tc[idxNbTi] = cable.Tc0_NbTi * (1 - B / cable.Bc20_NbTi) ** cable.alpha_NbTi
            Tcs[idxNbTi] = (1 - I / (cable.c1_Ic_NbTi + cable.c2_Ic_NbTi * B)) * Tc
        if len(idxNb3Sn) > 0:
            [Tc[idxNb3Sn], Tcs[idxNb3Sn]] = self.__Tc_Tcs_Nb3Sn_approx(I / cable.A_SC[idxNb3Sn], B[idxNb3Sn],
                                                                cable.Jc_Nb3Sn0[idxNb3Sn], cable.Tc0_Nb3Sn[idxNb3Sn],
                                                                cable.Bc20_Nb3Sn[idxNb3Sn])

        Ts = (Tcs + Tc) / 2
        cp_ST = np.zeros(B.shape)
        cp_ST[idxCu_ST] = self.__cpCu_nist_mat(Ts)
        cp_SC = np.zeros(B.shape)
        if len(idxNbTi) > 0:
            cp_SC[idxNbTi] = self.__cpNbTi_cudi_mat(Ts, B)
        if len(idxNb3Sn) > 0:
            cp_SC[idxNb3Sn] = self.__cpNb3Sn_alternative_mat(Ts[idxNb3Sn], B[idxNb3Sn], cable.Tc0_Nb3Sn[idxNb3Sn], cable.Bc20_Nb3Sn[idxNb3Sn])
        cp = cp_ST * f_ST_inStrand + cp_SC * f_SC_inStrand
        vQ = J_op / cp * ((L0 * Ts) / (Ts - T_bath))**0.5
        idxInfQuenchVelocity=np.where(Tcs <= T_bath)
        vQ[idxInfQuenchVelocity]=1E6

        ### Calculate MPZ
        rhoCu = self.__rhoCu_nist(Ts, B)
        kCu = self.__kCu_WiedemannFranz(rhoCu, Ts)
        l = np.sqrt((2 * kCu * (Tc - T_bath)) / (J_op ** 2 * rhoCu))

        # Calculate thermal conductivity of insulations
        idxKapton = np.where(self.Inputs.insulationType_inGroup == 2, 1, 0)
        idxKapton = np.where(np.repeat(idxKapton, self.Inputs.nT.astype(int))==1)[0]
        idxG10 = np.where(self.Inputs.insulationType_inGroup == 1, 1, 0)
        idxG10 = np.where(np.repeat(idxG10, self.Inputs.nT.astype(int))==1)[0]
        kIns = np.zeros(Ts.size)
        kIns[idxKapton] = self.__kKapton(Ts[idxKapton])
        kIns[idxG10] = self.__kG10(Ts[idxG10])
        cpIns = np.zeros(Ts.size)
        cpIns[idxKapton] = self.__cpKapton(Ts[idxKapton])
        cpIns[idxG10] = self.__cpG10(Ts[idxG10])
        cp_full = (cp* (A_CableBare/cable.A_CableInsulated) + cpIns*(1-A_CableBare/cable.A_CableInsulated))/2

        ### Calculate delta T transversal
        [th_con_h, delta_t_h, th_con_w, delta_t_w] = self.__calculateTransversalDelay(cp_full, kIns, Tc, Tcs, T_bath)
        return [vQ, l, th_con_h, delta_t_h, th_con_w, delta_t_w]

    def __reorderROXIEFiles(self, ROXIE_File):
        orderedROXIE = []
        for i in range(len(ROXIE_File)-1):
            prefix = 'E'+str(i)
            for j in range(len(ROXIE_File)):
                if prefix in ROXIE_File[j]:
                    orderedROXIE.append(ROXIE_File[j])
        for j in range(len(ROXIE_File)):
            if 'All_WithIron_WithSelfField' in ROXIE_File[j]:
                orderedROXIE.append(ROXIE_File[j])
        return orderedROXIE

    def _acquireBField(self, ROXIE_File):
        if ROXIE_File.endswith('.map2d'):
            ROXIE_File = [ROXIE_File]
            N = 1
        else:
            ROXIE_File1 = [f for f in os.listdir(ROXIE_File) if os.path.isfile(os.path.join(ROXIE_File, f))]
            ROXIE_File1 = [f for f in ROXIE_File1 if f.endswith('.map2d')]
            ROXIE_File1 = [f for f in ROXIE_File1 if 'WithIron' in f]
            for i in range(len(ROXIE_File1)):
                ROXIE_File1[i] = os.path.join(ROXIE_File, ROXIE_File1[i])
            ROXIE_File = ROXIE_File1
            ROXIE_File = self.__reorderROXIEFiles(ROXIE_File)
            N = len(ROXIE_File)
            if N>1:
                print('Reading ', N, ' Field maps. This may take a while.')
            else:
                print('Reading Field map.')

        for i in trange(N, file=sys.stdout, desc='Field maps'):
            Inom = self.Options.Iref
            reader = csv.reader(open(ROXIE_File[i]))
            B_Field = np.array([])
            stack = 0
            for row in reader:
                if not row: continue
                row_s = np.array(row[0].split())
                if not stack:
                    B_Field = np.array(row_s[1:])
                    stack = 1
                else:
                    B_Field = np.vstack((B_Field, np.array(row_s)))
            B_Field = B_Field[1:].astype(float)
            if i == 0:
                BX = (B_Field[:, 5].transpose()/ Inom)
                BY = (B_Field[:, 6].transpose()/ Inom)
            elif i == N-1:
                BX_All = B_Field[:, 5].transpose()
                BY_All = B_Field[:, 6].transpose()
            else:
                BX = BX + (B_Field[:, 5].transpose() / Inom)
                BY = BY + (B_Field[:, 6].transpose() / Inom)
        f_mag = (BX** 2 + BY** 2) ** 0.5
        if N>1:
            B_E_All = (BX_All** 2 + BY_All** 2) ** 0.5
            peakB_Superposition = max(f_mag * Inom)
            peakB_Real = max(B_E_All)
            f_peakReal_peakSuperposition = peakB_Real / peakB_Superposition
        else: f_peakReal_peakSuperposition = 1

        B = f_mag*self.Inputs.I00*f_peakReal_peakSuperposition

        B[B > 10E6]=10E-6
        return B

    def __repeatCable(self, cable):
        nT = self.Inputs.nT
        nT = nT.astype(int)
        newCable = Cable()
        for attribute in cable.__annotations__:
            if attribute == 'alpha_NbTi': continue
            x = np.ndarray([])
            x = getattr(cable, attribute)
            x = np.repeat(x, nT)
            setattr(newCable, attribute, x)
        return newCable

    def adjust_vQ(self, ROXIE_File, Return  = False):
        cable = Cable()
        cable.A_CableInsulated = (self.Inputs.wBare_inGroup+2*self.Inputs.wIns_inGroup) \
                               * (self.Inputs.hBare_inGroup+2*self.Inputs.hIns_inGroup)
        B = self._acquireBField(ROXIE_File)

        if max(self.Inputs.nStrands_inGroup) > 1:
            strandCount = 0
            GroupCount = 0
            nStrands_inGroup = self.Inputs.nStrands_inGroup
            ds_inGroup = self.Inputs.ds_inGroup
            if any(nStrands_inGroup % 2 == 1) and any(nStrands_inGroup != 1):
                for g in range(len(self.Inputs.nT)):
                    if (nStrands_inGroup[g] % 2 == 1) & (nStrands_inGroup[g] > 1):
                        ds_inGroup[g] = ds_inGroup[g] * np.sqrt(nStrands_inGroup[g] / (nStrands_inGroup[g] - 1))
                        nStrands_inGroup[g] = nStrands_inGroup[g] - 1

            Bcopy = np.zeros(int(sum(self.Inputs.nT)))
            for i in range(int(sum(self.Inputs.nT))):
                Bcopy[i] = sum(B[int(strandCount):int(strandCount+nStrands_inGroup[GroupCount])])/nStrands_inGroup[int(GroupCount)]
                TurnSum = sum(self.Inputs.nT[0:GroupCount+1])
                strandCount = strandCount + nStrands_inGroup[GroupCount]
                if i>TurnSum: GroupCount = GroupCount + 1
            B = Bcopy

            cable.f_SC = self.Inputs.f_SC_strand_inGroup * \
                         (nStrands_inGroup* (np.pi/4)*(ds_inGroup**2)) / cable.A_CableInsulated
            cable.f_ST = (1 - self.Inputs.f_SC_strand_inGroup) * \
                         (nStrands_inGroup* (np.pi/4)*(ds_inGroup**2)) / cable.A_CableInsulated
        else:
            cable.f_SC = self.Inputs.f_SC_strand_inGroup * \
                         (self.Inputs.wBare_inGroup * self.Inputs.hBare_inGroup) / cable.A_CableInsulated
            cable.f_ST = (1 - self.Inputs.f_SC_strand_inGroup) * \
                         (self.Inputs.wBare_inGroup * self.Inputs.hBare_inGroup) / cable.A_CableInsulated

        I = self.Inputs.I00
        T_bath = self.Inputs.T00
        cable.A_SC =cable.A_CableInsulated * cable.f_SC
        cable.SCtype = self.Inputs.SCtype_inGroup
        cable.STtype = self.Inputs.STtype_inGroup
        cable.Tc0_NbTi = self.Inputs.Tc0_NbTi_ht_inGroup
        cable.Bc20_NbTi = self.Inputs.Bc2_NbTi_ht_inGroup
        cable.c1_Ic_NbTi = self.Inputs.c1_Ic_NbTi_inGroup
        cable.c2_Ic_NbTi = self.Inputs.c2_Ic_NbTi_inGroup
        cable.alpha_NbTi = .59
        cable.Jc_Nb3Sn0 = self.Inputs.Jc_Nb3Sn0_inGroup
        cable.Tc0_Nb3Sn = self.Inputs.Tc0_Nb3Sn_inGroup
        cable.Bc20_Nb3Sn = self.Inputs.Bc2_Nb3Sn_inGroup
        cable = self.__repeatCable(cable)

        [vQ, l, th_con_h, delta_t_h, th_con_w, delta_t_w] = self._quenchPropagationVelocity(I, B, T_bath, cable)
        self.setAttribute(getattr(self, "Inputs"), "vQ_iStartQuench", vQ)
        #self.setAttribute(getattr(self, "Inputs"), "lengthHotSpot_iStartQuench", l)
        if Return:
            return [vQ, l, th_con_h, delta_t_h, th_con_w, delta_t_w]
        else:
            return

    def addVariablesInputs(self,
                           T00, l_magnet, I00, M_m,
                           fL_I, fL_L,
                           GroupToCoilSection, polarities_inGroup, nT, nStrands_inGroup, l_mag_inGroup, ds_inGroup,
                           f_SC_strand_inGroup, f_ro_eff_inGroup, Lp_f_inGroup, RRR_Cu_inGroup,
                           SCtype_inGroup, STtype_inGroup, insulationType_inGroup, internalVoidsType_inGroup,
                           externalVoidsType_inGroup,
                           wBare_inGroup, hBare_inGroup, wIns_inGroup, hIns_inGroup, Lp_s_inGroup, R_c_inGroup,
                           Tc0_NbTi_ht_inGroup, Bc2_NbTi_ht_inGroup, c1_Ic_NbTi_inGroup, c2_Ic_NbTi_inGroup,
                           Tc0_Nb3Sn_inGroup, Bc2_Nb3Sn_inGroup, Jc_Nb3Sn0_inGroup,
                           el_order_half_turns,
                           alphasDEG, rotation_block, mirror_block, mirrorY_block,
                           iContactAlongWidth_From, iContactAlongWidth_To, iContactAlongHeight_From,
                           iContactAlongHeight_To,
                           iStartQuench, tStartQuench, lengthHotSpot_iStartQuench, vQ_iStartQuench,
                           R_circuit, R_crowbar, Ud_crowbar, t_PC, t_PC_LUT, I_PC_LUT,
                           tEE, R_EE_triggered,
                           tCLIQ, directionCurrentCLIQ, nCLIQ, U0, C, Rcapa,
                           tQH, U0_QH, C_QH, R_warm_QH, w_QH, h_QH, s_ins_QH, type_ins_QH, s_ins_QH_He, type_ins_QH_He, l_QH, f_QH,
                           iQH_toHalfTurn_From, iQH_toHalfTurn_To,
                           tQuench, initialQuenchTemp,
                           HalfTurnToInductanceBlock, M_InductanceBlock_m, *args,**kwargs
                           ):
        '''
            **Adds all LEDET parameters to be written in the "Inputs" sheet **

            Function to add "Inputs" LEDET parameters
            Parameter represent the LEDET equivalents

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            try:
                self.setAttribute(self.Inputs, attribute, ins[attribute])
            except:
                print('Could not store ',attribute, ', not found.')
        for attribute in kwargs:
            try:
                self.setAttribute(self.Inputs, attribute, kwargs[attribute])
            except:
                pass

    def addVariablesOptions(self,
                            time_vector_params,
                            Iref, flagIron, flagSelfField, headerLines, columnsXY, columnsBxBy, flagPlotMTF,
                            flag_calculateInductanceMatrix, flag_useExternalInitialization, flag_initializeVar,
                            flag_fastMode, flag_controlCurrent, flag_automaticRefinedTimeStepping, flag_IronSaturation,
                            flag_InvertCurrentsAndFields, flag_ScaleDownSuperposedMagneticField, flag_HeCooling,
                            fScaling_Pex, fScaling_Pex_AlongHeight,
                            fScaling_MR, flag_scaleCoilResistance_StrandTwistPitch, flag_separateInsulationHeatCapacity,
                            flag_ISCL, fScaling_Mif, fScaling_Mis, flag_StopIFCCsAfterQuench, flag_StopISCCsAfterQuench,
                            tau_increaseRif, tau_increaseRis,
                            fScaling_RhoSS, maxVoltagePC, flag_symmetricGroundingEE, flag_removeUc, BtX_background,
                            BtY_background,
                            flag_showFigures, flag_saveFigures, flag_saveMatFile, flag_saveTxtFiles,
                            flag_generateReport,
                            flag_hotSpotTemperatureInEachGroup, *args,**kwargs
                           ):
        '''
            **Adds all LEDET parameters to be written in the "Options" sheet **

            Function to add "Options" LEDET parameters
            Parameter represent the LEDET equivalents

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            try:
                self.setAttribute(self.Options, attribute, ins[attribute])
            except:
                print('Could not store ',attribute, ', not found.')
        for attribute in kwargs:
            try:
                self.setAttribute(self.Options, attribute, kwargs[attribute])
            except:
                pass

    def addVariablesPlots(self,
                          suffixPlot, typePlot, outputPlotSubfolderPlot, variableToPlotPlot, selectedStrandsPlot,
                          selectedTimesPlot,
                          labelColorBarPlot, minColorBarPlot, maxColorBarPlot, MinMaxXYPlot, flagSavePlot,
                          flagColorPlot, flagInvisiblePlot, *args,**kwargs
                           ):
        '''
            **Adds all LEDET parameters to be written in the "Plots" sheet **

            Function to add "Plots" LEDET parameters
            Parameter represent the LEDET equivalents

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            try:
                self.setAttribute(self.Plots, attribute, ins[attribute])
            except:
                print('Could not store ', attribute, ', not found.')
        for attribute in kwargs:
            try:
                self.setAttribute(self.Plots, attribute, kwargs[attribute])
            except:
                pass
    def addVariablesVariables(self,
                              variableToSaveTxt, typeVariableToSaveTxt, variableToInitialize, *args,**kwargs
                              ):
        '''
            **Adds all LEDET parameters to be written in the "Variables" sheet **

            Function to add "Variables" LEDET parameters

            :param variableToSaveTxt: Array including LEDET variable
            :type variableToSaveTxt: np.ndarray()
            :param typeVariableToSaveTxt: Array including LEDET variable
            :type typeVariableToSaveTxt: np.ndarray()
            :param variableToInitialize: Array including LEDET variable
            :type variableToInitialize: np.ndarray()

            :return: None
        '''
        ins = locals()
        for attribute in ins:
            try:
                self.setAttribute(self.Variables, attribute, ins[attribute])
            except:
                print('Could not store ', attribute, ', not found.')
        for attribute in kwargs:
            try:
                self.setAttribute(self.Variables, attribute, kwargs[attribute])
            except:
                pass


    def printVariableDescNameValue(self, variableGroup, variableLabels):
        """

           **Print variable description, variable name, and variable value**

           Function prints variable description, variable name, and variable value

           :param variableGroup: Dataclass containing all the attributes of the LEDET object
           [obsolete, but still supported: list of tuples; each tuple has two elements: the first element is a string defining
           the variable name, and the second element is either an integer, a float, a list, or a numpy.ndarray
           defining the variable value :type variableGroup: list :param variableLabels: dictionary assigning a
           description to each variable name]
           :type variableLabels: dataclass [obsolete, but still supported: dict]

           :return: None

           [Example for usage of obsolete dictionary-version]
            import numpy as np
            variableGroup = []
            variableGroup.append( ('x1', 12) )
            variableGroup.append( ('x2', 23.42) )
            variableGroup.append( ('x3', [2, 4, 6]) )
            variableGroup.append( ('x3', np.array([2, 4, 6])) )
            variableLabels = {'x1': '1st variable', 'x2': '2nd variable', 'x3': '3rd variable'}
            utils.printVariableDescNameValue(variableGroup, variableLabels)
            # >>> 					1st variable x1 12
            # >>> 					2nd variable x2 23.42
            # >>> 					3rd variable x3 [2, 4, 6]
            # >>> 					3rd variable x3 [2 4 6]

        """
        if(variableGroup == self.variableGroupInputs):
            variableGroup=self.Inputs
        if (variableGroup == self.variableGroupOptions):
            variableGroup = self.Options
        if (variableGroup == self.variableGroupPlots):
            variableGroup = self.Plots
        if (variableGroup == self.variableGroupVariables):
            variableGroup = self.Variables

        if(type(variableGroup) != dict):
            for k in variableGroup.__annotations__:
                if k == 'overwrite_f_internalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                if k == 'overwrite_f_externalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                print(variableLabels[k])
                print(k, self.getAttribute(variableGroup, k))
        else:
            for k in variableGroup:
                if k == 'overwrite_f_internalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                if k == 'overwrite_f_externalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                print(variableLabels[k])
                print(k, variableGroup[k])

    def __getNumberOfCoilSections(self):
        '''
            **Consistency check of LEDET Inputs - Helper function**

            Returns the number of CoilSections from Mutual-inductance matrix

            :return: int
        '''
        k = self.Inputs.M_m
        if k.shape == (1,): return k.shape[0]
        try:
            if k.shape[0] != k.shape[1]: print("M_m is not square")
        except:
            print("M_m is not square")
            return -1
        k2 = max(self.Inputs.GroupToCoilSection)
        if k.shape[0] != k2:
            print('M_m matrix does have size: ',k.shape[0], ' but you assign a Coil-Section: ',k2)
            return -1
        return k.shape[0]

    def __checkM_InductanceBlock_m(self, Turns):
        '''
            **Consistency check of LEDET Inputs - Inductance Matrix **

            Check if Inductance matrix is squared. Issues a warning if it is not the size of the number of turns. Returns result as bool

            :param Turns: Number of turns of the LEDET object
            :type arr: int
            :return: bool
        '''
        if type(self.Inputs.M_InductanceBlock_m) != np.ndarray:
            k = np.array(self.Inputs.M_InductanceBlock_m)
        else:
            k = self.Inputs.M_InductanceBlock_m
        ## Account for the option to set the matrix to 0
        if k.shape == (1,):
            return True
        try:
            if k.shape[0] == k.shape[1]:
                if k.shape[0] != Turns:
                    print("M_InductanceBlock_m is squared, but its size unequal to the number of turns")
                return True
        except:
            print("M_InductanceBlock_m is not correct!")
            return False
        print("M_InductanceBlock_m is not correct!")
        return False

    def __checkHeFraction(self, Groups):
        '''
            **Consistency check of LEDET Inputs - Helium options check **

            Check if Helium options are both set and have the correct and same size, returns bool of result

            :param Groups: Number of groups of the LEDET object
            :type arr: int
            :return: bool
        '''
        k = self.Inputs.overwrite_f_externalVoids_inGroup
        k2 = self.Inputs.overwrite_f_internalVoids_inGroup
        if len(k) > 0:
            if len(k) != len(k2):
                print("Helium section was set but is corrupted.")
                return False
            if len(k) != Groups:
                print("Helium section was set but is wrong length.")
                return False
        elif len(k2) > 0:
            print("Helium section was set but is corrupted.")
            return False
        return True

    def __checkMonotony(self, arr, Invert=False):
        '''
            **Consistency check of LEDET Inputs - Monotony check **

            Check if given array is monotone or not. Returns bool of result.

            :param arr: Given array to be checked
            :type arr: np.ndarray()
            :param Invert: flag that determines which direction the array should be interpreted [True= from the last towards the first, False= from the first towards the last]
            :type Invert: bool
            :return: bool
        '''
        if Invert:
            arr = np.flip(arr)
        b = all(x <= y for x, y in zip(arr, arr[1:]))
        return b

    def __checkTimes(self):
        '''
            **Consistency check of LEDET Inputs - check Times **

            Function that checks if times in LEDET are all set accordingly, otherwise adjusts the time_vector
            :return: none
        '''

        ### Check start of time_vector
        # obtain all times from LEDET object
        try: t1 = min(self.Inputs.tQuench)
        except: t1 = min(np.array([self.Inputs.tQuench]))
        t2 = self.Inputs.t_PC_LUT[0]
        try: t3 = min(self.Inputs.tStartQuench)
        except: t3 = min(np.array([self.Inputs.tStartQuench]))
        t4 = self.Options.time_vector_params[0]
        try:  t5 = min(self.Inputs.tQH)
        except: t5 = min(np.array([self.Inputs.tQH]))
        t6 = self.Inputs.tCLIQ
        t7 = self.Inputs.tEE

        # Check if times are all after the beginning of the simulation, otherwise set the beginning to the earliest time used
        if any(x < t4 for x in [t1,t2,t3, t5, t6, t7]):
            print("You're using a time, that is before the start of the simulation. Corrected Time-Vector.")
            self.Options.time_vector_params[0] = np.min([t1, t2, t3, t5, t6, t7])-0.01
            self.Inputs.t_PC_LUT[0] = np.min([t1, t2, t3, t5, t6, t7])-0.01
            self.Inputs.tQuench = np.zeros((len(self.Inputs.tQuench),))+np.min([t1, t2, t3, t5, t6, t7])

        ### Check end of time vector
        # obtain all times from LEDET object [if times are above > 999, they are interpreted as not set]
        try: t1 = max(self.Inputs.tQuench)
        except: t1 = max(np.array([self.Inputs.tQuench]))
        if t1 >= 999: t1 = 0
        t2 = self.Inputs.t_PC_LUT[-1]
        if t2 >= 999: t2 = 0
        t3 = self.Inputs.tStartQuench
        try: t3 = [0 if x>=999 else x for x in t3]
        except:
            t3 = [t3]
            t3 = [0 if x >= 999 else x for x in t3]
        t3 = max(t3)
        t4 = self.Options.time_vector_params[-1]
        t5 = self.Inputs.tQH
        try: t5 = [0 if x>=999 else x for x in t5]
        except:
            t5 = [t5]
            t5 = t5 = [0 if x>=999 else x for x in t5]
        t5 = max(t5)
        t6 = self.Inputs.tCLIQ
        if t6 >= 999: t6 = 0
        t7 = self.Inputs.tEE
        if t7 >= 999: t7 = 0

        # Check if times are all before the end of the simulation, otherwise extend the time_vector
        if any(x > t4 for x in [t1, t2, t3, t5, t6, t7]):
            print("You're using a time, that is after the end of the simulation. Corrected Time-Vector.")
            self.Options.time_vector_params[-1] = np.max([t1, t2, t3, t5, t6, t7]) + 1
            self.Inputs.t_PC_LUT[-1] = np.max([t1, t2, t3, t5, t6, t7]) + 1
        return 1

    def __checkPersistentCurrents(self):
        if len(self.Inputs.df_inGroup)==1:
            if self.Options.flag_persistentCurrents == 0: return True
            else:
                print('Persistent current parameters flag is set, but parameters not. I set the flag to 0.')
                self.Options.flag_persistentCurrents = 0
                return True
        else:
            if self.Options.flag_persistentCurrents==0:
                print('Persistent current parameters are set but flag is not. Continuing.')
            maxFit = np.max(self.Inputs.selectedFit_inGroup)
            shp = self.Inputs.fitParameters_inGroup.shape
            if maxFit == 1:
                if not shp[0]>0 or not shp[1]==len(self.Inputs.nT):
                    print('You selected constant Jc, but fitParameters are not set. Abort.')
                    return False
                else: return True
            elif maxFit == 2:
                if not shp[0]>1 or not shp[1]==len(self.Inputs.nT):
                    print('You selected Botturas fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            elif maxFit == 3:
                if not shp[0] > 7 or not shp[1] == len(self.Inputs.nT):
                    print('You selected CUDI fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            elif maxFit == 4:
                if not shp[0] > 2 or not shp[1] == len(self.Inputs.nT):
                    print('You selected Summers fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            else:
                print('Unknown fit. Please check! Abort.')
                return False

    def _consistencyCheckLEDET(self):
        '''
            **Consistency check of LEDET Inputs - Main function **

            Function applies different consistency checks on LEDET inputs to see if the values are set properly
            Applied checks:
                - Length checks [checking Inputs that require the same size]
                - checkM_InductanceBlock_m [checking if inductance matrix is squared]
                - checkHeFraction [checking if both Helium options are set]
                - checkMonotony [check Inputs that require monotony in themselves]
                - checkTimes [check if TimeVector fits to other times in the Inputs]

            :return Break: flag, showing if Inputs are consistent or not
            :type break: bool
        '''
        # Define groups that require the same size, number in each list contains the row-number of the attribute
        ## 0 Single - 1 CoilSections - 2 Groups - 3 Half-Turns - 4 doesn't matter - 5 iContactAlongWidth - 6 iContactAlongHeight - 7 vQlength
        ## 8 Quench Heater, 9 QH_QuenchToFrom, 10 CLIQ,
        VarsSameInput = [['T00','l_magnet','I00','R_circuit','R_crowbar','Ud_crowbar','t_PC', 't_EE', 'R_EE_triggered', 'sim3D_uThreshold',
                          'sim3D_f_cooling_down','sim3D_f_cooling_up','sim3D_f_cooling_left','sim3D_f_cooling_right','sim3D_fExToIns',
                          'sim3D_fExUD','sim3D_fExLR','sim3D_min_ds_coarse','sim3D_min_ds_fine','sim3D_min_nodesPerStraightPart',
                          'sim3D_min_nodesPerEndsPart','sim3D_idxFinerMeshHalfTurn','sim3D_Tpulse_sPosition','sim3D_Tpulse_peakT',
                          'sim3D_Tpulse_width','sim3D_durationGIF','sim3D_flag_saveFigures','sim3D_flag_saveGIF',
                          'sim3D_flag_VisualizeGeometry3D','sim3D_flag_SaveGeometry3D'],
                         ['M_m', 'directionCurrentCLIQ', 'tQuench', 'initialQuenchTemp'],
                         ['GroupToCoilSection','polarities_inGroup','nT','nStrands_inGroup','l_mag_inGroup','ds_inGroup','f_SC_strand_inGroup',\
                         'f_ro_eff_inGroup','Lp_f_inGroup','RRR_Cu_inGroup','SCtype_inGroup','STtype_inGroup','insulationType_inGroup', \
                         'internalVoidsType_inGroup', 'externalVoidsType_inGroup', 'wBare_inGroup', 'hBare_inGroup','wIns_inGroup','hIns_inGroup',\
                         'Lp_s_inGroup', 'R_c_inGroup', 'Tc0_NbTi_ht_inGroup', 'Bc2_NbTi_ht_inGroup', 'c1_Ic_NbTi_inGroup','c2_Ic_NbTi_inGroup',\
                         'Tc0_Nb3Sn_inGroup','Bc2_Nb3Sn_inGroup','Jc_Nb3Sn0_inGroup','el_order_half_turns'],
                        ['el_order_half_turns', 'alphasDEG', 'rotation_block', 'mirror_block', 'mirrorY_block','HalfTurnToInductanceBlock'],
                        ['fL_I', 'fL_L', 'overwrite_f_internalVoids_inGroup', 'overwrite_f_externalVoids_inGroup','t_PC_LUT','I_PC_LUT'],
                        ['iContactAlongWidth_From', 'iContactAlongWidth_To'],
                        ['iContactAlongHeight_From', 'iContactAlongHeight_To'],
                        ['iStartQuench', 'tStartQuench', 'lengthHotSpot_iStartQuench', 'vQ_iStartQuench'],
                        ['tQH', 'U0_QH', 'C_QH', 'R_warm_QH', 'w_QH', 'h_QH', 's_ins_QH', 'type_ins_QH', 's_ins_QH_He', 'type_ins_QH_He',\
                         'l_QH', 'f_QH'],
                        ['iQH_toHalfTurn_From','iQH_toHalfTurn_To'],
                        ['tCLIQ','nCLIQ','U0', 'C', 'Rcapa'],
                        ['df_inGroup', 'selectedFit_inGroup', 'fitParameters_inGroup']]


        slicesSameInput = []
        for i in range(len(VarsSameInput)):
            slicesSameInput.append([])

        counter = 1
        for l in self.Inputs.__annotations__:
            for i in range(len(VarsSameInput)):
                try:
                    _ = VarsSameInput[i].index(l)
                    slicesSameInput[i].append(counter-1)
                except:
                    pass
            counter = counter + 1

        # Acquire representative sizes for the defined groups
        lengthInputs = len(self.Inputs.__annotations__)
        sizeInputs = np.zeros((lengthInputs,1))
        sizeInputs[slicesSameInput[0]] = 1 #single Valued
        sizeInputs[slicesSameInput[1]] = self.__getNumberOfCoilSections() #Number of CoilSections

        if sizeInputs[slicesSameInput[1][0]] == -1:
            print("M_m or Number of Coilsections is corrupted. please check.")
            return True
        sizeInputs[slicesSameInput[2]] = len(self.Inputs.nT) #Number of Groups
        sizeInputs[slicesSameInput[3]] = sum(self.Inputs.nT) #Number of Turns
        sizeInputs[slicesSameInput[4]] = 0 #Unchecked
        sizeInputs[slicesSameInput[5]] = len(self.Inputs.iContactAlongWidth_From)
        sizeInputs[slicesSameInput[6]] = len(self.Inputs.iContactAlongHeight_From)
        sizeInputs[slicesSameInput[7]] = len(self.Inputs.iStartQuench)
        sizeInputs[slicesSameInput[8]] = len(self.Inputs.tQH)
        sizeInputs[slicesSameInput[9]] = len(self.Inputs.iQH_toHalfTurn_From)
        try:
            sizeInputs[slicesSameInput[10]] = len(self.Inputs.tCLIQ)
        except:
            sizeInputs[slicesSameInput[10]] = 1
        if len(self.Inputs.df_inGroup)>1:
            sizeInputs[slicesSameInput[11]] = len(self.Inputs.nT)
        else:
            sizeInputs[slicesSameInput[11]] = len(self.Inputs.df_inGroup)

        # Checks for types and lengths/sizes
        Count = 0
        Break = 0
        for k in self.Inputs.__annotations__:
            if sizeInputs[Count] == 0:
                Count = Count + 1
                continue
            cC = self.getAttribute(self.Inputs, k)
            if type(cC) == list:
                if not len(cC)==sizeInputs[Count]:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                    Break = 1
            elif type(cC) == np.ndarray:
                if len(cC.shape)==1:
                    if not len(cC)==sizeInputs[Count]:
                        print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                        Break = 1
                else:
                    if not cC.shape[1]==sizeInputs[Count]:
                        print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",cC.shape[1],"! Please check.")
                        Break = 1
            elif type(cC) == float or type(cC) == int or type(cC) == np.float64:
                if not sizeInputs[Count]== 1:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is", type(cC),". Please check.")
                    Break = 1
            else:
                print("Variable ", k, " has the wrong data-type set! Please check.")
                Break = 1
            Count = Count + 1

        ## Remaining checks in functions
        if not self.__checkHeFraction(len(self.Inputs.nT)):
            Break = 1
        if not self.__checkMonotony(self.Inputs.t_PC_LUT):
            print("t_PC_LUT is not monotonic")
            Break = 1
        if not self.__checkMonotony(self.Inputs.fL_I):
            print("fL_I is not monotonic")
            Break = 1
        if not self.__checkTimes():
            Break = 1
        if not self.__checkPersistentCurrents():
            Break = 1
        if not self.__checkM_InductanceBlock_m(int(sum(self.Inputs.nT)/2)):
            Break = 1
        return Break

    def writeFileLEDET(self, nameFileLEDET, verbose: bool = False, SkipConsistencyCheck: bool = False):
        '''
            **Writes LEDET input file **

            Function to write a LEDET input file composed of "Inputs", "Options", "Plots", and "Variables" sheets

            :param nameFileLEDET: String defining the name of the LEDET input file to be written
            :type nameFileLEDET: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool
            :param SkipConsistencyCheck: flag that determines, whether the parameters shall be checked for consistency or not [False = Apply checks, True = Skip checks]
            :type SkipConsistencyCheck: bool

            :return: None
        '''
        if not SkipConsistencyCheck:
            if self._consistencyCheckLEDET():
                print("Variables are not consistent! Writing aborted - ", nameFileLEDET)
                return
            else:
                print("Preliminary consistency check was successful! - ", nameFileLEDET)
        else:
            print("Skipped consistency checks.")

        workbook = openpyxl.Workbook()
        workbook.properties.creator = 'STEAM-Team'

        if verbose:
            print('')
            print('### Write "Variables" sheet ###')
        self.writeLEDETInputsNew(workbook, "Variables", self.Variables, self.variablesVariables)

        if verbose:
            print('')
            print('### Write "Plots" sheet ###')
        self.writeLEDETInputsNew(workbook, "Plots", self.Plots, self.variablesPlots)

        if verbose:
            print('')
            print('### Write "Options" sheet ###')
        self.writeLEDETInputsNew(workbook, "Options", self.Options, self.variablesOptions)

        if verbose:
            print('### Write "Inputs" sheet ###')
        self.writeLEDETInputsNew(workbook, "Inputs", self.Inputs, self.variablesInputs)

        # Save the workbook
        std = workbook['Sheet']
        workbook.remove(std)
        workbook.save(nameFileLEDET)

        # Display time stamp and end run
        currentDT = datetime.datetime.now()
        print(' ')
        print('Time stamp: ' + str(currentDT))
        print('New file ' + nameFileLEDET + ' generated.')
        return

    def writeLEDETInputsNew(self, book, sheet, variableGroup, variableLabels):
        """
            **Write one sheet of a LEDET input file**

            Function writes one sheet of a LEDET input file

            :param book: workbook object to write
            :type book: openpyxl.Workbook
            :param sheet: name of the sheet to write
            :type sheet: string
            :param variableGroup: dataclass containing the attributes and values to be written into sheet
            :type variableGroup: dataclass
            :param variableLabels: dictionary assigning a description to each variable name
            :type variableLabels: dict
            :return:
        """
        def styled_cells(data):
            for c in data:
                c = Cell(sheet1, column="A", row=1, value=c)
                c.font = Font(size=14,bold=True)
                yield c

        # Loop through sheets to activate the correct sheet to write to
        book.create_sheet(index = 1 , title = sheet)
        for s in range(len(book.sheetnames)):
            if book.sheetnames[s] == sheet: break
        book.active = s
        sheet1 = book.active

        # if type(variableGroup)==LEDETInputs:
        #     if 'Start_I' in self.sectionTitles:
        #         sheet1.append(styled_cells([self.sectionTitles['Start_I']]))
        # if type(variableGroup)==LEDETOptions:
        #     if 'Start_O' in self.sectionTitles:
        #         sheet1.append(styled_cells([self.sectionTitles['Start_O']]))
        # if type(variableGroup)==LEDETPlots:
        #     if 'Start_P' in self.sectionTitles:
        #         sheet1.append(styled_cells([self.sectionTitles['Start_P']]))
        # if type(variableGroup)==LEDETVariables:
        #     if 'Start_V' in self.sectionTitles:
        #         sheet1.append(styled_cells([self.sectionTitles['Start_V']]))

        # Correct section titles, if Helium option is present
        if "overwrite_f_externalVoids_inGroup" in variableGroup.__annotations__:
            ofiVg = self.getAttribute(variableGroup, "overwrite_f_externalVoids_inGroup")
            if len(ofiVg) != 0:
                self.sectionTitles["overwrite_f_externalVoids_inGroup"] = self.sectionTitles["fitParameters_inGroup"]
                del self.sectionTitles["fitParameters_inGroup"]

        # Write to the sheet of the workbook
        # Loop through all attributes in the given variableGroup
        for attribute in variableGroup.__annotations__:
            # Only write the helium options if they are set, otherwise skip
            if (attribute == "overwrite_f_internalVoids_inGroup"):
                ofiVg = self.getAttribute(variableGroup, attribute)
                if len(ofiVg) == 0: continue
            if (attribute == "overwrite_f_externalVoids_inGroup"):
                ofiVg = self.getAttribute(variableGroup, attribute)
                if len(ofiVg) == 0: continue

            # Check if size of list is < 16382 (max size of xlsx - 2 rows for descriptions), if so: convert to np.ndarray
            if isinstance(self.getAttribute(variableGroup, attribute), list):
                if len(self.getAttribute(variableGroup, attribute)) > 16382:
                    self.setAttribute(variableGroup, attribute, np.array(self.getAttribute(variableGroup, attribute)).reshape(-1, 1))
            if isinstance(self.getAttribute(variableGroup, attribute), np.ndarray):
                if self.getAttribute(variableGroup, attribute).shape[0] > 16382:
                    self.setAttribute(variableGroup, attribute, np.array(self.getAttribute(variableGroup, attribute)).reshape(-1, 1))

            # Actual writing process. Check which datatype the attribute is and append it to the sheet
            # If datatype is matrix, a for loop executes all rows/columns before continuing to next attribute
            varDesc = variableLabels.get(str(attribute))
            if isinstance(self.getAttribute(variableGroup, attribute), np.ndarray):
                if self.getAttribute(variableGroup, attribute).ndim > 1:
                    for i in range(self.getAttribute(variableGroup, attribute).shape[0]):
                        values = self.getAttribute(variableGroup, attribute)[i,:].tolist()
                        if i == 0: sheet1.append([varDesc, attribute] + values)
                        else: sheet1.append([None, None] + values)
                else:
                    values = np.array(self.getAttribute(variableGroup, attribute)).tolist()
                    sheet1.append([varDesc, attribute] + values)
            elif isinstance(self.getAttribute(variableGroup, attribute), list):
                values = self.getAttribute(variableGroup, attribute)
                sheet1.append([varDesc, attribute] + values)
            else:
                values = [self.getAttribute(variableGroup, attribute)]
                sheet1.append([varDesc, attribute]+ values)

            if attribute in self.sectionTitles.keys():
                sheet1.append([None])
                sheet1.append(styled_cells([self.sectionTitles[attribute]]))

            # Setting the width of each cells in the workbook [only for good view]
            width = [80.7109375, 40.7109375, 20.7109375]
            sheet1.column_dimensions['A'].width = width[0]
            sheet1.column_dimensions['B'].width = width[1]

            if sheet1.max_column+1 > 18278:
                smc = 18278
            else:
                smc = sheet1.max_column+1
            for i in range(3,smc):
                cl = get_column_letter(i)
                sheet1.column_dimensions[cl].width = width[2]