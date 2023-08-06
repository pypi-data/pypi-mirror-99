from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from w.services.abstract_service import AbstractService
from w.services.technical.filesystem_service import FilesystemService


class ExcelService(AbstractService):
    @classmethod
    def is_excel(cls, filename):
        """ Check if filename is excel (from extension) """
        return FilesystemService.has_suffix(filename, [".xls", ".xlsx"])

    @classmethod
    def check_is_excel(cls, filename) -> Path:
        """
        Check if filename is excel (from extension) or raise RuntimeError
        """
        if cls.is_excel(filename):
            return filename
        raise RuntimeError(f"{Path(filename).name} is not a valid excel file")

    @classmethod
    def open_workbook(cls, filename, **options):
        """
        Open filename is valid excel (check exists too)

        Args:
            filename(str): excel full path filename
            **options:
             - read_only(bool): default False, optimised for reading, content cannot be
               edited
             - keep_vba(bool): default False, preserve vba content
             - data_only(bool): default False, controls whether cells with formulae have
               either the formula (default) or the value stored the last time Excel read
               the sheet
            - keep_links: default True, whether links to external workbooks should be
              preserved.

        Returns:
            Workbook

        Raises:
            RuntimeError: if filename does not exists or is not valid excel
        """
        FilesystemService.check_file_exists(filename)
        cls.check_is_excel(filename)
        try:
            return openpyxl.load_workbook(filename, **options)
        except InvalidFileException:
            raise RuntimeError(f"{Path(filename).name} is not a valid excel file")

    @classmethod
    def load(cls, excel_filename, mapping_columns, sheet_name=None) -> list:
        """
        Load active or specific sheet of excel file

        Args:
            excel_filename(str|Path): excel file
            mapping_columns(dict): csv header columns mapping to wanted attributes
            sheet_name(str) : sheet name (default load active sheet)

        Returns:
            list: [{"<mapping name>": < row col value>, ...}, ...]

        Raises:
            RuntimeError :
                - filename is not csv
                - filename does not exists
                - incorrect or missing header
        """
        wb = cls.open_workbook(excel_filename, read_only=True, data_only=True)

        if sheet_name is not None and sheet_name not in wb.sheetnames:
            raise RuntimeError(f"{sheet_name} is not a valid sheet name")

        sheet = wb[sheet_name] if sheet_name is not None else wb.active

        rows = sheet.iter_rows()
        headers = [c.value.strip() for c in next(rows)]
        required_headers = list(mapping_columns.keys())
        if required_headers != headers:
            raise RuntimeError(
                f"incorrect or missing header, expected '{', '.join(required_headers)}'"
                f" got '{', '.join(headers)}'"
            )
        return [
            {mapping_columns[headers[i]]: cell.value for i, cell in enumerate(row)}
            for row in rows
        ]
