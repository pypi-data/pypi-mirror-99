""" This modules reads and stores the data configuration file """

import fnmatch
import openpyxl


def _read_configuration_sheet(path, sheet):

    # Initialize the return value
    config = {}

    # process the data line by line
    end = False
    row = 1
    config['file'] = path
    config['classes'] = []
    while not end:

        if isinstance(sheet.cell(row=row, column=1).value, str):
            newclass = {}
            newclass['classname'] = sheet.cell(row=row, column=1).value
            newclass['columns'] = []
            config['classes'].append(newclass)

        elif isinstance(sheet.cell(row=row, column=1).value, int):
            number = int(sheet.cell(row=row, column=1).value)
            column = {}
            column['number'] = number
            name = sheet.cell(row=row, column=2).value
            if name.startswith("id:"):
                column['name'] = name[3:]
                column['id'] = True
            else:
                column['name'] = name
                column['id'] = False

            column['type'] = sheet.cell(row=row, column=3).value
            column['entity'] = bool(sheet.cell(row=row, column=4).value)
            column['indexInverted'] = bool(sheet.cell(row=row, column=5).value)

            newclass['columns'].append(column)

        elif sheet.cell(row=row, column=1).value is None:
            end = True

        row += 1

    return config


def read_data_configuration_file(path):
    """ Read the data configuration file """

    config = None

    if fnmatch.fnmatch(path, "*.xls") or fnmatch.fnmatch(path, "*.xlsx"):
        workbook = openpyxl.load_workbook(path, data_only=True)
        if workbook is not None:
            sheet = workbook.active
            if sheet is not None:
                config = _read_configuration_sheet(path, sheet)

    return config
