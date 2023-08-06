# coding: utf-8


"""
    xlTables - Load/generate table data with Excel
    from python dictionary structures

    cunningr - 2020

    Requires openpyxl >= 2.6.2, jsonschema


"""

import os
import openpyxl
from openpyxl import Workbook
from sdtables import xlTables
from tabulate import tabulate


class SdTables:
    def __init__(self, wb=None):
        self.sheetnames = []
        self.table_names = {}
        self.schemas = {}
        self.validation_results = {}

        if wb is not None:
            self.load_xlsx_file(wb)
        else:
            self.wb = Workbook()
            ws = self.wb.active
            self.wb.remove(ws)

    def load_xlsx_file(self, file, data_only=False):
        """
        Method used to load an xlsx file containing one or more tables
        :return:
        """
        self.wb = openpyxl.load_workbook(filename=file, data_only=data_only)
        self.sheetnames = self.wb.sheetnames
        self._get_xl_table_data()

    def _get_xl_table_data(self):
        """
        Internal method used to index tables from openpyxl workbook object
        :return:
        """
        _tables_dict = {}
        for sheet in self.wb.sheetnames:
            for table in self.wb[sheet].tables.values():
                _tables_dict.update({table.name: sheet})

        self.table_names = _tables_dict

    def get_table_as_dict(self, table_name, fill_empty=False, string_only=False):
        """
        Takes a table name and returns the data as list of dictionaries

        Args:
            table_name: Name of the table name
            fill_empty: By default and empty cell will have a value None.
                    fill_empty will replace None with the empty string ""
            string_only: Enforce that all cell values convert to strings

        Returns:
            A dictionary (key=table_name) with a list of dictionaries (rows)

        """
        worksheet_name = self.table_names[table_name]
        ws = self.wb[worksheet_name]

        return xlTables.build_dict_from_table(ws, table_name, fill_empty=fill_empty, string_only=string_only)

    def get_all_tables_as_dict(self, flatten=False, squash=False, fill_empty=False, string_only=False):
        """
        Returns all table data.  When dealing with tables from xlsx, by default each table is nested in a dictionary
        using the worksheet names as the key E.g.

        { "worksheet_name":
            [
                { "table_name": [{"col1": "value", "col2": "value"}]}
            ]
        }

        Args:
            flatten: Removes the worksheet_name hierarchy from the returned dictionary when using xlxs as source
            squash: Replaces the table_name with the worksheet_name when using xlsx as source.
                    Only one table per worksheet allowed and ignores additional tables
            fill_empty: By default and empty cell will have a value None.
                    fill_empty will replace None with the empty string ""
            string_only: Enforce that all cell values convert to strings (E.g. for xlsx formulae)

        Returns:
            A list of dictionaries (rows)

        """

        _dict = {}
        for table_name, worksheet_name in self.table_names.items():
            ws = self.wb[worksheet_name]
            table_dict = xlTables.build_dict_from_table(ws, table_name, fill_empty=fill_empty, string_only=string_only)

            if flatten:
                if squash:
                    print('ERROR: Do not set flatten=True and squash=True together')
                    return
                _dict.update(table_dict)
            elif squash:
                _dict_key = list(table_dict.keys())[0]
                _dict.update({worksheet_name: table_dict[table_name]})
            else:
                if not _dict.get(worksheet_name):
                    _dict.update({worksheet_name: {}})
                _dict[worksheet_name].update(table_dict)

        return _dict

    def add_xlsx_table_from_data(self, table_name, data, worksheet_name='Sheet1', table_style='TableStyleMedium2', row_offset=2, col_offset=1):
        if type(table_name) is not str or type(data) is not list:
            print('ERROR: table name must be of type str and data of type list')
        if worksheet_name not in self.wb.sheetnames:
            _ws = self.wb.create_sheet(worksheet_name)
        else:
            _ws = self.wb[worksheet_name]

        schema = {'properties': xlTables._build_schema_from_row(data[0])}
        xlTables.add_schema_table_to_worksheet(_ws, table_name, schema, data=data, table_style=table_style, row_offset=row_offset, col_offset=col_offset)
        self._get_xl_table_data()

    def update_xlsx_table_data(self, table_name, data, append=True, schema=None):
        print('WARNING: update data is experimental and is known to break data validation')
        self._get_xl_table_data()
        if self.table_names.get(table_name):
            worksheet_name = self.table_names.get(table_name)
        else:
            print('ERROR: table with name {} not found'.format(table_name))
            return

        xlTables.update_table_data(self.wb, worksheet_name, table_name, data, append=append, schema=schema)

    def add_xlsx_table_from_schema(self, table_name, schema, worksheet_name='default', data=None, table_style='TableStyleMedium2', row_offset=2, col_offset=1):
        if type(table_name) is not str or type(schema) is not dict:
            print('ERROR: table name must be of type str and schema of type dict')

        if worksheet_name not in self.wb.sheetnames:
            _ws = self.wb.create_sheet(worksheet_name)
        else:
            _ws = self.wb[worksheet_name]

        self.schemas.update({table_name: schema})

        return xlTables.add_schema_table_to_worksheet(_ws, table_name, schema, data=data, table_style=table_style, row_offset=row_offset, col_offset=col_offset)

    def validate_table_data_with_schema(self, table_name, schema):
        self._get_xl_table_data()
        ws = self.wb[self.table_names[table_name]]
        data = xlTables.build_dict_from_table(ws, table_name, fill_empty=False, string_only=False)

        results = {'results.summary': {}, 'results.details': {}}
        _validate_results = xlTables.validate_data(schema, data[table_name])
        results['results.summary'] = {'table': table_name, 'result': _validate_results['result']}
        results['results.details'] = {'table': table_name, 'result': _validate_results['details']}

        return results

    def validate_table_data(self, stdout=False):
        self._get_xl_table_data()

        results = {'results.summary': [], 'results.details': []}
        for table_name, worksheet_name in self.table_names.items():
            if table_name in self.schemas.keys():
                _validate_results = self.validate_table_data_with_schema(table_name, self.schemas[table_name])
                results['results.summary'].append(_validate_results['results.summary'])
                results['results.details'].append(_validate_results['results.details'])
            else:
                print('WARNING: No schema found for table {}'.format(table_name))

        self.validation_results.update(results)
        if stdout:
            self.print_validation_results()

        return results

    def delete_xlsx_table(self, table_name, row_offset=2, col_offset=1):
        self._get_xl_table_data()
        if self.table_names.get(table_name):
            worksheet_name = self.table_names[table_name]
            xlTables.delete_table(self.wb, worksheet_name, table_name, row_offset=row_offset, col_offset=col_offset)
            self.table_names.pop(table_name, None)

    def add_schema(self, schema_name, schema):
        self.schemas.update({schema_name: schema})

    def delete_schema(self, schema_name):
        self.schemas.pop(schema_name, None)

    def save_xlsx(self, filename):
        xlsx_filename = '{}/{}.xlsx'.format(os.getcwd(), filename)
        self.wb.save(xlsx_filename)

    def print_validation_results(self):
        print('\nValidation Summary:\n')
        print(tabulate(self.validation_results['results.summary'], headers='keys', tablefmt="grid"))

        print('\nValidation Details:\n')
        for table in self.validation_results['results.details']:
            print('Table: {}'.format(table['table']))
            print(tabulate(table['result'], headers='keys', tablefmt="grid"))

# Retrieve a list of schema names under a given worksheet
# list(filter(lambda item: "network_settings" in item.keys(), meme.schemanames))