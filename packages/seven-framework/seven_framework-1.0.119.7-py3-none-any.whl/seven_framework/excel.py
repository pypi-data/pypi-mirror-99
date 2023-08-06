# -*- coding: utf-8 -*-
"""
:Author: ChenCheng
:Date: 2020-05-21 11:20:06
:LastEditTime: 2020-11-04 17:28:31
:LastEditors: ChenXiaolei
:Description: excel help
"""


import openpyxl
import os
import sys
import re
import xlwt
import xlrd
from xlutils.copy import copy

from seven_framework import *


class ExcelHelper(object):
    """
    :Description: 对excel基本操作进行封装的类
    :last_editors: ChenCheng
    """

    @classmethod
    def __get_filename_by_xlsx(self):
        return os.path.abspath('.') + r'/export.xlsx'

    @classmethod
    def __error_file_type(self):
        raise TypeError  # "文件类型错误"

    @classmethod
    def __get_filename_by_xls(self):
        return os.path.abspath('.') + r'/export.xls'

    @classmethod
    def judge_format(self, filename):
        """
        :Description: 判断文件格式
        :param filename: 文件名
        :return: 文件格式 'xls' or 'xlsx'
        :last_editors: ChenCheng
        """
        file_type = re.search('xls[x]*', filename)
        if file_type:
            if file_type.group() == 'xls':
                return 'xls'
            else:
                return 'xlsx'
        else:
            self.__error_file_type()

    @classmethod
    def create_file(self, filename=''):
        """
        :Description: 创建excel文件
        :param filename: 文件名（包含完整路径）,默认在当前文件夹下创建一个"新建Excel.xlsx"
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            filename = self.__get_filename_by_xlsx()

        file_type = self.judge_format(filename)
        if file_type == 'xls':
            self._create_file_by_xls(filename)
        elif file_type == 'xlsx':
            self._create_file_by_xlsx(filename)
        else:
            self.__error_file_type()

    @classmethod
    def input(self, filename, sheet_name='Sheet'):
        """
        :Description: 从excel导入数据
        :param filename：文件名，需要完整路径
        :return data:二维数组形式输出数组
        :last_editors: ChenCheng
        """
        cell_num_begin = [1, 1]
        max_row = self.get_max_row(filename, sheet_name=sheet_name)
        max_column = self.get_max_column(filename, sheet_name=sheet_name)
        cell_num_end = [max_row, max_column]
        data = self.read_cell(
            cell_num_begin, cell_num_end, filename, sheet_name)
        return data

    @classmethod
    def export(self, data, filename=''):
        """
        :Description: 数据导出到excel
        :param data: 字典数组 
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx" 
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            filename = self.__get_filename_by_xlsx()
        else:
            if self.judge_format(filename) == 'xls':
                f = filename + 'x'
                filename = f
        print(f'开始：{filename}')
        self._create_file_by_xlsx(filename)

        # 生成二维数组格式data
        columns = 0
        row_first = dict()

        for dict_data in data:
            for item in dict_data:
                if item not in row_first:
                    columns += 1
                    row_first[item] = columns

        # 数组初始化
        list_list_data = [['' for j in range(columns)]
                          for i in range(len(data) + 1)]
        for item, value in row_first.items():
            list_list_data[0][value - 1] = item

        column = 1
        for dict_data in data:
            column += 1
            for item, value in dict_data.items():
                list_list_data[column - 1][row_first[item] - 1] = value

        cell_num_begin = [1, 1]
        self.updata_cells(list_list_data, cell_num_begin, filename)

    @classmethod
    def updata_cell(self, row, column, value, filename='', sheet_name='Sheet'):
        """
        :Description: 单个单元格更新数据
        :param row, column：单元格坐标（row, column）
        :param value：要更新的数据
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx" 
        :param sheet_name: 表单名
        :return: 
        :last_editors: ChenCheng
        """
        data = [[value, ], ]
        cell_num_begin = [row, column]

        self.updata_cells(data, cell_num_begin, filename, sheet_name)

    @classmethod
    def updata_cells(self, data, cell_num_begin, filename='', sheet_name='Sheet'):
        """
        :Description: 更新单元格数据
        :param data: 需要更新的数据, 二维数组
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx" 
        :param sheet_name: 表单名
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xlsx(filename)
            filename = self.__get_filename_by_xlsx()

        file_type = self.judge_format(filename)
        if file_type == 'xls':
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            self._updata_cell_by_xls(
                data, cell_num_begin, filename, sheet_name)
        elif file_type == 'xlsx':
            self._updata_cell_by_xlsx(
                data, cell_num_begin, filename, sheet_name)
        else:
            self.__error_file_type()

    @classmethod
    def read_cell(self, cell_num_begin, cell_num_end, filename='', sheet_name='Sheet'):
        """
        :Description: 读取指定位置单元格数据
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param cell_num_end: 单元格j结束位置编号,如[2,2]
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xls" 
        :param sheet_name: 表单名
        :return data:返回输入范围内的数据，二维数组，格式为data[[],[],...]
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xls(filename)
            filename = self.__get_filename_by_xlsx()

        file_type = self.judge_format(filename)
        if file_type == 'xls':
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            return self._read_cell_by_xls(cell_num_begin, cell_num_end, filename, sheet_name)
        elif file_type == 'xlsx':
            return self._read_cell_by_xlsx(cell_num_begin, cell_num_end, filename, sheet_name)
        else:
            self.__error_file_type()

    @classmethod
    def delete_cell(self, cell_num_begin, cell_num_end, filename, sheet_name='Sheet'):
        """
        :Description: 读取指定位置单元格数据
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param cell_num_end: 单元格j结束位置编号,如[2,2]
        :param filename: 文件名（包含地址）
        :param sheet_name: 表单名
        :return:
        :last_editors: ChenCheng
        """
        file_type = self.judge_format(filename)
        if file_type == 'xls':
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            self._delete_cell_by_xls(
                cell_num_begin, cell_num_end, filename, sheet_name)
        elif file_type == 'xlsx':
            self._delete_cell_by_xlsx(
                cell_num_begin, cell_num_end, filename, sheet_name)
        else:
            self.__error_file_type()

    @classmethod
    def get_max_row(self, filename, sheet_name='Sheet'):
        """
        :Description: 获取最大行数
        :param filename: 文件名
        :param sheet_name: 表单名
        :return: 返回选定表单有数据的最大行数
        :last_editors: ChenCheng
        """
        if self.judge_format(filename) == 'xls':
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            excel = xlrd.open_workbook(filename)
            sheet = excel.sheet_by_name(sheet_name)
            return sheet.nrows
        else:
            excel = openpyxl.load_workbook(filename)
            sheet = excel[sheet_name]
            return sheet.max_row

    @classmethod
    def get_max_column(self, filename, sheet_name='Sheet'):
        """
        :Description: 获取最大列数
        :param filename: 文件名
        :param sheet_name: 表单名
        :return: 返回选定表单有数据的最大列数
        :last_editors: ChenCheng
        """
        if self.judge_format(filename) == 'xls':
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            excel = xlrd.open_workbook(filename)
            sheet = excel.sheet_by_name(sheet_name)
            return sheet.ncols
        else:
            excel = openpyxl.load_workbook(filename)
            sheet = excel[sheet_name]
            return sheet.max_column

    @classmethod
    def create_sheet(self, sheet_name, filename=''):
        """
        :Description: 创建表单，只支持xlsx
        :param sheet_name: 表单名字
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx"
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            self.create_file_by_xls(filename)
            filename = self.__get_filename_by_xls()

        if self.judge_format(filename) == 'xlsx':
            excel = openpyxl.load_workbook(filename)
            excel.createe_sheet(sheet_name)
            excel.save(filename)
        else:
            print('只支持xlsx创建表单')
            self.__error_file_type()

    @classmethod
    def delete_sheet(self, sheet_name, filename):
        """
        :Description: 删除表单，只支持xlsx
        :param sheet_name: 表单名字
        :param filename: 文件名
        :return: 
        :last_editors: ChenCheng
        """
        if self.judge_format(filename) == 'xlsx':
            excel = openpyxl.load_workbook(filename)
            sheet = excel[sheet_name]
            excel.remove(sheet)
            excel.save(filename)
        else:
            print('只支持xlsx删除表单')
            self.__error_file_type()

    @classmethod
    def _create_file_by_xls(self, filename=''):
        """
        :Description: 创建xls文件
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xls"
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            filename = self.__get_filename_by_xls()

        if self.judge_format(filename) == 'xls':
            excel = xlwt.Workbook()
            sheet = excel.add_sheet('Sheet1')
            excel.save(filename)
        else:
            self.__error_file_type()

    @classmethod
    def _create_file_by_xlsx(self, filename=''):
        """
        :Description: 创建xlsx文件
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx"
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            filename = self.__get_filename_by_xlsx()

        if self.judge_format(filename) == 'xlsx':
            excel = openpyxl.Workbook()
            sheet = excel.active
            excel.save(filename)
        else:
            self.__error_file_type()

    @classmethod
    def _updata_cell_by_xls(self, data, cell_num_begin,  filename='', sheet_id=0):
        """
        :Description: 更新单元格数据
        :param data: 需要更新的数据, 二维数组
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx" 
        :param sheet_id: 表单id
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xls(filename)
            filename = self.__get_filename_by_xls()

        if self.judge_format(filename) == 'xls':
            excel = xlrd.open_workbook(filename)
            excel_copy = copy(excel)
            #sheet = excel_copy.sheet_by_index(sheet_name)
            # updata_cell（）的默认值传入，修改默认值
            if sheet_id == 'Sheet1':
                sheet_id = 0
            sheet = excel_copy.get_sheet(sheet_id)
            # 导入数据
            for i in range(cell_num_begin[0], cell_num_begin[0] + len(data)):
                for j in range(cell_num_begin[1], cell_num_begin[1] + len(data[0])):
                    sheet.write(
                        i-1, j-1, data[i - cell_num_begin[0]][j - cell_num_begin[1]])
            # 保存
            excel_copy.save(filename)
        else:
            self.__error_file_type()

    @classmethod
    def _updata_cell_by_xlsx(self, data, cell_num_begin,  filename='', sheet_name='Sheet'):
        """
        :Description: 更新单元格数据
        :param data: 需要更新的数据, 二维数组
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx" 
        :param sheet_name: 表单名
        :return: 
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xls(filename)
            filename = self.__get_filename_by_xlsx()

        if self.judge_format(filename) == 'xlsx':
            excel = openpyxl.load_workbook(filename)
            sheet = excel[sheet_name]
            # 导入数据
            for i in range(cell_num_begin[0], cell_num_begin[0] + len(data)):
                for j in range(cell_num_begin[1], cell_num_begin[1] + len(data[0])):
                    cell = openpyxl.utils.get_column_letter(j) + str(i)
                    sheet[cell] = data[i - cell_num_begin[0]
                                       ][j - cell_num_begin[1]]
            # 保存
            excel.save(filename)
        else:
            self.__error_file_type()

    @classmethod
    def _read_cell_by_xls(self, cell_num_begin, cell_num_end, filename='', sheet_name='Sheet1'):
        """
        :Description: 读取指定位置单元格数据
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param cell_num_end: 单元格j结束位置编号,如[2,2]
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xls" 
        :param sheet_name: 表单名
        :return data:返回输入范围内的数据，格式为data[[],[],...]
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xls(filename)
            filename = self.__get_filename_by_xls()

        if self.judge_format(filename) == 'xls':
            excel = xlrd.open_workbook(filename)
            sheet = excel.sheet_by_name(sheet_name)
            # 读取数据
            data = []
            for i in range(cell_num_begin[0], cell_num_end[0] + 1):
                data_row = []
                for j in range(cell_num_begin[1], cell_num_end[1] + 1):
                    data_row.append(sheet.cell(i-1, j-1).value)
                data.append(data_row)
            return data
        else:
            self.__error_file_type()

    @classmethod
    def _read_cell_by_xlsx(self, cell_num_begin, cell_num_end, filename='', sheet_name='Sheet'):
        """
        :Description: 读取指定位置单元格数据
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param cell_num_end: 单元格j结束位置编号,如[2,2]
        :param filename: 文件名（包含地址）,默认在当前文件夹下创建一个"新建Excel.xlsx" 
        :param sheet_name: 表单名
        :return data:返回输入范围内的数据，二维数组，格式为data[[],[],...]
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xlsx(filename)
            filename = self.__get_filename_by_xlsx()

        if self.judge_format(filename) == 'xlsx':
            excel = openpyxl.load_workbook(filename)
            sheet = excel[sheet_name]
            # 读取数据
            data = []
            for i in range(cell_num_begin[0], cell_num_end[0] + 1):
                data_row = []
                for j in range(cell_num_begin[1], cell_num_end[1] + 1):
                    cell = openpyxl.utils.get_column_letter(j) + str(i)
                    data_row.append(sheet[cell].value)
                data.append(data_row)
            return data
        else:
            self.__error_file_type()

    @classmethod
    def _delete_cell_by_xls(self, cell_num_begin, cell_num_end, filename, sheet_name='Sheet1'):
        """
        :Description: 读取指定位置单元格数据
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param cell_num_end: 单元格j结束位置编号,如[2,2]
        :param filename: 文件名（包含地址）
        :param sheet_name: 表单名
        :return :
        :last_editors: ChenCheng
        """
        if not filename:
            self._create_file_by_xlsx(filename)

        if self.judge_format(filename) == 'xls':
            data = []
            for i in range(cell_num_begin[0], cell_num_end[0] + 1):
                data_row = []
                for j in range(cell_num_begin[1], cell_num_end[1] + 1):
                    data_row.append('')
                data.append(data_row)
            self._updata_cell_by_xls(
                data, cell_num_begin, filename, sheet_name)
        else:
            self.__error_file_type()

    @classmethod
    def _delete_cell_by_xlsx(self, cell_num_begin, cell_num_end, filename, sheet_name='Sheet'):
        """
        :Description: 读取指定位置单元格数据
        :param cell_num_begin: 单元格起始位置编号,如[1,1]，注意：第一格坐标为[1,1]
        :param cell_num_end: 单元格j结束位置编号,如[2,2]
        :param filename: 文件名（包含地址）
        :param sheet_name: 表单名
        :return:
        :last_editors: ChenCheng
        """
        if self.judge_format(filename) == 'xlsx':
            excel = openpyxl.load_workbook(filename)
            sheet = excel[sheet_name]
            # 删除数据
            for i in range(cell_num_begin[0], cell_num_end[0] + 1):
                for j in range(cell_num_begin[1], cell_num_end[1] + 1):
                    cell = openpyxl.utils.get_column_letter(j) + str(i)
                    sheet[cell] = ''
            excel.save(filename)
        else:
            self.__error_file_type()
