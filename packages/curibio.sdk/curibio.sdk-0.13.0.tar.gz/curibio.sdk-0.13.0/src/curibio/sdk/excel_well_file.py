# -*- coding: utf-8 -*-
"""Classes and functions for finding and managing excel files."""
import datetime
from typing import Any
from typing import Optional
from uuid import UUID

from labware_domain_models import LabwareDefinition
from mantarray_file_manager import CURI_BIO_ACCOUNT_UUID
from mantarray_file_manager import CURI_BIO_USER_ACCOUNT_ID
from mantarray_file_manager import MANTARRAY_SERIAL_NUMBER_UUID
from mantarray_file_manager import PLATE_BARCODE_UUID
from mantarray_file_manager import TISSUE_SAMPLING_PERIOD_UUID
from mantarray_file_manager import UTC_BEGINNING_RECORDING_UUID
from mantarray_file_manager import WELL_NAME_UUID
from mantarray_file_manager import WellFile
from nptyping import NDArray
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from stdlib_utils import get_current_file_abs_directory
from xlsxwriter.utility import xl_cell_to_rowcol

from .constants import EXCEL_OPTICAL_METADATA_CELLS
from .constants import INTERPOLATION_VALUE_UUID
from .constants import METADATA_UUID_DESCRIPTIONS
from .constants import TWITCHES_POINT_UP_UUID
from .exceptions import MetadataNotFoundError

PATH_OF_CURRENT_FILE = get_current_file_abs_directory()


def _get_col_as_array(
    sheet: Worksheet,
    zero_based_row: int,
    zero_based_col: int,
) -> NDArray[(2, Any), float]:
    col_array = []
    result = _get_cell_value(sheet, zero_based_row, zero_based_col)
    zero_based_row += 1
    while result:
        col_array.append(float(result))
        result = _get_cell_value(sheet, zero_based_row, zero_based_col)
        zero_based_row += 1
    return np.array(col_array)


def _get_single_sheet(file_name: str) -> Any:
    work_book = load_workbook(file_name)
    return work_book[work_book.sheetnames[0]]


def _get_cell_value(
    sheet: Worksheet,
    zero_based_row: int,
    zero_based_col: int,
) -> Optional[str]:
    result = sheet.cell(row=zero_based_row + 1, column=zero_based_col + 1).value
    if result is None:
        return result
    return str(result)


class ExcelWellFile(WellFile):
    # pylint:disable=too-many-ancestors # Eli (2/8/21): We had to subclass the basic H5 file, which already has a lot of ancestors...I don't see a clear way around this
    """Wrapper around an Excel file for a single well of optical data.

    Args:
        file_name: The path of the excel file to open.

    Attributes:
        _excel_sheet: The opened excel sheet.
    """

    def __init__(self, file_name: str) -> None:
        # pylint: disable=super-init-not-called
        self._excel_sheet = _get_single_sheet(file_name)
        self._file_name = file_name
        self._file_version = "0.1.1"
        self._raw_tissue_reading: Optional[NDArray[(2, Any), int]] = None
        self._raw_ref_reading: Optional[NDArray[(2, Any), int]] = np.zeros(
            self.get_raw_tissue_reading().shape
        )

    def get_excel_metadata_value(self, metadata_uuid: UUID) -> Optional[str]:
        """Return a user-entered metadata value."""
        metadata_description = METADATA_UUID_DESCRIPTIONS[metadata_uuid]
        cell_name = EXCEL_OPTICAL_METADATA_CELLS.get(metadata_uuid, None)
        if cell_name is None:
            raise NotImplementedError(
                f"Metadata value for {metadata_description} is not contained in excel files of well data"
            )
        row, col = xl_cell_to_rowcol(cell_name)
        result = _get_cell_value(self._excel_sheet, row, col)
        if result is None and metadata_uuid != INTERPOLATION_VALUE_UUID:
            raise MetadataNotFoundError(
                f"Metadata entry not found for {metadata_description}"
            )
        return result

    def get_h5_file(self) -> None:
        raise NotImplementedError("ExcelWellFiles do not store an H5 file")

    def get_h5_attribute(self, attr_name: str) -> Any:
        raise NotImplementedError(
            "ExcelWellFiles do not store an H5 file and therefore cannot get H5 attributes"
        )

    def get_well_name(self) -> str:
        return str(self.get_excel_metadata_value(WELL_NAME_UUID))

    def get_well_index(self) -> int:
        twenty_four_well = LabwareDefinition(row_count=4, column_count=6)
        return int(twenty_four_well.get_well_index_from_well_name(self.get_well_name()))

    def get_plate_barcode(self) -> str:
        return str(self.get_excel_metadata_value(PLATE_BARCODE_UUID))

    def get_user_account(self) -> UUID:
        if not isinstance(CURI_BIO_USER_ACCOUNT_ID, UUID):
            # Tanner (10/13/20): Making mypy happy
            raise NotImplementedError(
                "CURI_BIO_USER_ACCOUNT_ID should always be a UUID"
            )
        return CURI_BIO_USER_ACCOUNT_ID

    def get_timestamp_of_beginning_of_data_acquisition(self) -> datetime.datetime:
        return self.get_begin_recording()

    def get_customer_account(self) -> UUID:
        if not isinstance(CURI_BIO_ACCOUNT_UUID, UUID):
            # Tanner (10/13/20): Making mypy happy
            raise NotImplementedError("CURI_BIO_ACCOUNT_UUID should always be a UUID")
        return CURI_BIO_ACCOUNT_UUID

    def get_mantarray_serial_number(self) -> str:
        return str(self.get_excel_metadata_value(MANTARRAY_SERIAL_NUMBER_UUID))

    def get_begin_recording(self) -> datetime.datetime:
        timestamp_str = str(self.get_excel_metadata_value(UTC_BEGINNING_RECORDING_UUID))
        timestamp = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
        return timestamp

    def get_timestamp_of_first_tissue_data_point(self) -> datetime.datetime:
        return self.get_begin_recording()

    def get_timestamp_of_first_ref_data_point(self) -> datetime.datetime:
        return self.get_begin_recording()

    def get_tissue_sampling_period_microseconds(self) -> int:
        value = self.get_excel_metadata_value(TISSUE_SAMPLING_PERIOD_UUID)
        if value is None:
            raise NotImplementedError(
                "Tissue Sampling Period should never be None here. A MetadataNotFoundError should have been raised by get_excel_metadata_value"
            )
        sampling_period_seconds = 1 / float(value)
        return int(round(sampling_period_seconds, 6) * 1e6)

    def get_reference_sampling_period_microseconds(self) -> int:
        return 0

    def get_recording_start_index(self) -> int:
        return 0

    def get_twitches_point_up(self) -> bool:
        return "y" in str(self.get_excel_metadata_value(TWITCHES_POINT_UP_UUID)).lower()

    def get_raw_tissue_reading(self) -> NDArray[(2, Any), float]:
        if self._raw_tissue_reading is None:
            self._raw_tissue_reading = np.array(
                (
                    _get_col_as_array(self._excel_sheet, 1, 0),
                    _get_col_as_array(self._excel_sheet, 1, 1),
                )
            )
        return self._raw_tissue_reading

    def get_raw_reference_reading(self) -> NDArray[(2, Any), float]:
        return self._raw_ref_reading

    def get_interpolation_value(self) -> float:
        result = self.get_excel_metadata_value(INTERPOLATION_VALUE_UUID)
        if result is None:
            return self.get_tissue_sampling_period_microseconds()
        return float(result) * 1e6

    def __del__(self) -> None:
        """Destroy the object without using the parent class.

        The ``BasicWellFile`` parent class closes the H5 file during
        destruction, but this subclass does not have that attribute.
        """
